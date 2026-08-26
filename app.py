import os, io, csv, json, hashlib, hmac, datetime, urllib.parse, html, base64, re, secrets, uuid, shutil, subprocess, threading, time, decimal
from pathlib import Path
from email.message import EmailMessage
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, abort
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.exceptions import RequestEntityTooLarge
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from sqlalchemy import func, or_, inspect
from dateutil.relativedelta import relativedelta
import requests
import qrcode
from PIL import Image as PILImage, ImageOps, ImageEnhance, ImageFilter, ImageStat
from zoneinfo import ZoneInfo

try:
    import pillow_heif
    pillow_heif.register_heif_opener()
except Exception:
    pillow_heif=None

from agreement_core import PRESETS, DEFAULTS, FIELDS, FORMAT_PROFILES, build_agreement_text, build_agreement_text_hindi
from electricity_core import normalize_bill_payload, bill_dedupe_key, reminder_status, transition_payment_status, build_electricity_csv, build_electricity_xlsx, fetch_bill_from_provider
from electricity_providers import load_seed_providers, seed_electricity_providers, safe_official_url
from vault_core import encrypt_secret, decrypt_secret, mask_secret, validate_secret_type, ALLOWED_SECRET_TYPES, encrypt_blob, decrypt_blob
from integrations_core import category_module, user_can_access_category, safe_connection_summary, validate_integration_secret_name, normalize_nonsecret_config
from integrations_catalog import load_integration_catalog, seed_integration_providers, legacy_connection_status, provider_workflow_url
from letterhead_ai import classify_request, missing_required_fields, build_ai_draft_request, parse_structured_draft, rewrite_action, deterministic_draft
from letterhead_integrations import get_ai_client, send_email as letterhead_send_email, send_whatsapp as letterhead_send_whatsapp
from letterhead_pdf import render_letterhead_pdf, merge_annexures, sha256_bytes
from letterhead_delivery import can_retry_delivery
from letterhead_core import validate_structured_content, audit_safe_metadata, format_reference_number, build_reference_prefix, can_transition_document
from letterhead_sources import SourceCandidate, resolve_sources, minimize_for_ai, can_access_protected_source
from letterhead_templates import template_is_usable, signature_is_usable, next_template_version_no, starter_template_definitions
from party_master_core import (MASTER_FIELD_SET, SENSITIVE_FIELDS, DOCUMENT_CATEGORIES, LANDLORD_AGREEMENT_MAP, TENANT_AGREEMENT_MAP, normalize_master_payload, safe_master_summary, identifier_lookup_hash, identifier_lookup_hashes, mask_identifier, master_display_payload, validate_master_document, legacy_profile_to_master, apply_master_mapping, parse_annexure_ids)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OS_NAME = 'Tesla OS 27'
OS_VERSION = '27.0.1'
OS_BUILD = '27A101'
APP_VERSION = OS_VERSION
app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'change-this-secret-before-production')
raw_db = os.getenv('DATABASE_URL', '')
if raw_db.startswith('postgres://'):
    raw_db = raw_db.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = raw_db or ('sqlite:///' + os.path.join(BASE_DIR, 'instance', 'livenza_web.db'))
# Render + Supabase Session Pooler resilience. pool_pre_ping prevents the first
# navigation after an idle/stale pooled connection from throwing a transient 500.
if raw_db.startswith('postgresql'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 180,
        'pool_timeout': 20,
        'pool_size': 3,
        'max_overflow': 3,
        'pool_use_lifo': True,
    }
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Browser-to-storage resumable Video Wall uploads bypass this application-body
# limit. Keep ordinary form uploads bounded so one request cannot exhaust the
# web worker while leaving enough headroom for a 50 MB storage object.
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_UPLOAD_MB','64')) * 1024 * 1024
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
if os.getenv('FORCE_HTTPS', '1') == '1':
    app.config['SESSION_COOKIE_SECURE'] = True

db = SQLAlchemy(app)

@app.errorhandler(RequestEntityTooLarge)
def request_too_large(error):
    if request.path.startswith('/agreements/aadhaar-extract'):
        return jsonify(ok=False,error='The Aadhaar file is larger than 10 MB. Use a clear JPG/PNG or a smaller PDF.'),413
    if request.path.startswith('/video-wall/'):
        return jsonify(ok=False,error='This direct form upload is too large. Keep this page open and use the resumable Video Wall uploader.'),413
    return jsonify(ok=False,error='The uploaded file is larger than this request allows.'),413

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(300), nullable=False)
    role = db.Column(db.String(30), default='manager')
    full_name = db.Column(db.String(180), default='')
    photo_data_uri = db.Column(db.Text, default='')
    avatar_data_uri = db.Column(db.Text, default='')
    avatar_generation_mode = db.Column(db.String(40), default='')
    avatar_updated_at = db.Column(db.DateTime, nullable=True)
    aadhaar_last4 = db.Column(db.String(4), default='')
    aadhaar_name = db.Column(db.String(180), default='')
    aadhaar_verification_status = db.Column(db.String(40), default='Not verified')
    aadhaar_verification_method = db.Column(db.String(80), default='')
    aadhaar_verification_ref = db.Column(db.String(180), default='')
    aadhaar_verified_at = db.Column(db.DateTime, nullable=True)
    permissions_json = db.Column(db.Text, default='[]')
    capabilities_json = db.Column(db.Text, default='[]')
    pattern_hash = db.Column(db.Text, default='')
    webauthn_enabled = db.Column(db.Boolean, default=False)
    webauthn_enrolled_at = db.Column(db.DateTime, nullable=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class City(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    code = db.Column(db.String(30), default='')
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class Agreement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(250), nullable=False)
    preset = db.Column(db.String(120), default='Strong Residential - 11 Months')
    data_json = db.Column(db.Text, nullable=False, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

    @property
    def data(self):
        try: return json.loads(self.data_json or '{}')
        except Exception: return {}

class AgreementPartyProfile(db.Model):
    """Encrypted reusable party details for the Agreement Studio."""
    id = db.Column(db.Integer, primary_key=True)
    profile_type = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(180), nullable=False)
    data_ciphertext = db.Column(db.Text, nullable=False, default='')
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('profile_type','name', name='uq_agreement_party_profile_name'),)

    @property
    def data(self):
        try:
            raw=_integration_cipher().decrypt((self.data_ciphertext or '').encode('ascii')).decode('utf-8')
            value=json.loads(raw)
            return value if isinstance(value,dict) else {}
        except Exception:
            return {}

class LandlordMaster(db.Model):
    __tablename__='landlord_master'
    id=db.Column(db.Integer,primary_key=True)
    master_code=db.Column(db.String(32),unique=True,nullable=False,index=True)
    profile_name=db.Column(db.String(180),nullable=False,index=True)
    party_type=db.Column(db.String(40),default='individual',index=True)
    legal_name=db.Column(db.String(220),default='',index=True)
    primary_mobile=db.Column(db.String(40),default='',index=True)
    email=db.Column(db.String(220),default='',index=True)
    city=db.Column(db.String(120),default='',index=True)
    state=db.Column(db.String(120),default='',index=True)
    country=db.Column(db.String(120),default='India')
    verification_status=db.Column(db.String(40),default='unverified',index=True)
    tags=db.Column(db.String(500),default='',index=True)
    search_text=db.Column(db.Text,default='')
    identifier_lookup_json=db.Column(db.Text,default='[]')
    active=db.Column(db.Boolean,default=True,index=True)
    encrypted_payload=db.Column(db.Text,nullable=False,default='')
    encrypted_nonce=db.Column(db.Text,nullable=False,default='')
    legacy_profile_id=db.Column(db.Integer,unique=True,nullable=True)
    created_by_user_id=db.Column(db.Integer,db.ForeignKey('user.id'),nullable=True)
    updated_by_user_id=db.Column(db.Integer,db.ForeignKey('user.id'),nullable=True)
    created_at=db.Column(db.DateTime,default=datetime.datetime.utcnow)
    updated_at=db.Column(db.DateTime,default=datetime.datetime.utcnow,onupdate=datetime.datetime.utcnow)

class TenantMaster(db.Model):
    __tablename__='tenant_master'
    id=db.Column(db.Integer,primary_key=True)
    master_code=db.Column(db.String(32),unique=True,nullable=False,index=True)
    profile_name=db.Column(db.String(180),nullable=False,index=True)
    party_type=db.Column(db.String(40),default='individual',index=True)
    legal_name=db.Column(db.String(220),default='',index=True)
    primary_mobile=db.Column(db.String(40),default='',index=True)
    email=db.Column(db.String(220),default='',index=True)
    city=db.Column(db.String(120),default='',index=True)
    state=db.Column(db.String(120),default='',index=True)
    country=db.Column(db.String(120),default='India')
    verification_status=db.Column(db.String(40),default='unverified',index=True)
    tags=db.Column(db.String(500),default='',index=True)
    search_text=db.Column(db.Text,default='')
    identifier_lookup_json=db.Column(db.Text,default='[]')
    active=db.Column(db.Boolean,default=True,index=True)
    encrypted_payload=db.Column(db.Text,nullable=False,default='')
    encrypted_nonce=db.Column(db.Text,nullable=False,default='')
    legacy_profile_id=db.Column(db.Integer,unique=True,nullable=True)
    created_by_user_id=db.Column(db.Integer,db.ForeignKey('user.id'),nullable=True)
    updated_by_user_id=db.Column(db.Integer,db.ForeignKey('user.id'),nullable=True)
    created_at=db.Column(db.DateTime,default=datetime.datetime.utcnow)
    updated_at=db.Column(db.DateTime,default=datetime.datetime.utcnow,onupdate=datetime.datetime.utcnow)

class MasterDocument(db.Model):
    __tablename__='master_document'
    id=db.Column(db.Integer,primary_key=True)
    owner_type=db.Column(db.String(20),nullable=False,index=True)
    landlord_master_id=db.Column(db.Integer,db.ForeignKey('landlord_master.id'),nullable=True,index=True)
    tenant_master_id=db.Column(db.Integer,db.ForeignKey('tenant_master.id'),nullable=True,index=True)
    category=db.Column(db.String(80),nullable=False,index=True)
    display_label=db.Column(db.String(180),nullable=False)
    storage_id=db.Column(db.String(64),unique=True,nullable=False,index=True)
    extension=db.Column(db.String(16),nullable=False)
    mime_type=db.Column(db.String(120),nullable=False)
    ciphertext=db.Column(db.Text,nullable=False)
    nonce=db.Column(db.Text,nullable=False)
    issue_date=db.Column(db.Date,nullable=True)
    expiry_date=db.Column(db.Date,nullable=True,index=True)
    verification_status=db.Column(db.String(40),default='unverified')
    replaced_document_id=db.Column(db.Integer,nullable=True)
    active=db.Column(db.Boolean,default=True,index=True)
    uploaded_by_user_id=db.Column(db.Integer,db.ForeignKey('user.id'),nullable=True)
    uploaded_at=db.Column(db.DateTime,default=datetime.datetime.utcnow)

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    city = db.Column(db.String(120), default='')
    property_name = db.Column(db.String(180), default='')
    premises = db.Column(db.Text, default='')
    room_no = db.Column(db.String(60), nullable=False)
    room_type = db.Column(db.String(120), default='')
    capacity = db.Column(db.String(30), default='')
    standard_tariff = db.Column(db.String(60), default='')
    status_override = db.Column(db.String(40), default='')
    notes = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('property_name','room_no', name='uq_room_property'),)

class Tenant(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tenant_name = db.Column(db.String(180), default='')
    tenant_mobile = db.Column(db.String(40), default='')
    tenant_whatsapp = db.Column(db.String(40), default='')
    tenant_email = db.Column(db.String(180), default='')
    tenant_id_type = db.Column(db.String(40), default='')
    tenant_id_no = db.Column(db.String(100), default='')
    city = db.Column(db.String(120), default='')
    property_name = db.Column(db.String(180), default='')
    premises = db.Column(db.Text, default='')
    room_unit_no = db.Column(db.String(60), default='')
    room_type = db.Column(db.String(120), default='')
    tariff = db.Column(db.String(60), default='')
    security_deposit = db.Column(db.String(60), default='')
    joining_date = db.Column(db.String(40), default='')
    leaving_date = db.Column(db.String(40), default='')
    status = db.Column(db.String(40), default='Occupied')
    agreement_id = db.Column(db.Integer, db.ForeignKey('agreement.id'), nullable=True)
    agreement_reference = db.Column(db.String(120), default='')
    notes = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class Review(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    business_name = db.Column(db.String(180), default='')
    business_type = db.Column(db.String(120), default='')
    experience = db.Column(db.Text, default='')
    output = db.Column(db.Text, default='')
    language = db.Column(db.String(40), default='English')
    google_review_url = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class FoodOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    platform = db.Column(db.String(50), default='Direct')
    order_id = db.Column(db.String(120), default='')
    outlet = db.Column(db.String(180), default='')
    customer = db.Column(db.String(180), default='')
    order_time = db.Column(db.String(60), default='')
    status = db.Column(db.String(50), default='New')
    payment_mode = db.Column(db.String(50), default='')
    gross = db.Column(db.Float, default=0)
    commission = db.Column(db.Float, default=0)
    fees = db.Column(db.Float, default=0)
    taxes = db.Column(db.Float, default=0)
    net = db.Column(db.Float, default=0)
    settlement_status = db.Column(db.String(50), default='Pending')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)


class BankDocument(db.Model):
    """Encrypted bank statements and reusable reconciliation templates."""
    __tablename__ = 'bank_document'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    document_type = db.Column(db.String(24), nullable=False, default='statement')
    bank_name = db.Column(db.String(160), default='')
    account_label = db.Column(db.String(160), default='')
    title = db.Column(db.String(180), default='')
    file_name = db.Column(db.String(255), nullable=False)
    mime_type = db.Column(db.String(120), default='application/octet-stream')
    encrypted_blob = db.Column(db.LargeBinary, nullable=False)
    parsed_ciphertext = db.Column(db.Text, default='')
    row_count = db.Column(db.Integer, default=0)
    parse_status = db.Column(db.String(80), default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class BankReconciliationRun(db.Model):
    __tablename__ = 'bank_reconciliation_run'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    statement_id = db.Column(db.Integer, db.ForeignKey('bank_document.id'), nullable=False)
    template_id = db.Column(db.Integer, db.ForeignKey('bank_document.id'), nullable=False)
    summary_json = db.Column(db.Text, default='{}')
    result_ciphertext = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class ElectricityProvider(db.Model):
    __tablename__ = 'electricity_provider'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(180), nullable=False)
    state = db.Column(db.String(120), nullable=False, default='')
    city = db.Column(db.String(120), nullable=False, default='')
    official_website_url = db.Column(db.Text, default='')
    official_payment_url = db.Column(db.Text, default='')
    official_login_url = db.Column(db.Text, default='')
    identifier_types_json = db.Column(db.Text, default='[]')
    bbps_biller_id = db.Column(db.String(120), default='')
    supports_bbps_fetch = db.Column(db.Boolean, default=False)
    supports_bbps_payment = db.Column(db.Boolean, default=False)
    embedding_mode = db.Column(db.String(24), default='external')
    workflow_mode = db.Column(db.String(24), default='portal')
    active = db.Column(db.Boolean, default=True)
    notes = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    __table_args__=(db.UniqueConstraint('name','state','city',name='uq_electricity_provider_scope'),)

    @property
    def identifier_types(self):
        try:
            value=json.loads(self.identifier_types_json or '[]')
            return value if isinstance(value,list) else []
        except Exception: return []

class VaultSecret(db.Model):
    __tablename__ = 'vault_secret'
    id = db.Column(db.Integer, primary_key=True)
    secret_type = db.Column(db.String(60), nullable=False)
    label = db.Column(db.String(180), nullable=False)
    username_masked = db.Column(db.String(180), default='')
    ciphertext = db.Column(db.Text, nullable=False)
    nonce = db.Column(db.Text, nullable=False)
    key_version = db.Column(db.String(24), default='v1')
    linked_provider_id = db.Column(db.Integer, db.ForeignKey('electricity_provider.id'), nullable=True)
    linked_connection_id = db.Column(db.Integer, nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    updated_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class IntegrationProvider(db.Model):
    __tablename__ = 'integration_provider'
    id = db.Column(db.Integer, primary_key=True)
    provider_key = db.Column(db.String(80), nullable=False, unique=True, index=True)
    display_name = db.Column(db.String(180), nullable=False)
    category = db.Column(db.String(60), nullable=False, index=True)
    workflow_module = db.Column(db.String(60), nullable=False, default='integrations')
    portal_url = db.Column(db.Text, default='')
    developer_url = db.Column(db.Text, default='')
    embed_mode = db.Column(db.String(24), default='external')
    capabilities_json = db.Column(db.Text, default='[]')
    active = db.Column(db.Boolean, default=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

    @property
    def capabilities(self):
        try:
            value=json.loads(self.capabilities_json or '[]')
            return value if isinstance(value,list) else []
        except Exception:
            return []

class IntegrationConnection(db.Model):
    __tablename__ = 'integration_connection'
    id = db.Column(db.Integer, primary_key=True)
    provider_id = db.Column(db.Integer, db.ForeignKey('integration_provider.id'), nullable=False, index=True)
    display_name = db.Column(db.String(180), nullable=False)
    property_scope = db.Column(db.String(180), default='', index=True)
    source_mode = db.Column(db.String(24), default='native')
    status = db.Column(db.String(32), default='unconfigured', index=True)
    nonsecret_config_json = db.Column(db.Text, default='{}')
    last_test_status = db.Column(db.String(32), default='')
    last_test_message = db.Column(db.Text, default='')
    last_tested_at = db.Column(db.DateTime, nullable=True)
    last_success_status = db.Column(db.String(32), default='')
    last_success_message = db.Column(db.Text, default='')
    last_success_at = db.Column(db.DateTime, nullable=True)
    active = db.Column(db.Boolean, default=True, index=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    updated_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    provider = db.relationship('IntegrationProvider', foreign_keys=[provider_id])

class IntegrationSecretRef(db.Model):
    __tablename__ = 'integration_secret_ref'
    id = db.Column(db.Integer, primary_key=True)
    connection_id = db.Column(db.Integer, db.ForeignKey('integration_connection.id'), nullable=False, index=True)
    secret_name = db.Column(db.String(80), nullable=False)
    vault_secret_id = db.Column(db.Integer, db.ForeignKey('vault_secret.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('connection_id','secret_name',name='uq_integration_connection_secret'),)

class LetterheadTemplate(db.Model):
    __tablename__ = 'letterhead_template'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False)
    slug = db.Column(db.String(180), nullable=False, unique=True, index=True)
    entity_scope = db.Column(db.String(160), default='')
    property_scope = db.Column(db.String(160), default='')
    document_family_scope = db.Column(db.String(160), default='')
    status = db.Column(db.String(24), nullable=False, default='draft', index=True)
    current_published_version_id = db.Column(db.Integer, nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class LetterheadTemplateVersion(db.Model):
    __tablename__ = 'letterhead_template_version'
    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey('letterhead_template.id'), nullable=False, index=True)
    version_no = db.Column(db.Integer, nullable=False)
    lifecycle_state = db.Column(db.String(24), nullable=False, default='draft', index=True)
    layout_json = db.Column(db.Text, nullable=False, default='{}')
    scope_json = db.Column(db.Text, nullable=False, default='{}')
    content_hash = db.Column(db.String(64), nullable=False, default='')
    submitted_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    submitted_at = db.Column(db.DateTime, nullable=True)
    published_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    published_at = db.Column(db.DateTime, nullable=True)
    rejection_comment = db.Column(db.Text, nullable=False, default='')
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('template_id','version_no',name='uq_letterhead_template_version'),)

class LetterheadAsset(db.Model):
    __tablename__ = 'letterhead_asset'
    id = db.Column(db.Integer, primary_key=True)
    asset_kind = db.Column(db.String(40), nullable=False, index=True)
    owner_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    mime_type = db.Column(db.String(80), nullable=False)
    encrypted_asset = db.Column(db.LargeBinary, nullable=False)
    sha256 = db.Column(db.String(64), nullable=False, index=True)
    display_name = db.Column(db.String(240), nullable=False, default='')
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)

class SignatureAsset(db.Model):
    __tablename__ = 'signature_asset'
    id = db.Column(db.Integer, primary_key=True)
    asset_kind = db.Column(db.String(24), nullable=False, default='signature')
    signatory_name = db.Column(db.String(160), nullable=False)
    designation = db.Column(db.String(160), nullable=False, default='')
    scope_json = db.Column(db.Text, nullable=False, default='{}')
    encrypted_asset = db.Column(db.LargeBinary, nullable=False)
    mime_type = db.Column(db.String(80), nullable=False)
    effective_date = db.Column(db.Date, nullable=True)
    expires_at = db.Column(db.Date, nullable=True)
    revoked_at = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True, index=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)

class LetterheadDocument(db.Model):
    __tablename__ = 'letterhead_document'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(240), nullable=False)
    document_family = db.Column(db.String(80), nullable=False, default='custom', index=True)
    lifecycle_state = db.Column(db.String(24), nullable=False, default='draft', index=True)
    creator_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    property_ref = db.Column(db.String(160), nullable=False, default='', index=True)
    entity_ref = db.Column(db.String(160), nullable=False, default='', index=True)
    source_refs_json = db.Column(db.Text, nullable=False, default='[]')
    current_revision_id = db.Column(db.Integer, nullable=True)
    finalized_revision_id = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class LetterheadDocumentRevision(db.Model):
    __tablename__ = 'letterhead_document_revision'
    id = db.Column(db.Integer, primary_key=True)
    document_id = db.Column(db.Integer, db.ForeignKey('letterhead_document.id'), nullable=False, index=True)
    revision_no = db.Column(db.Integer, nullable=False)
    structured_content_json = db.Column(db.Text, nullable=False, default='{}')
    template_version_id = db.Column(db.Integer, db.ForeignKey('letterhead_template_version.id'), nullable=True)
    signature_asset_id = db.Column(db.Integer, db.ForeignKey('signature_asset.id'), nullable=True)
    reference_number = db.Column(db.String(160), nullable=False, default='', index=True)
    status = db.Column(db.String(24), nullable=False, default='draft', index=True)
    encrypted_pdf = db.Column(db.LargeBinary, nullable=True)
    pdf_sha256 = db.Column(db.String(64), nullable=False, default='')
    approved_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    finalized_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('document_id','revision_no',name='uq_letterhead_document_revision'),)

class DocumentAttachmentLink(db.Model):
    __tablename__ = 'document_attachment_link'
    id = db.Column(db.Integer, primary_key=True)
    revision_id = db.Column(db.Integer, db.ForeignKey('letterhead_document_revision.id'), nullable=False, index=True)
    source_kind = db.Column(db.String(80), nullable=False)
    source_id = db.Column(db.String(120), nullable=False)
    suggested_by_ai = db.Column(db.Boolean, nullable=False, default=False)
    approved_by_user = db.Column(db.Boolean, nullable=False, default=False)
    approved_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)

class DocumentDelivery(db.Model):
    __tablename__ = 'document_delivery'
    id = db.Column(db.Integer, primary_key=True)
    revision_id = db.Column(db.Integer, db.ForeignKey('letterhead_document_revision.id'), nullable=False, index=True)
    channel = db.Column(db.String(24), nullable=False, index=True)
    recipient = db.Column(db.String(320), nullable=False)
    state = db.Column(db.String(24), nullable=False, default='pending', index=True)
    provider_name = db.Column(db.String(80), nullable=False, default='')
    provider_reference = db.Column(db.String(240), nullable=False, default='')
    attempt_no = db.Column(db.Integer, nullable=False, default=1)
    error_code = db.Column(db.String(120), nullable=False, default='')
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    completed_at = db.Column(db.DateTime, nullable=True)

class DocumentSequence(db.Model):
    __tablename__ = 'document_sequence'
    id = db.Column(db.Integer, primary_key=True)
    sequence_key = db.Column(db.String(220), nullable=False, unique=True, index=True)
    next_value = db.Column(db.Integer, nullable=False, default=1)

class MascotPreference(db.Model):
    __tablename__ = 'mascot_preference'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, unique=True, index=True)
    enabled = db.Column(db.Boolean, default=True)
    intensity = db.Column(db.String(24), default='full')
    size = db.Column(db.String(16), default='medium')
    position = db.Column(db.String(24), default='bottom-right')
    operational_updates = db.Column(db.Boolean, default=True)
    motivational_messages = db.Column(db.Boolean, default=True)
    weather_reactions = db.Column(db.Boolean, default=True)
    weather_city = db.Column(db.String(120), default='Gurugram')
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class ElectricityConnection(db.Model):
    __tablename__ = 'electricity_connection'
    id = db.Column(db.Integer, primary_key=True)
    city_id = db.Column(db.Integer, db.ForeignKey('city.id'), nullable=True)
    property_name = db.Column(db.String(180), default='')
    provider_id = db.Column(db.Integer, db.ForeignKey('electricity_provider.id'), nullable=False)
    connection_name = db.Column(db.String(180), default='')
    consumer_name = db.Column(db.String(180), default='')
    identifier_primary = db.Column(db.String(180), nullable=False)
    identifier_primary_type = db.Column(db.String(40), nullable=False, default='CONSUMER_NO')
    identifier_secondary = db.Column(db.String(180), default='')
    identifier_secondary_type = db.Column(db.String(40), default='')
    meter_number = db.Column(db.String(120), default='')
    billing_cycle = db.Column(db.String(80), default='Monthly')
    reminder_days_before = db.Column(db.Integer, default=5)
    vault_credential_id = db.Column(db.Integer, db.ForeignKey('vault_secret.id'), nullable=True)
    status = db.Column(db.String(32), default='active')
    last_fetch_status = db.Column(db.String(48), default='')
    last_fetch_at = db.Column(db.DateTime, nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    provider = db.relationship('ElectricityProvider', foreign_keys=[provider_id])
    city_ref = db.relationship('City', foreign_keys=[city_id])

class ElectricityBill(db.Model):
    __tablename__ = 'electricity_bill'
    id = db.Column(db.Integer, primary_key=True)
    connection_id = db.Column(db.Integer, db.ForeignKey('electricity_connection.id'), nullable=False, index=True)
    provider_id = db.Column(db.Integer, db.ForeignKey('electricity_provider.id'), nullable=False, index=True)
    dedupe_key = db.Column(db.String(320), unique=True, nullable=False)
    bill_month = db.Column(db.String(24), default='')
    billing_period_start = db.Column(db.Date, nullable=True)
    billing_period_end = db.Column(db.Date, nullable=True)
    bill_number = db.Column(db.String(140), default='')
    bill_date = db.Column(db.Date, nullable=True)
    due_date = db.Column(db.Date, nullable=True)
    consumer_name = db.Column(db.String(180), default='')
    meter_number = db.Column(db.String(120), default='')
    units_consumed = db.Column(db.Numeric(14,3), nullable=True)
    previous_reading = db.Column(db.Numeric(14,3), nullable=True)
    current_reading = db.Column(db.Numeric(14,3), nullable=True)
    current_charges = db.Column(db.Numeric(14,2), default=0)
    arrears_amount = db.Column(db.Numeric(14,2), default=0)
    late_fee_amount = db.Column(db.Numeric(14,2), default=0)
    net_amount = db.Column(db.Numeric(14,2), default=0)
    total_due_amount = db.Column(db.Numeric(14,2), default=0)
    status = db.Column(db.String(48), default='unpaid')
    source_type = db.Column(db.String(48), default='manual_entry')
    raw_source_meta_json = db.Column(db.Text, default='{}')
    receipt_file_path_or_token = db.Column(db.Text, default='')
    bill_file_path_or_token = db.Column(db.Text, default='')
    bill_file_name = db.Column(db.String(255), default='')
    bill_mime_type = db.Column(db.String(120), default='')
    encrypted_bill_blob = db.Column(db.LargeBinary, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    connection = db.relationship('ElectricityConnection', foreign_keys=[connection_id])
    provider = db.relationship('ElectricityProvider', foreign_keys=[provider_id])

class ElectricityPayment(db.Model):
    __tablename__ = 'electricity_payment'
    id = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('electricity_bill.id'), nullable=False, index=True)
    connection_id = db.Column(db.Integer, db.ForeignKey('electricity_connection.id'), nullable=False)
    provider_id = db.Column(db.Integer, db.ForeignKey('electricity_provider.id'), nullable=False)
    payment_provider = db.Column(db.String(120), default='')
    payment_reference = db.Column(db.String(180), default='')
    provider_txn_id = db.Column(db.String(180), default='')
    paid_amount = db.Column(db.Numeric(14,2), default=0)
    initiated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    confirmed_at = db.Column(db.DateTime, nullable=True)
    status = db.Column(db.String(48), default='initiated')
    receipt_path_or_token = db.Column(db.Text, default='')
    meta_json = db.Column(db.Text, default='{}')
    bill = db.relationship('ElectricityBill', foreign_keys=[bill_id])

class ReminderItem(db.Model):
    __tablename__ = 'reminder_item'
    id = db.Column(db.Integer, primary_key=True)
    module = db.Column(db.String(60), nullable=False, default='electricity')
    entity_id = db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(240), nullable=False)
    severity = db.Column(db.String(24), default='info')
    due_at = db.Column(db.DateTime, nullable=True)
    status = db.Column(db.String(32), default='active')
    payload_json = db.Column(db.Text, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    __table_args__=(db.UniqueConstraint('module','entity_id',name='uq_reminder_entity'),)

class AuditEvent(db.Model):
    __tablename__ = 'audit_event'
    id = db.Column(db.Integer, primary_key=True)
    actor_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    module = db.Column(db.String(60), nullable=False)
    action = db.Column(db.String(120), nullable=False)
    target_type = db.Column(db.String(80), default='')
    target_id = db.Column(db.Integer, nullable=True)
    status = db.Column(db.String(32), default='success')
    note = db.Column(db.Text, default='')
    meta_json = db.Column(db.Text, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class FoodIntegration(db.Model):
    __tablename__ = 'food_integration'
    id = db.Column(db.Integer, primary_key=True)
    platform = db.Column(db.String(60), nullable=False, default='Other')
    display_name = db.Column(db.String(160), default='')
    outlet_id = db.Column(db.String(180), default='')
    account_identifier = db.Column(db.String(180), default='')
    portal_url = db.Column(db.Text, default='')
    developer_url = db.Column(db.Text, default='')
    api_base_url = db.Column(db.Text, default='')
    api_token_env = db.Column(db.String(120), default='')
    api_key_env = db.Column(db.String(120), default='')
    webhook_enabled = db.Column(db.Boolean, default=True)
    api_enabled = db.Column(db.Boolean, default=False)
    active = db.Column(db.Boolean, default=True)
    last_sync_at = db.Column(db.DateTime, nullable=True)
    last_sync_status = db.Column(db.Text, default='')
    last_sync_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)


BANK_PORTALS = [
    # Public sector banks — current official corporate domains; .bank.in is preferred where live.
    {'key':'sbi','name':'State Bank of India','category':'Public Sector','url':'https://sbi.bank.in/'},
    {'key':'bob','name':'Bank of Baroda','category':'Public Sector','url':'https://bankofbaroda.bank.in/'},
    {'key':'boi','name':'Bank of India','category':'Public Sector','url':'https://bankofindia.co.in/'},
    {'key':'bom','name':'Bank of Maharashtra','category':'Public Sector','url':'https://bankofmaharashtra.in/'},
    {'key':'canara','name':'Canara Bank','category':'Public Sector','url':'https://www.canarabank.bank.in/'},
    {'key':'central','name':'Central Bank of India','category':'Public Sector','url':'https://www.centralbankofindia.co.in/'},
    {'key':'indian','name':'Indian Bank','category':'Public Sector','url':'https://www.indianbank.in/'},
    {'key':'iob','name':'Indian Overseas Bank','category':'Public Sector','url':'https://www.iob.in/'},
    {'key':'pnb','name':'Punjab National Bank','category':'Public Sector','url':'https://pnb.bank.in/'},
    {'key':'psb','name':'Punjab & Sind Bank','category':'Public Sector','url':'https://punjabandsindbank.co.in/'},
    {'key':'uco','name':'UCO Bank','category':'Public Sector','url':'https://www.ucobank.com/'},
    {'key':'union','name':'Union Bank of India','category':'Public Sector','url':'https://www.unionbankofindia.co.in/'},
    # Private sector banks.
    {'key':'axis','name':'Axis Bank','category':'Private Sector','url':'https://www.axis.bank.in/'},
    {'key':'bandhan','name':'Bandhan Bank','category':'Private Sector','url':'https://bandhanbank.com/'},
    {'key':'csb','name':'CSB Bank','category':'Private Sector','url':'https://www.csb.co.in/'},
    {'key':'cityunion','name':'City Union Bank','category':'Private Sector','url':'https://www.cityunionbank.com/'},
    {'key':'dcb','name':'DCB Bank','category':'Private Sector','url':'https://www.dcbbank.com/'},
    {'key':'dhanlaxmi','name':'Dhanlaxmi Bank','category':'Private Sector','url':'https://www.dhanbank.com/'},
    {'key':'federal','name':'Federal Bank','category':'Private Sector','url':'https://www.federal.bank.in/'},
    {'key':'hdfc','name':'HDFC Bank','category':'Private Sector','url':'https://www.hdfc.bank.in/'},
    {'key':'icici','name':'ICICI Bank','category':'Private Sector','url':'https://www.icici.bank.in/'},
    {'key':'indusind','name':'IndusInd Bank','category':'Private Sector','url':'https://www.indusind.com/'},
    {'key':'idfc','name':'IDFC FIRST Bank','category':'Private Sector','url':'https://www.idfcfirst.bank.in/'},
    {'key':'jk','name':'Jammu & Kashmir Bank','category':'Private Sector','url':'https://www.jkbank.com/'},
    {'key':'karnataka','name':'Karnataka Bank','category':'Private Sector','url':'https://karnatakabank.com/'},
    {'key':'kvb','name':'Karur Vysya Bank','category':'Private Sector','url':'https://www.kvb.co.in/'},
    {'key':'kotak','name':'Kotak Mahindra Bank','category':'Private Sector','url':'https://www.kotak.bank.in/'},
    {'key':'nainital','name':'Nainital Bank','category':'Private Sector','url':'https://www.nainitalbank.co.in/'},
    {'key':'rbl','name':'RBL Bank','category':'Private Sector','url':'https://www.rblbank.com/'},
    {'key':'sib','name':'South Indian Bank','category':'Private Sector','url':'https://www.southindianbank.com/'},
    {'key':'tmb','name':'Tamilnad Mercantile Bank','category':'Private Sector','url':'https://www.tmb.in/'},
    {'key':'yes','name':'YES BANK','category':'Private Sector','url':'https://www.yes.bank.in/'},
    {'key':'idbi','name':'IDBI Bank','category':'Private Sector','url':'https://www.idbibank.in/'},
    {'key':'dbs','name':'DBS Bank India','category':'Private / Foreign Subsidiary','url':'https://www.dbs.com/in/'},
    # Small finance banks from the RBI banking list.
    {'key':'au','name':'AU Small Finance Bank','category':'Small Finance','url':'https://www.au.bank.in/'},
    {'key':'capital','name':'Capital Small Finance Bank','category':'Small Finance','url':'https://www.capitalbank.co.in/'},
    {'key':'equitas','name':'Equitas Small Finance Bank','category':'Small Finance','url':'https://www.equitasbank.com/'},
    {'key':'esaf','name':'ESAF Small Finance Bank','category':'Small Finance','url':'https://www.esafbank.com/'},
    {'key':'jana','name':'Jana Small Finance Bank','category':'Small Finance','url':'https://www.janabank.com/'},
    {'key':'shivalik','name':'Shivalik Small Finance Bank','category':'Small Finance','url':'https://shivalikbank.com/'},
    {'key':'suryoday','name':'Suryoday Small Finance Bank','category':'Small Finance','url':'https://www.suryodaybank.com/'},
    {'key':'ujjivan','name':'Ujjivan Small Finance Bank','category':'Small Finance','url':'https://www.ujjivansfb.in/'},
    {'key':'unity','name':'Unity Small Finance Bank','category':'Small Finance','url':'https://theunitybank.com/'},
    {'key':'utkarsh','name':'Utkarsh Small Finance Bank','category':'Small Finance','url':'https://www.utkarsh.bank/'},
    {'key':'slice','name':'slice Small Finance Bank','category':'Small Finance','url':'https://slice.bank.in/'},
    # Payments banks.
    {'key':'airtel','name':'Airtel Payments Bank','category':'Payments Bank','url':'https://www.airtel.in/bank/'},
    {'key':'ippb','name':'India Post Payments Bank','category':'Payments Bank','url':'https://www.ippbonline.com/'},
    {'key':'fino','name':'Fino Payments Bank','category':'Payments Bank','url':'https://www.finobank.com/'},
    {'key':'jio','name':'Jio Payments Bank','category':'Payments Bank','url':'https://www.jiopaymentsbank.com/'},
    {'key':'paytm','name':'Paytm Payments Bank','category':'Payments Bank','url':'https://www.paytmbank.com/'},
]

BANK_ALLOWED_EXTENSIONS={'.csv','.xlsx','.xls','.pdf'}
BANK_MAX_FILE_BYTES=16*1024*1024


def _bank_encrypt_bytes(raw):
    return _integration_cipher().encrypt(raw)


def _bank_decrypt_bytes(raw):
    try: return _integration_cipher().decrypt(bytes(raw or b''))
    except Exception: return b''


def _bank_encrypt_json(value):
    packed=json.dumps(value,ensure_ascii=False,separators=(',',':')).encode('utf-8')
    return _integration_cipher().encrypt(packed).decode('ascii')


def _bank_decrypt_json(value):
    try:
        raw=_integration_cipher().decrypt((value or '').encode('ascii')).decode('utf-8')
        data=json.loads(raw); return data if isinstance(data,(list,dict)) else []
    except Exception: return []


def _bank_clean_name(name):
    clean=secure_filename(name or 'bank-document') or 'bank-document'
    return clean[:240]


def _bank_cell(value):
    if value is None: return ''
    if isinstance(value,(datetime.datetime,datetime.date)): return value.strftime('%Y-%m-%d')
    return str(value).strip()


def _bank_rows_from_matrix(matrix):
    rows=[[ _bank_cell(v) for v in row ] for row in matrix if any(_bank_cell(v) for v in row)]
    if not rows: return []
    header_index=0
    # Bank exports often place account metadata above the real transaction header.
    hints=('date','narration','description','particular','debit','credit','withdraw','deposit','amount','balance','reference','ref','utr')
    best=(-1,0)
    for i,row in enumerate(rows[:30]):
        joined=' '.join(x.lower() for x in row)
        score=sum(1 for h in hints if h in joined)
        if score>best[0]: best=(score,i)
    if best[0]>=2: header_index=best[1]
    headers=[]; seen={}
    for idx,val in enumerate(rows[header_index]):
        h=(val or f'Column {idx+1}').strip() or f'Column {idx+1}'
        key=h; n=seen.get(h.lower(),0)+1; seen[h.lower()]=n
        if n>1: key=f'{h} {n}'
        headers.append(key)
    output=[]
    for row in rows[header_index+1:]:
        if not any(str(x).strip() for x in row): continue
        if len(row)<len(headers): row=row+['']*(len(headers)-len(row))
        output.append({headers[i]:_bank_cell(row[i]) for i in range(len(headers))})
        if len(output)>=15000: break
    return output


def _bank_parse_pdf(raw):
    try:
        import fitz
        doc=fitz.open(stream=raw,filetype='pdf'); lines=[]
        for page_index in range(min(80,doc.page_count)):
            page=doc.load_page(page_index)
            lines.extend([x.strip() for x in (page.get_text('text') or '').splitlines() if x.strip()])
        date_re=re.compile(r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}[- ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]\d{2,4})\b',re.I)
        money_re=re.compile(r'(?<!\d)(?:₹\s*)?-?\d[\d,]*\.\d{2}(?:\s*(?:CR|DR))?',re.I)
        out=[]
        for line in lines:
            dm=date_re.search(line); amounts=money_re.findall(line)
            if not dm or not amounts: continue
            rest=(line[:dm.start()]+' '+line[dm.end():]).strip()
            for a in amounts: rest=rest.replace(a,' ',1)
            rest=re.sub(r'\s+',' ',rest).strip(' -|')
            item={'Date':dm.group(0),'Narration':rest,'Amount':'','Debit':'','Credit':'','Balance':''}
            clean_amounts=[a.replace('₹','').strip() for a in amounts]
            if len(clean_amounts)>=2: item['Amount']=clean_amounts[-2]; item['Balance']=clean_amounts[-1]
            else: item['Amount']=clean_amounts[0]
            token=' '.join(amounts).upper()
            if 'DR' in token: item['Debit']=item['Amount']
            elif 'CR' in token: item['Credit']=item['Amount']
            out.append(item)
        return out
    except Exception:
        return []


def _bank_parse_rows(raw, filename):
    ext=Path(filename or '').suffix.lower()
    try:
        if ext=='.csv':
            text=''
            for enc in ('utf-8-sig','utf-8','cp1252','latin-1'):
                try: text=raw.decode(enc); break
                except Exception: pass
            sample=text[:5000]
            try: dialect=csv.Sniffer().sniff(sample,delimiters=',;\t|')
            except Exception: dialect=csv.excel
            return _bank_rows_from_matrix(list(csv.reader(io.StringIO(text),dialect))[:16000])
        if ext=='.xlsx':
            from openpyxl import load_workbook
            wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
            ws=wb[wb.sheetnames[0]]
            return _bank_rows_from_matrix([list(row) for row in ws.iter_rows(values_only=True)])
        if ext=='.xls':
            try:
                import xlrd
                book=xlrd.open_workbook(file_contents=raw,on_demand=True); sheet=book.sheet_by_index(0)
                return _bank_rows_from_matrix([sheet.row_values(i) for i in range(sheet.nrows)])
            except Exception: return []
        if ext=='.pdf': return _bank_parse_pdf(raw)
    except Exception as exc:
        app.logger.warning('Bank statement parse failed for %s: %s',filename,str(exc)[:180])
    return []


def _bank_key(value):
    return re.sub(r'[^a-z0-9]','',str(value or '').lower())


def _bank_float(value):
    raw=str(value or '').strip().upper().replace('₹','').replace('INR','').replace(',','')
    negative=raw.startswith('(') and raw.endswith(')')
    raw=raw.replace('CR','').replace('DR','').replace('(','').replace(')','').strip()
    raw=re.sub(r'[^0-9.\-]','',raw)
    try: n=float(raw or 0); return -abs(n) if negative else n
    except Exception: return 0.0


def _bank_date(value):
    raw=str(value or '').strip()
    if not raw: return None
    raw=re.sub(r'\s+',' ',raw)
    for fmt in ('%Y-%m-%d','%d/%m/%Y','%d-%m-%Y','%d.%m.%Y','%d/%m/%y','%d-%m-%y','%d %b %Y','%d-%b-%Y','%d %B %Y'):
        try: return datetime.datetime.strptime(raw[:20],fmt).date()
        except Exception: pass
    return None


def _bank_pick(row, synonyms):
    keyed={_bank_key(k):v for k,v in (row or {}).items()}
    for syn in synonyms:
        sk=_bank_key(syn)
        for k,v in keyed.items():
            if k==sk or sk in k: return v
    return ''


def _bank_canonical(row):
    date=_bank_pick(row,['transaction date','txn date','posting date','value date','date'])
    debit=_bank_pick(row,['debit amount','withdrawal amount','withdrawals','debit','withdrawal'])
    credit=_bank_pick(row,['credit amount','deposit amount','deposits','credit','deposit'])
    generic=_bank_pick(row,['transaction amount','txn amount','amount'])
    dtype=str(_bank_pick(row,['dr/cr','type','transaction type'])).upper()
    debit_n=abs(_bank_float(debit)); credit_n=abs(_bank_float(credit)); amount=_bank_float(generic)
    if debit_n: amount=-debit_n
    elif credit_n: amount=credit_n
    elif generic:
        raw=str(generic).upper()
        if 'DR' in raw or dtype.startswith('D'): amount=-abs(amount)
        elif 'CR' in raw or dtype.startswith('C'): amount=abs(amount)
    desc=_bank_pick(row,['narration','description','particulars','particular','remarks','details','transaction details'])
    ref=_bank_pick(row,['utr number','utr','reference number','reference','ref no','transaction id','txn id','cheque number','chq no'])
    balance=_bank_pick(row,['closing balance','running balance','balance'])
    return {'date':_bank_date(date),'date_raw':str(date or ''),'amount':round(amount,2),'abs_amount':round(abs(amount),2),'description':str(desc or ''),'reference':str(ref or ''),'balance':_bank_float(balance),'source':row}


def _bank_tokens(value):
    stop={'the','and','for','from','to','in','of','by','a','an','upi','neft','imps','rtgs','transfer','payment','txn'}
    return {x for x in re.findall(r'[a-z0-9]{3,}',str(value or '').lower()) if x not in stop}


def _bank_match_score(expected, actual):
    score=0.0; weight=0.0; reasons=[]
    if expected['abs_amount']>0:
        weight+=.55
        diff=abs(expected['abs_amount']-actual['abs_amount'])
        if diff<=.01: score+=.55; reasons.append('amount')
        elif diff<=1: score+=.35
    if expected['date']:
        weight+=.22
        if actual['date']:
            days=abs((expected['date']-actual['date']).days)
            if days==0: score+=.22; reasons.append('date')
            elif days<=1: score+=.16
            elif days<=3: score+=.08
    er=_bank_key(expected['reference'])
    if er:
        weight+=.18; ar=_bank_key(actual['reference'])
        if ar and (er==ar or er in ar or ar in er): score+=.18; reasons.append('reference')
    et=_bank_tokens(expected['description'])
    if et:
        weight+=.15; at=_bank_tokens(actual['description'])
        overlap=len(et & at)/max(1,len(et))
        score+=.15*min(1,overlap)
        if overlap>=.5: reasons.append('description')
    if weight<=0: return 0.0,[]
    return min(1.0,score/weight),reasons


def _bank_reconcile(statement_rows, template_rows):
    statement=[_bank_canonical(r) for r in statement_rows]
    expected=[_bank_canonical(r) for r in template_rows]
    used=set(); matches=[]
    for idx,e in enumerate(expected):
        best_score=0; best_i=None; best_reasons=[]
        for j,a in enumerate(statement):
            if j in used: continue
            sc,reasons=_bank_match_score(e,a)
            if sc>best_score: best_score,best_i,best_reasons=sc,j,reasons
        status='missing'; actual=None
        if best_i is not None and best_score>=.72:
            status='matched'; used.add(best_i); actual=statement[best_i]
        elif best_i is not None and best_score>=.50:
            status='review'; used.add(best_i); actual=statement[best_i]
        matches.append({'index':idx+1,'status':status,'score':round(best_score*100),'reasons':best_reasons,'expected':_bank_serializable(e),'actual':_bank_serializable(actual) if actual else None})
    extras=[_bank_serializable(statement[i]) for i in range(len(statement)) if i not in used]
    summary={'expected':len(expected),'statement':len(statement),'matched':sum(1 for x in matches if x['status']=='matched'),'review':sum(1 for x in matches if x['status']=='review'),'missing':sum(1 for x in matches if x['status']=='missing'),'extra':len(extras)}
    summary['match_rate']=round((summary['matched']/max(1,summary['expected']))*100)
    return {'summary':summary,'matches':matches,'extras':extras}


def _bank_serializable(row):
    if not row: return {}
    return {'date':row['date'].isoformat() if row.get('date') else row.get('date_raw',''),'amount':row.get('amount',0),'description':row.get('description',''),'reference':row.get('reference',''),'balance':row.get('balance',0),'source':row.get('source',{})}


OFFICIAL_FOOD_PORTALS = {
    'Swiggy': {
        'portal_url': 'https://partner.swiggy.com/login',
        'developer_url': 'https://developers.swiggy.com/login',
    },
    'Zomato': {
        'portal_url': 'https://www.zomato.com/partners/onlineordering/orders/',
        'developer_url': 'https://www.zomato.com/business/merchant-app',
    },
    'Toing': {
        'portal_url': 'https://www.toingit.com/',
        'developer_url': '',
    },
}

STALE_FOOD_PORTAL_URLS = {
    'https://partner.swiggy.com/v2/': OFFICIAL_FOOD_PORTALS['Swiggy']['portal_url'],
    'https://partner.swiggy.com/v2': OFFICIAL_FOOD_PORTALS['Swiggy']['portal_url'],
    'https://www.zomato.com/partners': OFFICIAL_FOOD_PORTALS['Zomato']['portal_url'],
    'https://www.zomato.com/partners/': OFFICIAL_FOOD_PORTALS['Zomato']['portal_url'],
}


def ensure_default_food_integrations():
    # Seed the three official starting portals only for a brand-new database.
    # Once the user has configured/removed integrations, respect that choice.
    if FoodIntegration.query.count():
        # Upgrade only obsolete built-in URLs; never overwrite a custom portal.
        changed=False
        for row in FoodIntegration.query.all():
            replacement=STALE_FOOD_PORTAL_URLS.get((row.portal_url or '').strip())
            if replacement:
                row.portal_url=replacement;changed=True
        if changed: db.session.commit()
        return
    for platform,urls in OFFICIAL_FOOD_PORTALS.items():
        db.session.add(FoodIntegration(
            platform=platform,
            display_name=f'{platform} Restaurant Partner',
            portal_url=urls.get('portal_url',''),
            developer_url=urls.get('developer_url',''),
            webhook_enabled=True,
            active=True,
        ))
    db.session.commit()


def _food_float(value):
    try:
        if isinstance(value,str): value=value.replace(',','').replace('₹','').strip()
        return float(value or 0)
    except Exception:
        return 0.0


def _food_get(row, *keys, default=''):
    if not isinstance(row,dict): return default
    for key in keys:
        cur=row
        ok=True
        for part in str(key).split('.'):
            if isinstance(cur,dict) and part in cur:
                cur=cur.get(part)
            else:
                ok=False;break
        if ok and cur not in (None,''):
            return cur
    return default


def _food_records(payload):
    if isinstance(payload,list): return [x for x in payload if isinstance(x,dict)]
    if not isinstance(payload,dict): return []
    for key in ('orders','records','results','items','data'):
        val=payload.get(key)
        if isinstance(val,list): return [x for x in val if isinstance(x,dict)]
        if isinstance(val,dict):
            for nested in ('orders','records','results','items','data'):
                n=val.get(nested)
                if isinstance(n,list): return [x for x in n if isinstance(x,dict)]
    return [payload]


def _upsert_food_order(platform, row, default_outlet=''):
    platform=(platform or 'Direct').strip()[:50]
    order_id=str(_food_get(row,'order_id','orderId','orderID','id','order.id','order.uuid',default=''))[:120]
    outlet=str(_food_get(row,'outlet','outlet_name','restaurant_name','restaurant.name','store.name',default=default_outlet))[:180]
    customer=str(_food_get(row,'customer','customer_name','customer.name','user.name',default=''))[:180]
    order_time=str(_food_get(row,'order_time','ordered_at','created_at','createdAt','order.created_at',default=''))[:60]
    status=str(_food_get(row,'status','order_status','order.status',default='New'))[:50]
    payment_mode=str(_food_get(row,'payment_mode','paymentMode','payment.method','payment_type',default=''))[:50]
    gross=_food_float(_food_get(row,'gross','amount','order_total','orderTotal','total','bill.total','order.amount',default=0))
    commission=_food_float(_food_get(row,'commission','commission_amount','charges.commission',default=0))
    fees=_food_float(_food_get(row,'fees','fee','platform_fee','charges.fees',default=0))
    taxes=_food_float(_food_get(row,'taxes','tax','gst','charges.taxes',default=0))
    net=_food_float(_food_get(row,'net','net_amount','netAmount','settlement_amount','payout_amount',default=0))
    if not net: net=gross-commission-fees-taxes
    settlement=str(_food_get(row,'settlement_status','settlementStatus','payout_status',default='Pending'))[:50]
    obj=None
    if order_id:
        obj=FoodOrder.query.filter(func.lower(FoodOrder.platform)==platform.lower(),FoodOrder.order_id==order_id).first()
    if not obj:
        obj=FoodOrder(platform=platform,order_id=order_id)
        db.session.add(obj)
    obj.outlet=outlet;obj.customer=customer;obj.order_time=order_time;obj.status=status;obj.payment_mode=payment_mode
    obj.gross=gross;obj.commission=commission;obj.fees=fees;obj.taxes=taxes;obj.net=net;obj.settlement_status=settlement
    return obj


def _ingest_food_payload(platform,payload,default_outlet=''):
    rows=_food_records(payload);count=0
    for row in rows:
        _upsert_food_order(platform,row,default_outlet=default_outlet);count+=1
    if count: db.session.commit()
    return count


class VideoAsset(db.Model):
    __tablename__ = 'video_asset'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(220), default='', nullable=False)
    media_type = db.Column(db.String(20), default='video', nullable=False)
    storage_path = db.Column(db.Text, default='')
    public_url = db.Column(db.Text, default='', nullable=False)
    mime_type = db.Column(db.String(120), default='')
    file_size = db.Column(db.BigInteger, default=0)
    uploaded_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class VideoScreen(db.Model):
    __tablename__ = 'video_screen'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(180), nullable=False)
    player_token = db.Column(db.String(180), nullable=False, unique=True)
    city = db.Column(db.String(120), default='')
    location_name = db.Column(db.String(220), default='')
    device_label = db.Column(db.String(180), default='')
    current_asset_id = db.Column(db.Integer, db.ForeignKey('video_asset.id'), nullable=True)
    playlist_json = db.Column(db.Text, default='[]')
    rotation_degrees = db.Column(db.Integer, default=0)
    fit_mode = db.Column(db.String(20), default='contain')
    loop_media = db.Column(db.Boolean, default=True)
    muted = db.Column(db.Boolean, default=True)
    enabled = db.Column(db.Boolean, default=True)
    slide_duration_seconds = db.Column(db.Integer, default=10)
    last_seen_at = db.Column(db.DateTime, nullable=True)
    last_ip = db.Column(db.String(120), default='')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class FestiveSession(db.Model):
    __tablename__ = 'festive_session'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(180), default='Festive Takeover', nullable=False)
    asset_id = db.Column(db.Integer, db.ForeignKey('video_asset.id'), nullable=True)
    active = db.Column(db.Boolean, default=False)
    started_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    started_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    ended_at = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text, default='')

class QueryLead(db.Model):
    __tablename__ = 'query_lead'
    id = db.Column(db.Integer, primary_key=True)
    source = db.Column(db.String(60), default='Manual')
    external_id = db.Column(db.String(180), default='')
    city = db.Column(db.String(120), default='')
    property_name = db.Column(db.String(180), default='')
    customer_name = db.Column(db.String(180), default='')
    mobile = db.Column(db.String(40), default='')
    whatsapp = db.Column(db.String(40), default='')
    email = db.Column(db.String(180), default='')
    query_text = db.Column(db.Text, default='')
    budget = db.Column(db.String(80), default='')
    move_in_date = db.Column(db.String(40), default='')
    stay_type = db.Column(db.String(80), default='')
    status = db.Column(db.String(40), default='Live')
    heat = db.Column(db.String(20), default='Warm')
    score = db.Column(db.Integer, default=50)
    assigned_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    next_follow_up = db.Column(db.String(60), default='')
    notes = db.Column(db.Text, default='')
    raw_json = db.Column(db.Text, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class QueryTemplate(db.Model):
    __tablename__ = 'query_template'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False)
    category = db.Column(db.String(60), default='General')
    message = db.Column(db.Text, default='')
    whatsapp_template_name = db.Column(db.String(160), default='')
    sources_json = db.Column(db.Text, default='[]')
    statuses_json = db.Column(db.Text, default='[]')
    auto_send = db.Column(db.Boolean, default=False)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class QueryActivity(db.Model):
    __tablename__ = 'query_activity'
    id = db.Column(db.Integer, primary_key=True)
    query_id = db.Column(db.Integer, db.ForeignKey('query_lead.id'), nullable=False)
    action = db.Column(db.String(80), default='Update')
    details = db.Column(db.Text, default='')
    actor_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class WebAuthnCredential(db.Model):
    __tablename__ = 'web_authn_credential'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    credential_id = db.Column(db.LargeBinary, unique=True, nullable=False)
    public_key = db.Column(db.LargeBinary, nullable=False)
    sign_count = db.Column(db.BigInteger, default=0, nullable=False)
    transports = db.Column(db.Text, default='[]')
    device_name = db.Column(db.String(180), default='Windows Hello / fingerprint')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    last_used_at = db.Column(db.DateTime, nullable=True)

class WhatsAppMessage(db.Model):
    __tablename__ = 'whatsapp_message'
    id = db.Column(db.Integer, primary_key=True)
    direction = db.Column(db.String(12), nullable=False, default='outbound')
    contact_name = db.Column(db.String(180), default='')
    wa_id = db.Column(db.String(40), nullable=False, default='', index=True)
    message_id = db.Column(db.String(220), unique=True, nullable=True)
    message_type = db.Column(db.String(40), default='text')
    body = db.Column(db.Text, default='')
    status = db.Column(db.String(40), default='sent')
    raw_json = db.Column(db.Text, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, index=True)

class DriveFile(db.Model):
    __tablename__ = 'drive_file'
    id = db.Column(db.Integer, primary_key=True)
    provider_file_id = db.Column(db.String(220), unique=True, nullable=False)
    name = db.Column(db.String(300), nullable=False)
    mime_type = db.Column(db.String(180), default='application/octet-stream')
    file_size = db.Column(db.BigInteger, default=0)
    web_view_link = db.Column(db.Text, default='')
    source = db.Column(db.String(80), default='manual')
    uploaded_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class Setting(db.Model):
    key = db.Column(db.String(120), primary_key=True)
    value = db.Column(db.Text, default='')


def setting(key, default=''):
    try:
        row = db.session.get(Setting, key)
        return row.value if row else default
    except Exception as exc:
        # Settings are presentation/runtime preferences. A stale or temporarily
        # unavailable settings table must never take down login or the shell.
        try:
            db.session.rollback()
        except Exception:
            pass
        try:
            app.logger.warning('Setting lookup failed for %s: %s', key, str(exc)[:180])
        except Exception:
            pass
        return default

def set_setting(key, value):
    row = db.session.get(Setting, key)
    if row: row.value = value
    else: db.session.add(Setting(key=key, value=value))
    db.session.commit()


GOOGLE_SCOPES = [
    'https://www.googleapis.com/auth/drive.file',
    'https://www.googleapis.com/auth/gmail.readonly',
    'https://www.googleapis.com/auth/gmail.send',
]

def _integration_cipher():
    """Encrypt provider refresh tokens at rest; biometrics are never stored here."""
    from cryptography.fernet import Fernet
    configured=os.getenv('INTEGRATION_ENCRYPTION_KEY','').strip()
    if configured:
        try: return Fernet(configured.encode('ascii'))
        except Exception:
            derived=base64.urlsafe_b64encode(hashlib.sha256(configured.encode()).digest())
            return Fernet(derived)
    derived=base64.urlsafe_b64encode(hashlib.sha256(str(app.config['SECRET_KEY']).encode()).digest())
    return Fernet(derived)

def _encrypted_setting_get(key):
    raw=setting(key,'')
    if not raw: return ''
    try: return _integration_cipher().decrypt(raw.encode('ascii')).decode('utf-8')
    except Exception: return ''

def _encrypted_setting_set(key, value):
    set_setting(key,_integration_cipher().encrypt(value.encode('utf-8')).decode('ascii') if value else '')

def google_oauth_configured():
    return bool(os.getenv('GOOGLE_CLIENT_ID','').strip() and os.getenv('GOOGLE_CLIENT_SECRET','').strip())

def google_connected():
    return bool(_encrypted_setting_get('google_oauth_token'))

def _google_token_data(refresh=True):
    raw=_encrypted_setting_get('google_oauth_token')
    if not raw: return {}
    try: data=json.loads(raw)
    except Exception: return {}
    expires=float(data.get('expires_at') or 0)
    if refresh and (not data.get('access_token') or expires < datetime.datetime.utcnow().timestamp()+90):
        refresh_token=data.get('refresh_token')
        if not (refresh_token and google_oauth_configured()): return {}
        try:
            r=requests.post('https://oauth2.googleapis.com/token',data={
                'client_id':os.getenv('GOOGLE_CLIENT_ID','').strip(),
                'client_secret':os.getenv('GOOGLE_CLIENT_SECRET','').strip(),
                'refresh_token':refresh_token,'grant_type':'refresh_token'
            },timeout=25)
            if not r.ok: return {}
            fresh=r.json(); data.update(fresh); data['refresh_token']=refresh_token
            data['expires_at']=datetime.datetime.utcnow().timestamp()+int(fresh.get('expires_in') or 3600)
            _encrypted_setting_set('google_oauth_token',json.dumps(data))
        except Exception: return {}
    return data

def _google_headers():
    token=_google_token_data().get('access_token','')
    return {'Authorization':f'Bearer {token}'} if token else {}

def ensure_google_drive_folder():
    existing=(setting('google_drive_folder_id','') or os.getenv('GOOGLE_DRIVE_FOLDER_ID','')).strip()
    if existing: return existing
    headers=_google_headers()
    if not headers: return ''
    try:
        r=requests.post('https://www.googleapis.com/drive/v3/files',headers=dict(headers,**{'Content-Type':'application/json'}),params={'fields':'id'},json={'name':'Livenza Back Office','mimeType':'application/vnd.google-apps.folder'},timeout=25)
        if r.ok and r.json().get('id'):
            set_setting('google_drive_folder_id',r.json()['id']); return r.json()['id']
    except Exception: pass
    return ''

def google_drive_upload_bytes(data, filename, mime_type='application/octet-stream', source='manual', uploaded_by=None):
    """Upload through Drive's resumable API and keep non-secret metadata locally."""
    headers=_google_headers()
    if not headers: return None, 'Connect Google in Admin first.'
    folder=ensure_google_drive_folder()
    metadata={'name':filename}
    if folder: metadata['parents']=[folder]
    init_headers=dict(headers,**{'Content-Type':'application/json; charset=UTF-8','X-Upload-Content-Type':mime_type,'X-Upload-Content-Length':str(len(data))})
    try:
        start=requests.post('https://www.googleapis.com/upload/drive/v3/files',params={'uploadType':'resumable','fields':'id,name,mimeType,size,webViewLink'},headers=init_headers,json=metadata,timeout=25)
        if not start.ok or not start.headers.get('Location'):
            return None, f'Drive upload could not start ({start.status_code}): {start.text[:300]}'
        put=requests.put(start.headers['Location'],headers={'Content-Type':mime_type,'Content-Length':str(len(data))},data=data,timeout=180)
        if not put.ok: return None, f'Drive upload failed ({put.status_code}): {put.text[:300]}'
        info=put.json()
        row=DriveFile.query.filter_by(provider_file_id=info.get('id','')).first()
        if not row:
            row=DriveFile(provider_file_id=info.get('id',''),name=info.get('name') or filename)
            db.session.add(row)
        row.name=info.get('name') or filename; row.mime_type=info.get('mimeType') or mime_type
        row.file_size=int(info.get('size') or len(data)); row.web_view_link=info.get('webViewLink') or ''
        row.source=source; row.uploaded_by_user_id=(uploaded_by.id if uploaded_by else None)
        db.session.commit()
        return row, ''
    except Exception as exc:
        db.session.rollback(); return None, f'Drive upload failed: {exc}'

def _pattern_value(value):
    nodes=[]
    for part in str(value or '').split('-'):
        if part.isdigit() and 0 <= int(part) <= 8 and part not in nodes: nodes.append(part)
    return '-'.join(nodes) if len(nodes)>=4 else ''

def _b64url_decode(value):
    value=str(value or '')
    return base64.urlsafe_b64decode(value+'='*((4-len(value)%4)%4))

def _webauthn_context():
    rp_id=os.getenv('WEBAUTHN_RP_ID','').strip() or request.host.split(':')[0]
    origin=os.getenv('WEBAUTHN_ORIGIN','').strip() or request.host_url.rstrip('/')
    return rp_id,origin

MARKET_QUOTE_CACHE = {}
WEATHER_CACHE = {}

COMPANION_LOCATIONS = {
    'Gurugram': (28.4595, 77.0266),
    'Jaipur': (26.9124, 75.7873),
    'Delhi': (28.6139, 77.2090),
    'Mumbai': (19.0760, 72.8777),
    'Bengaluru': (12.9716, 77.5946),
}

COMPANION_QUOTES = [
    'Small improvements become remarkable systems.',
    'Hospitality is care made visible.',
    'A calm workspace creates confident decisions.',
    'Consistency turns everyday service into a trusted brand.',
    'Make the next useful move, then make it beautifully.',
    'Great operations feel effortless because the details are intentional.',
    'Today is a good day to make someone feel at home.',
    'Progress grows wherever attention and action meet.',
]

def _weather_description(code):
    code=int(code or 0)
    if code==0: return 'Clear sky'
    if code in (1,2): return 'Partly cloudy'
    if code==3: return 'Overcast'
    if code in (45,48): return 'Foggy'
    if code in (51,53,55,56,57): return 'Light drizzle'
    if code in (61,63,65,66,67): return 'Rain'
    if code in (71,73,75,77,85,86): return 'Snow'
    if code in (80,81,82): return 'Rain showers'
    if code in (95,96,99): return 'Thunderstorm'
    return 'Changing weather'

def _weather_effect(code, is_day=True):
    code=int(code or 0)
    if code in (95,96,99): return 'storm'
    if code in (51,53,55,56,57,61,63,65,66,67,80,81,82): return 'rain'
    if code in (71,73,75,77,85,86): return 'snow'
    if code in (45,48): return 'fog'
    if code in (1,2,3): return 'clouds'
    return 'sun' if is_day else 'night'

def _companion_weather(city):
    city=next((name for name in COMPANION_LOCATIONS if name.lower()==str(city or '').strip().lower()),None) or 'Gurugram'
    latitude,longitude=COMPANION_LOCATIONS[city]
    if city==setting('companion_default_city','Gurugram'):
        try:
            latitude=float(os.getenv('WEATHER_LATITUDE',latitude));longitude=float(os.getenv('WEATHER_LONGITUDE',longitude))
        except (TypeError,ValueError): pass
    now=datetime.datetime.utcnow().timestamp();cached=WEATHER_CACHE.get(city)
    if cached and now-cached['at']<600: return cached['value']
    params={
        'latitude':latitude,'longitude':longitude,
        'current':'temperature_2m,relative_humidity_2m,apparent_temperature,is_day,precipitation,rain,weather_code,wind_speed_10m',
        'daily':'weather_code,temperature_2m_max,temperature_2m_min,precipitation_probability_max',
        'timezone':'Asia/Kolkata','forecast_days':4,
    }
    try:
        response=requests.get('https://api.open-meteo.com/v1/forecast',params=params,headers={'User-Agent':'LivenzaLife-OperationsCloud/1.5.13'},timeout=10)
        response.raise_for_status();payload=response.json();current=payload.get('current') or {};daily=payload.get('daily') or {}
        code=int(current.get('weather_code') or 0);is_day=bool(int(current.get('is_day',1) or 0))
        dates=daily.get('time') or [];codes=daily.get('weather_code') or [];highs=daily.get('temperature_2m_max') or [];lows=daily.get('temperature_2m_min') or [];rain_chance=daily.get('precipitation_probability_max') or []
        forecast=[]
        for index,date in enumerate(dates[:4]):
            day_code=int(codes[index] if index<len(codes) else 0)
            forecast.append({
                'date':date,'condition':_weather_description(day_code),'effect':_weather_effect(day_code,True),
                'high':round(float(highs[index])) if index<len(highs) and highs[index] is not None else None,
                'low':round(float(lows[index])) if index<len(lows) and lows[index] is not None else None,
                'rain_chance':round(float(rain_chance[index])) if index<len(rain_chance) and rain_chance[index] is not None else 0,
            })
        value={
            'available':True,'city':city,'temperature':round(float(current.get('temperature_2m') or 0)),
            'feels_like':round(float(current.get('apparent_temperature') or current.get('temperature_2m') or 0)),
            'humidity':round(float(current.get('relative_humidity_2m') or 0)),
            'wind':round(float(current.get('wind_speed_10m') or 0)),
            'precipitation':round(float(current.get('precipitation') or 0),1),
            'condition':_weather_description(code),'effect':_weather_effect(code,is_day),'is_day':is_day,
            'forecast':forecast,'source':'Open-Meteo','updated_at':current.get('time') or datetime.datetime.now(ZoneInfo('Asia/Kolkata')).isoformat(),
        }
        WEATHER_CACHE[city]={'at':now,'value':value};return value
    except Exception:
        return {'available':False,'city':city,'temperature':None,'condition':'Weather temporarily unavailable','effect':'none','forecast':[],'source':'Open-Meteo'}

def _companion_operations():
    try:
        active=Tenant.query.filter(~Tenant.status.in_(['Vacated','Cancelled','Terminated'])).count()
        beds=0
        for room in Room.query.all():
            if room_status(room)=='Vacant':
                try: beds+=max(1,int(float(room.capacity or 1)))
                except Exception: beds+=1
        earned=float(db.session.query(func.coalesce(func.sum(FoodOrder.net),0)).scalar() or 0)
        hot=QueryLead.query.filter_by(heat='Hot').count()
        return [
            {'label':'Current tenants','value':str(active),'tone':'green','icon':'◉'},
            {'label':'Vacant beds','value':str(beds),'tone':'gold','icon':'▦'},
            {'label':'Amount earned','value':'₹'+format(earned,',.0f'),'tone':'neutral','icon':'₹'},
            {'label':'Hot queries','value':str(hot),'tone':'pink','icon':'Q'},
        ]
    except Exception:
        db.session.rollback();return []

def _moneycontrol_quote(label, url):
    """Small, cached, fail-soft reader for a user-selected official Moneycontrol quote page."""
    parsed=urllib.parse.urlparse(url)
    host=(parsed.hostname or '').lower()
    if parsed.scheme!='https' or not (host=='moneycontrol.com' or host.endswith('.moneycontrol.com')):
        return None
    cached=MARKET_QUOTE_CACHE.get(url); now=datetime.datetime.utcnow().timestamp()
    if cached and now-cached['at']<120: return cached['value']
    try:
        r=requests.get(url,headers={'User-Agent':'Mozilla/5.0 (compatible; LivenzaBackOffice/1.5; +https://www.moneycontrol.com/)'},timeout=12)
        if not r.ok: return None
        raw=r.text
        price=''
        for pattern in (
            r'id=["\'](?:nsecp|bsecp)["\'][^>]*>\s*([0-9][0-9,.]*)',
            r'class=["\'][^"\']*(?:inprice1|lastprice)[^"\']*["\'][^>]*>\s*([0-9][0-9,.]*)',
            r'["\'](?:lastPrice|last_price|price)["\']\s*:\s*["\']?([0-9][0-9,.]*)',
        ):
            m=re.search(pattern,raw,re.I)
            if m: price=m.group(1); break
        if not price: return None
        change=''
        for pattern in (r'id=["\'](?:nsechange|bsechange)["\'][^>]*>\s*([^<]{1,40})',r'["\'](?:percentChange|pChange)["\']\s*:\s*["\']?([-+0-9.]+)'):
            m=re.search(pattern,raw,re.I)
            if m: change=re.sub(r'<[^>]+>','',m.group(1)).strip(); break
        value={'label':label[:40],'value':price+((f' ({change}%)' if change and '%' not in change else f' ({change})') if change else ''),'url':url,'source':'Moneycontrol'}
        MARKET_QUOTE_CACHE[url]={'at':now,'value':value}; return value
    except Exception: return None

def live_marquee_items(user=None):
    user=user or current_user(); items=[]
    if setting('marquee_show_username','1')=='1' and user:
        items.append({'label':'Signed in','value':user.full_name or user.username,'tone':'neutral'})
    if setting('marquee_show_tenants','1')=='1':
        active=Tenant.query.filter(~Tenant.status.in_(['Vacated','Cancelled','Terminated'])).count()
        items.append({'label':'Current tenants','value':str(active),'tone':'green'})
    rooms=Room.query.all()
    if setting('marquee_show_vacant_beds','1')=='1':
        vacant=[r for r in rooms if room_status(r)=='Vacant']; beds=0
        for r in vacant:
            try: beds+=max(1,int(float(r.capacity or 1)))
            except Exception: beds+=1
        items.append({'label':'Vacant beds','value':str(beds),'tone':'gold'})
    if setting('marquee_show_earnings','1')=='1':
        earned=float(db.session.query(func.coalesce(func.sum(FoodOrder.net),0)).scalar() or 0)
        manual=setting('marquee_manual_earnings','').strip()
        items.append({'label':'Amount earned','value':manual or ('₹'+format(earned,',.0f')),'tone':'green'})
    if setting('marquee_show_favorites','0')=='1' and setting('marquee_favorites','').strip():
        items.append({'label':'Favourites','value':setting('marquee_favorites','').strip()[:180],'tone':'pink'})
    if setting('marquee_custom_text','').strip():
        items.append({'label':'Live update','value':setting('marquee_custom_text','').strip()[:240],'tone':'neutral'})
    if setting('marquee_show_stocks','0')=='1':
        configured=setting('marquee_stock_pages','').splitlines()
        if not configured:
            configured=['NIFTY 50|https://www.moneycontrol.com/indian-indices/nifty-50-9.html','SENSEX|https://www.moneycontrol.com/indian-indices/sensex-4.html']
        for line in configured[:5]:
            label,sep,url=line.partition('|')
            quote=_moneycontrol_quote(label.strip() or 'Market',url.strip()) if sep else None
            if quote: quote['tone']='market'; items.append(quote)
    return items



MODULES = {
    'agreements': 'Agreement Studio',
    'rooms': 'Room Status & Tenants',
    'reviews': 'Google Review Generator',
    'food': 'Food Delivery Hub',
    'rentok': 'Livenza Billing Suite',
    'banking': 'Banking & Reconciliation Suite',
    'electricity': 'Electricity Bill Studio',
    'queries': 'Live Queries Manager',
    'video_wall': 'Video Wall Studio',
    'whatsapp': 'WhatsApp Workspace',
    'email': 'Email Workspace',
    'drive': 'Google Drive Files',
    'integrations': 'Integrations Center',
    'letterhead': 'Livenza Letterhead Studio',
}


SYSTEM_SETTINGS_PANES = [
    {'key':'account','label':'Account','group':'Personal','icon':'account','description':'Your Livenza identity, sign-in status and personal workspace preferences.'},
    {'key':'network','label':'Network','group':'Connectivity','icon':'wifi','description':'Livenza connectivity, backend reachability and integration health.'},
    {'key':'focus','label':'Focus','group':'Personal','icon':'focus','description':'Reduce non-critical alerts, mascot interruptions and live-status distractions.'},
    {'key':'general','label':'General','group':'System','icon':'general','description':'Version, region, language and core Livenza defaults.'},
    {'key':'appearance','label':'Appearance','group':'System','icon':'appearance','description':'macOS 27 application material, light/dark appearance, contrast and transparency.'},
    {'key':'accessibility','label':'Accessibility','group':'System','icon':'accessibility','description':'Motion, transparency, text and interaction accessibility preferences.'},
    {'key':'control-centre','label':'Control Centre','group':'System','icon':'control-centre','description':'Choose the quick controls available in the Livenza toolbar.'},
    {'key':'desktop-dock','label':'Desktop & Dock','group':'System','icon':'desktop-dock','description':'Dock size, magnification and automatic hiding.'},
    {'key':'wallpaper','label':'Wallpaper','group':'System','icon':'wallpaper','description':'Choose a vibrant Tesla OS desktop wallpaper or use your own image.'},
    {'key':'widgets','label':'Widgets','group':'System','icon':'widgets','description':'Manage operational widgets from System Settings without adding clutter to Home.'},
    {'key':'privacy-security','label':'Privacy & Security','group':'Security','icon':'privacy','description':'Passkeys, kiosk controls, secure sessions and sensitive-data visibility.'},
    {'key':'users-groups','label':'Users & Groups','group':'Administration','icon':'users','description':'Users, roles, permissions, profile identity and access controls.','admin_only':True},
    {'key':'internet-accounts','label':'Internet Accounts','group':'Connectivity','icon':'integrations','description':'Connected services, providers, secrets and integration workflows.','permission':'integrations'},
    {'key':'intelligence','label':'Livenza Intelligence','group':'Personal','icon':'intelligence','description':'Companion behavior, operational reactions, weather and assistance preferences.'},
    {'key':'automations','label':'Automations','group':'Administration','icon':'automation','description':'Live-status, reporting, query and first-party automation settings.','admin_only':True},
    {'key':'organisation','label':'Organisation & Admin','group':'Administration','icon':'organisation','description':'Vault, cities, provider administration and organisation-wide controls.','admin_only':True},
]


def allowed_settings_panes(user=None):
    user = user or current_user()
    if not user:
        return []
    admin = (user.role or '').lower() == 'admin'
    allowed = []
    for pane in SYSTEM_SETTINGS_PANES:
        if pane.get('admin_only') and not admin:
            continue
        permission = pane.get('permission')
        if permission and not admin and not can_access(permission, user):
            continue
        allowed.append(dict(pane))
    return allowed


def default_settings_pane(user=None):
    panes = allowed_settings_panes(user)
    return panes[0]['key'] if panes else 'account'


def settings_pane_url(pane, **values):
    return url_for('system_settings_pane', pane=pane, **values)


def _system_settings_server_keys():
    return ('food_webhook_token','whatsapp_recipient','empty_report_time','default_google_review_url','vacant_report_enabled','vacant_report_time','vacant_report_recipients','query_webhook_token',
            'marquee_enabled','marquee_show_username','marquee_show_tenants','marquee_show_vacant_beds','marquee_show_earnings','marquee_show_favorites','marquee_show_stocks','marquee_favorites','marquee_custom_text','marquee_manual_earnings','marquee_stock_pages','marquee_refresh_seconds',
            'companion_enabled','companion_weather_enabled','companion_weather_effects','companion_quotes_enabled','companion_operations_enabled','companion_default_city','companion_effect_seconds',
            'host3d_default_intensity','host3d_max_intensity','host3d_default_city','host3d_operational_updates_default','host3d_weather_default','host3d_motivational_default')


def _system_settings_server_values():
    defaults={'marquee_enabled':'1','marquee_show_username':'1','marquee_show_tenants':'1','marquee_show_vacant_beds':'1','marquee_show_earnings':'1','marquee_refresh_seconds':'60',
              'companion_enabled':'1','companion_weather_enabled':'1','companion_weather_effects':'1','companion_quotes_enabled':'1','companion_operations_enabled':'1','companion_default_city':'Gurugram','companion_effect_seconds':'11',
              'host3d_default_intensity':'full','host3d_max_intensity':'full','host3d_default_city':'Gurugram','host3d_operational_updates_default':'1','host3d_weather_default':'1','host3d_motivational_default':'1'}
    return {k:setting(k,defaults.get(k,'')) for k in _system_settings_server_keys()}


def settings_pane_context(pane, user=None):
    user = user or current_user()
    ctx = {'user':user, 'settings':_system_settings_server_values()}
    if pane in ('account','intelligence'):
        ctx.update(avatar_ai_ready=bool(os.getenv('OPENAI_API_KEY','').strip()), mascot_preferences=mascot_preferences_for(user))
    if pane == 'privacy-security':
        ctx.update(credentials=WebAuthnCredential.query.filter_by(user_id=user.id).order_by(WebAuthnCredential.id).all() if user else [], kiosk_enabled=setting('kiosk_mode_enabled','0')=='1')
    if pane == 'internet-accounts':
        ctx.update(_integration_center_context((request.args.get('category') or '').strip().lower(),(request.args.get('provider') or '').strip(),(request.args.get('workflow') or '').strip()))
    if pane == 'users-groups':
        users=User.query.order_by(User.username).all()
        ctx.update(users=users, credentials={u.id:WebAuthnCredential.query.filter_by(user_id=u.id).order_by(WebAuthnCredential.id).all() for u in users},
                   cities=City.query.order_by(City.name).all(), modules=MODULES, letterhead_capabilities=LETTERHEAD_CAPABILITIES,
                   query_templates=QueryTemplate.query.order_by(QueryTemplate.id.desc()).all(), aadhaar_provider_configured=bool(os.getenv('AADHAAR_AUTH_URL','').strip()),
                   google_oauth_ready=google_oauth_configured(), google_is_connected=google_connected(), drive_folder_id=setting('google_drive_folder_id',''),
                   drive_auto_backup=setting('google_drive_auto_backup','0')=='1', kiosk_enabled=setting('kiosk_mode_enabled','0')=='1', avatar_ai_ready=bool(os.getenv('OPENAI_API_KEY','').strip()))
    if pane == 'organisation':
        edit_secret_id=(request.args.get('edit_secret') or '').strip(); edit_provider_id=(request.args.get('edit_provider') or '').strip()
        ctx.update(entries=VaultSecret.query.order_by(VaultSecret.id.desc()).all(), audits=AuditEvent.query.filter(AuditEvent.module.in_(['vault','electricity'])).order_by(AuditEvent.id.desc()).limit(150).all(),
                   providers=_electricity_provider_rows(include_inactive=True), connections=ElectricityConnection.query.order_by(ElectricityConnection.property_name).all(),
                   allowed_secret_types=sorted(ALLOWED_SECRET_TYPES), vault_ready=bool(os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()),
                   edit_secret=db.session.get(VaultSecret,int(edit_secret_id)) if edit_secret_id.isdigit() else None,
                   edit_provider=db.session.get(ElectricityProvider,int(edit_provider_id)) if edit_provider_id.isdigit() else None)
    return ctx


LETTERHEAD_CAPABILITIES = {
    'letterhead_use': 'Use Letterhead Studio',
    'letterhead_ai': 'Use Ask Livenza AI',
    'letterhead_template_author': 'Create and edit letterhead template drafts',
    'letterhead_template_submit': 'Submit letterhead templates for approval',
    'letterhead_signature_use': 'Use protected signatures and seals',
    'letterhead_email_send': 'Send finalized documents by email',
    'letterhead_whatsapp_send': 'Send finalized documents by WhatsApp',
    'letterhead_vault_all': 'View all authorized Letterhead Document Vault records',
}

BASE_REQUIRED_AGREEMENT_FIELDS = []

AGREEMENT_GROUPS = [
    ('Agreement Format, Execution & Stamp Details', [x[0] for x in FIELDS[0:22]]),
    ('Landlord / Lessor / Management', [x[0] for x in FIELDS[22:32]]),
    ('Tenant / Lessee / Licensee / Guest', [x[0] for x in FIELDS[32:43]]),
    ('Corporate Client / Sponsor', [x[0] for x in FIELDS[43:51]]),
    ('Property, City & Financial Terms', ['city'] + [x[0] for x in FIELDS[51:75]]),
    ('Operations, Access, Services & Special Terms', [x[0] for x in FIELDS[75:88]]),
    ('Operating Model & Foreign Client Compliance', [x[0] for x in FIELDS[88:110]]),
    ('Licences, Verification & Regulatory References', [x[0] for x in FIELDS[110:120]]),
    ('Witnesses & Execution', [x[0] for x in FIELDS[120:122]]),
]

AGREEMENT_PARTY_PROFILE_FIELDS = {
    'landlord':[x[0] for x in FIELDS[22:32]],
    'tenant':[x[0] for x in FIELDS[32:43]],
}


def user_permissions(user=None):
    user = user or current_user()
    if not user:
        return set()
    if (user.role or '').lower() == 'admin':
        return set(MODULES)
    try:
        vals = json.loads(user.permissions_json or '[]')
        return {x for x in vals if x in MODULES}
    except Exception:
        return set()


def user_capabilities(user=None):
    user = user or current_user()
    if not user:
        return set()
    if (user.role or '').lower() == 'admin':
        return set(LETTERHEAD_CAPABILITIES)
    try:
        vals = json.loads(getattr(user, 'capabilities_json', '') or '[]')
        allowed = {x for x in vals if x in LETTERHEAD_CAPABILITIES}
    except Exception:
        allowed = set()
    if can_access('letterhead', user):
        allowed.add('letterhead_use')
    return allowed


def has_capability(capability, user=None):
    return capability in user_capabilities(user)


def capability_required(capability):
    def outer(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get('uid'):
                return redirect(url_for('login', next=request.path))
            user = current_user()
            if not user or not user.active:
                session.clear()
                return redirect(url_for('login'))
            if not has_capability(capability, user):
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return outer


def can_access(module, user=None):
    if module not in MODULES:
        return False
    return module in user_permissions(user)


def permission_required(module):
    def outer(fn):
        @wraps(fn)
        def wrapper(*a, **kw):
            if not session.get('uid'):
                return redirect(url_for('login', next=request.path))
            u = current_user()
            if not u or not u.active:
                session.clear()
                return redirect(url_for('login'))
            if not can_access(module, u):
                flash(f'Your account does not have access to {MODULES[module]}.', 'danger')
                return redirect(url_for('dashboard'))
            return fn(*a, **kw)
        return wrapper
    return outer


def admin_required(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        if not session.get('uid'):
            return redirect(url_for('login', next=request.path))
        u=current_user()
        if not u or not u.active or (u.role or '').lower()!='admin':
            abort(403)
        return fn(*a, **kw)
    return wrapper


def agreement_required_fields(preset_name, data):
    # Tesla OS 27: every Agreement Studio field is optional.
    return set()


def field_label_map():
    m={x[0]:x[1] for x in FIELDS}
    m['city']='City'
    return m


def missing_agreement_fields(preset_name, data):
    # Compatibility hook: no Agreement Studio fields are mandatory.
    return []


def normalize_whatsapp_number(value):
    raw=''.join(ch for ch in (value or '') if ch.isdigit())
    if len(raw)==10:
        raw='91'+raw
    return raw if 8 <= len(raw) <= 15 else ''


def share_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt='livenza-agreement-share-v1')

def _upload_image_bytes(file_storage, max_bytes=12*1024*1024):
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return b'', 'Choose a JPG, PNG, WebP, HEIC/HEIF, TIFF or BMP profile photo.'
    try:
        file_storage.stream.seek(0)
        raw=file_storage.stream.read(max_bytes+1)
        file_storage.stream.seek(0)
    except Exception:
        return b'', 'The selected photo could not be read.'
    if not raw:
        return b'', 'The selected photo is empty.'
    if len(raw)>max_bytes:
        return b'', 'The profile photo must be smaller than 12 MB.'
    try:
        probe=PILImage.open(io.BytesIO(raw)); probe.verify()
    except Exception:
        return b'', 'Use a valid JPG, PNG, WebP, HEIC/HEIF, TIFF or BMP profile photo.'
    return raw, ''

def _open_avatar_source(raw, max_side=1800):
    image=ImageOps.exif_transpose(PILImage.open(io.BytesIO(raw))).convert('RGB')
    image.thumbnail((max_side,max_side),PILImage.Resampling.LANCZOS)
    # Gentle adaptive recovery helps low-light, soft and phone-camera images without overprocessing identity.
    try:
        luminance=float(ImageStat.Stat(ImageOps.grayscale(image)).mean[0])
        if luminance < 72: image=ImageEnhance.Brightness(image).enhance(1.28)
        elif luminance < 105: image=ImageEnhance.Brightness(image).enhance(1.12)
        image=ImageEnhance.Contrast(image).enhance(1.045)
        image=ImageEnhance.Sharpness(image).enhance(1.10)
    except Exception:
        pass
    return image

def _preserved_square(image, size=768):
    # Preserve the whole source first instead of forcing a centred face crop.
    bg=ImageOps.fit(image,(size,size),method=PILImage.Resampling.LANCZOS,centering=(.5,.5)).filter(ImageFilter.GaussianBlur(max(8,size//28)))
    bg=ImageEnhance.Brightness(bg).enhance(.90)
    fg=ImageOps.contain(image,(int(size*.90),int(size*.90)),method=PILImage.Resampling.LANCZOS)
    x=(size-fg.width)//2; y=(size-fg.height)//2
    bg.paste(fg,(x,y))
    return bg

def _portrait_image(raw, size=768):
    image=_open_avatar_source(raw,max_side=max(1600,size*2))
    return _preserved_square(image,size)

def _avatar_reference_board(raw, size=1024):
    image=_open_avatar_source(raw,max_side=2200)
    canvas=PILImage.new('RGB',(size,size),(239,246,251))
    gutter=max(8,size//96); left=int(size*.66); right=size-left-gutter
    main=_preserved_square(image,left)
    canvas.paste(main,(0,(size-left)//2))
    slot_h=(size-gutter*2)//3
    for i,cx in enumerate((.30,.50,.70)):
        crop=ImageOps.fit(image,(right,slot_h),method=PILImage.Resampling.LANCZOS,centering=(cx,.42))
        canvas.paste(crop,(left+gutter,i*(slot_h+gutter)))
    return canvas

def _jpeg_data_uri(image, max_size=768, quality=88):
    image=image.convert('RGB')
    image.thumbnail((max_size,max_size),PILImage.Resampling.LANCZOS)
    buf=io.BytesIO(); image.save(buf,format='JPEG',quality=quality,optimize=True,progressive=True)
    if buf.tell()>950000:
        image.thumbnail((640,640),PILImage.Resampling.LANCZOS)
        buf=io.BytesIO(); image.save(buf,format='JPEG',quality=80,optimize=True,progressive=True)
    return 'data:image/jpeg;base64,'+base64.b64encode(buf.getvalue()).decode('ascii')

def profile_photo_data_uri_from_bytes(raw):
    try: return _jpeg_data_uri(_portrait_image(raw,640),640,86)
    except Exception: return ''

def polished_avatar_data_uri_from_bytes(raw):
    """Reliable private fallback: a clean Livenza-toned portrait with no cloud dependency."""
    try:
        portrait=_portrait_image(raw,768)
        portrait=ImageEnhance.Contrast(portrait).enhance(1.035)
        portrait=ImageEnhance.Color(portrait).enhance(1.045)
        portrait=ImageEnhance.Sharpness(portrait).enhance(1.08)
        background=PILImage.new('RGB',(768,768),(244,249,253))
        for y in range(768):
            t=y/767
            color=(int(248-16*t),int(251-12*t),int(253-4*t))
            background.paste(color,(0,y,768,y+1))
        # A restrained blue glass grade makes mixed source photos feel consistent.
        graded=PILImage.blend(portrait,background,.075)
        return _jpeg_data_uri(graded,768,89)
    except Exception:
        return ''

def image_data_uri(file_storage):
    raw,error=_upload_image_bytes(file_storage)
    return '' if error else profile_photo_data_uri_from_bytes(raw)

def _data_uri_bytes(value):
    try:
        encoded=(value or '').split(',',1)[1]
        return base64.b64decode(encoded,validate=True)
    except Exception:
        return b''

def _ai_avatar_data_uri(raw):
    key=os.getenv('OPENAI_API_KEY','').strip()
    if not key:
        return '', 'AI image service is not configured.'
    try:
        from openai import OpenAI
        source=_avatar_reference_board(raw,1024)
        source_buf=io.BytesIO(); source.save(source_buf,format='PNG',optimize=True); source_buf.seek(0)
        source_buf.name='livenza-avatar-reference-board.png'
        result=OpenAI(api_key=key,timeout=45).images.edit(
            model=os.getenv('OPENAI_AVATAR_MODEL','gpt-image-2'),
            image=source_buf,
            prompt=(
                'Create a premium animated Livenza male mascot based on the same person shown across this reference board. This must be '
                'a distinct character mascot, not a polished copy of the photo. The source may be a side profile, three-quarter angle, '
                'candid, full-body, off-centre, low-light, slightly soft, partially obstructed or not front-facing; use all visible '
                'references together to recover identity conservatively. Preserve recognizable facial geometry, skin tone, hairstyle, '
                'eyewear, approximate age and distinctive visible traits faithfully, but reinterpret the person as a clean 3D/semi-realistic '
                'animated character. Show a friendly three-quarter or full-body standing pose with natural proportions, a confident welcoming '
                'expression, and premium Livenza styling. Dress the mascot in a refined smart-casual Livenza outfit using white, navy and '
                'subtle ice-blue accents. Use a clean soft background or clean cutout feeling suitable for placing across a dashboard. '
                'Sophisticated rather than childish, charming rather than robotic. No text, no logo, no border, no extra people, no copied '
                'photo framing, no passport, and no realistic photography look.'
            ),
            size='1024x1024',
            quality=os.getenv('OPENAI_AVATAR_QUALITY','high'),
        )
        item=(getattr(result,'data',None) or [None])[0]
        encoded=getattr(item,'b64_json',None) if item else None
        generated=base64.b64decode(encoded) if encoded else b''
        if not generated and item and getattr(item,'url',None):
            response=requests.get(item.url,timeout=45); response.raise_for_status(); generated=response.content
        if not generated: return '', 'The AI image service returned no avatar.'
        return _jpeg_data_uri(_portrait_image(generated,768),768,89), ''
    except Exception as exc:
        app.logger.warning('Live mascot generation failed; default mascot will remain active: %s',str(exc)[:180])
        return '', 'AI mascot styling was unavailable.'

def create_live_avatar(raw, prefer_ai=True):
    if prefer_ai:
        generated,error=_ai_avatar_data_uri(raw)
        if generated: return generated,'ai_mascot','AI live mascot created and applied across the workspace.'
        return '','default','AI mascot generation was unavailable. The default Livenza mascot remains active.'
    return '','default','The default Livenza mascot remains active.'

def masked_aadhaar(last4):
    d=''.join(ch for ch in (last4 or '') if ch.isdigit())[-4:]
    return f'XXXX XXXX {d}' if len(d)==4 else ''

def _extract_json_object(text_value):
    raw=(text_value or '').strip()
    raw=re.sub(r'^```(?:json)?\s*|\s*```$','',raw,flags=re.I|re.S).strip()
    try:
        value=json.loads(raw)
        return value if isinstance(value,dict) else {}
    except Exception:
        m=re.search(r'\{.*\}',raw,re.S)
        if not m: return {}
        try:
            value=json.loads(m.group(0)); return value if isinstance(value,dict) else {}
        except Exception:
            return {}


def _normalize_aadhaar_extract(value):
    value=value if isinstance(value,dict) else {}
    out={}
    def clean(k,limit=800):
        return str(value.get(k,'') or '').strip()[:limit]
    out['tenant_name']=clean('name',180)
    out['tenant_father']=clean('father_or_spouse',180)
    out['tenant_dob']=clean('date_of_birth',40)
    out['tenant_address']=clean('address',1000)
    digits=''.join(ch for ch in clean('aadhaar_number',40) if ch.isdigit())
    if len(digits)==12:
        out['tenant_id_no']=f'{digits[:4]} {digits[4:8]} {digits[8:]}'
    else:
        out['tenant_id_no']=''
    out['tenant_id_type']='Aadhaar'
    out['gender']=clean('gender',30)
    return out


def _parse_aadhaar_text_fallback(raw_text):
    txt='\n'.join(x.strip() for x in (raw_text or '').splitlines() if x.strip())
    result={'name':'','father_or_spouse':'','date_of_birth':'','gender':'','address':'','aadhaar_number':''}
    m=re.search(r'(?<!\d)(\d{4})[\s\-]*(\d{4})[\s\-]*(\d{4})(?!\d)',txt)
    if m: result['aadhaar_number']=' '.join(m.groups())
    m=re.search(r'(?:DOB|Date\s*of\s*Birth|Year\s*of\s*Birth|YOB)\s*[:\-/]?\s*([0-3]?\d[\-/][01]?\d[\-/]\d{4}|\d{4})',txt,re.I)
    if m: result['date_of_birth']=m.group(1)
    m=re.search(r'\b(Male|Female|Transgender)\b',txt,re.I)
    if m: result['gender']=m.group(1).title()
    m=re.search(r'(?:S\s*/\s*[O0]|D\s*/\s*[O0]|W\s*/\s*[O0]|C\s*/\s*[O0])\s*[:\-]?\s*([^\n,]+)',txt,re.I)
    if m: result['father_or_spouse']=m.group(1).strip()
    m=re.search(r'(?:Address|पता)\s*[:\-]?\s*(.+?)(?=(?:\n\s*\d{4}\s*\d{4}\s*\d{4}|\Z))',txt,re.I|re.S)
    if m: result['address']=' '.join(m.group(1).split())[:1000]
    # Name heuristic: a short human-name line immediately preceding DOB/gender.
    lines=[x.strip() for x in txt.splitlines() if x.strip()]
    name_anchor=next((i for i,x in enumerate(lines) if re.search(r'\b(?:DOB|Date\s*of\s*Birth|Year\s*of\s*Birth|YOB)\b',x,re.I)),None)
    if name_anchor is None:name_anchor=next((i for i,x in enumerate(lines) if re.search(r'\b(?:Male|Female|Transgender)\b',x,re.I)),None)
    if name_anchor is not None:
        for cand in reversed(lines[max(0,name_anchor-5):name_anchor]):
            if 2 <= len(cand) <= 80 and not re.search(r'Government|भारत|UIDAI|Aadhaar|Enrollment|VID|Male|Female|Address',cand,re.I) and not re.search(r'\d{4}',cand):
                result['name']=cand; break
    return _normalize_aadhaar_extract(result)


_AADHAAR_OCR_ENGINE=None
_AADHAAR_OCR_LOCK=threading.Lock()


def _aadhaar_image_pages(file_bytes, filename, mimetype):
    """Return small bounded RGB pages without writing identity data to disk.

    Keeping at most two pages around 2.3 megapixels prevents the OCR model and
    a multi-page PDF raster from competing for memory on small cloud workers.
    """
    is_pdf=mimetype=='application/pdf' or (filename or '').lower().endswith('.pdf')
    if is_pdf:
        try:
            try:
                import pymupdf as fitz
            except ImportError:
                import fitz
            pages=[]
            with fitz.open(stream=file_bytes,filetype='pdf') as document:
                for page_number in range(min(2,document.page_count)):
                    page=document.load_page(page_number)
                    area=max(1,float(page.rect.width*page.rect.height))
                    scale=min(1.8,max(.35,(2_300_000/area) ** .5))
                    pixmap=page.get_pixmap(matrix=fitz.Matrix(scale,scale),alpha=False)
                    pages.append(PILImage.open(io.BytesIO(pixmap.tobytes('png'))).convert('RGB'))
            return pages
        except Exception:
            return []
    try:
        source=PILImage.open(io.BytesIO(file_bytes))
        if source.width*source.height>24_000_000:
            return []
        image=ImageOps.exif_transpose(source).convert('RGB')
        image.thumbnail((1900,1900),PILImage.Resampling.LANCZOS)
        return [image]
    except Exception:
        return []


def _aadhaar_reading_image(image):
    """Increase local contrast without retaining or exporting the document."""
    gray=ImageOps.grayscale(image)
    gray=ImageOps.autocontrast(gray,cutoff=1)
    if max(gray.size)<1300:
        scale=min(2.0,1300/max(1,max(gray.size)))
        gray=gray.resize((max(1,int(gray.width*scale)),max(1,int(gray.height*scale))),PILImage.Resampling.LANCZOS)
    return gray.convert('RGB')


def _aadhaar_tesseract_text(image):
    binary=shutil.which('tesseract')
    if not binary:
        return ''
    try:
        payload=io.BytesIO(); _aadhaar_reading_image(image).save(payload,format='PNG',optimize=True)
        run=subprocess.run(
            [binary,'stdin','stdout','-l','eng','--psm','6','-c','preserve_interword_spaces=1'],input=payload.getvalue(),
            stdout=subprocess.PIPE,stderr=subprocess.DEVNULL,timeout=18,check=False,
        )
        return run.stdout.decode('utf-8','ignore') if run.returncode==0 else ''
    except Exception:
        return ''


def _aadhaar_local_ocr_text(file_bytes, filename, mimetype):
    """Run bounded server-local OCR without a browser/device dependency."""
    pages=_aadhaar_image_pages(file_bytes,filename,mimetype)
    if not pages:
        return '', 'The document could not be rendered for automatic reading.'
    # The small native binary is the lowest-memory path and is installed by the
    # supplied Docker build. It also acts as a stable fallback if ONNX cannot
    # initialize on a constrained host.
    tesseract_text='\n'.join(filter(None,(_aadhaar_tesseract_text(image) for image in pages)))
    tesseract_markers=sum(bool(re.search(pattern,tesseract_text,re.I)) for pattern in (
        r'\d{4}[\s\-]*\d{4}[\s\-]*\d{4}',r'\b(?:DOB|YOB|Date\s*of\s*Birth)\b',r'\b(?:Male|Female)\b',r'\bAddress\b'))
    if tesseract_text.strip() and tesseract_markers>=2:
        return tesseract_text, ''
    rapid_error=''
    try:
        global _AADHAAR_OCR_ENGINE
        import numpy as np
        lines=[]
        with _AADHAAR_OCR_LOCK:
            if _AADHAAR_OCR_ENGINE is None:
                from rapidocr import RapidOCR
                _AADHAAR_OCR_ENGINE=RapidOCR()
            for image in pages:
                result=_AADHAAR_OCR_ENGINE(np.asarray(_aadhaar_reading_image(image)),use_det=True,use_cls=True,use_rec=True)
                texts=list(getattr(result,'txts',()) or ())
                scores=list(getattr(result,'scores',()) or ())
                lines.extend(text for index,text in enumerate(texts) if text and (index>=len(scores) or float(scores[index])>=.42))
        if lines:
            combined='\n'.join(lines)
            if tesseract_text.strip(): combined+='\n'+tesseract_text
            return combined, ''
    except Exception as exc:
        rapid_error=str(exc)[:120]
    if tesseract_text.strip():
        return tesseract_text, ''
    return '', ('The server OCR engine could not initialize.' if rapid_error else 'No readable text was detected.')


def _aadhaar_ai_extract(file_bytes, filename, mimetype):
    key=os.getenv('OPENAI_API_KEY','').strip()
    if not key:
        return {}, 'AI enhancement is unavailable.'
    from openai import OpenAI
    client=OpenAI(api_key=key)
    prompt=(
        'Extract identity fields from this Indian Aadhaar document for tenant-form autofill. '
        'Return ONLY strict JSON with keys: name, father_or_spouse, date_of_birth, gender, address, aadhaar_number. '
        'Use empty strings when a field is not clearly visible. Do not guess. Preserve the Aadhaar number only if all 12 digits are clearly legible. '
        'Do not add commentary. This extraction is not an authenticity verification.'
    )
    content=[{'type':'input_text','text':prompt}]
    if mimetype=='application/pdf' or filename.lower().endswith('.pdf'):
        content.append({'type':'input_file','filename':filename or 'aadhaar.pdf','file_data':base64.b64encode(file_bytes).decode('ascii')})
    else:
        mt=mimetype if mimetype in ('image/jpeg','image/png','image/webp') else 'image/jpeg'
        content.append({'type':'input_image','image_url':f'data:{mt};base64,'+base64.b64encode(file_bytes).decode('ascii'),'detail':'high'})
    resp=client.responses.create(
        model=os.getenv('OPENAI_AADHAAR_MODEL',os.getenv('OPENAI_REVIEW_MODEL','gpt-5.6-luna')),
        input=[{'role':'user','content':content}],
    )
    obj=_extract_json_object(getattr(resp,'output_text',''))
    return _normalize_aadhaar_extract(obj), '' if obj else 'The document could not be read reliably.'

def query_log(q, action, details='', actor=None):
    db.session.add(QueryActivity(query_id=q.id,action=action,details=details,actor_user_id=(actor.id if actor else None)))

def wa_number(value):
    return normalize_whatsapp_number(value)

def whatsapp_cloud_configured():
    return bool(os.getenv('WHATSAPP_CLOUD_TOKEN','').strip() and os.getenv('WHATSAPP_PHONE_NUMBER_ID','').strip())

def whatsapp_cloud_text(to, body):
    to=wa_number(to)
    token=os.getenv('WHATSAPP_CLOUD_TOKEN','').strip(); pid=os.getenv('WHATSAPP_PHONE_NUMBER_ID','').strip()
    if not (to and token and pid): return False, 'WhatsApp Cloud API is not configured.'
    ver=os.getenv('WHATSAPP_GRAPH_VERSION','v23.0')
    r=requests.post(f'https://graph.facebook.com/{ver}/{pid}/messages',headers={'Authorization':f'Bearer {token}','Content-Type':'application/json'},json={'messaging_product':'whatsapp','to':to,'type':'text','text':{'body':body,'preview_url':False}},timeout=25)
    if not r.ok: return False,r.text[:800]
    try: return True,str((r.json().get('messages') or [{}])[0].get('id') or 'Sent')
    except Exception: return True,'Sent'

def whatsapp_cloud_template(to, template_name, language='en'):
    to=wa_number(to); token=os.getenv('WHATSAPP_CLOUD_TOKEN','').strip(); pid=os.getenv('WHATSAPP_PHONE_NUMBER_ID','').strip()
    if not (to and token and pid and template_name): return False, 'Cloud API/template not configured.'
    ver=os.getenv('WHATSAPP_GRAPH_VERSION','v23.0')
    payload={'messaging_product':'whatsapp','to':to,'type':'template','template':{'name':template_name,'language':{'code':language}}}
    r=requests.post(f'https://graph.facebook.com/{ver}/{pid}/messages',headers={'Authorization':f'Bearer {token}','Content-Type':'application/json'},json=payload,timeout=25)
    return r.ok, (r.text[:800] if not r.ok else 'Sent')

def generate_empty_rooms_pdf_bytes():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf=io.BytesIO(); doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=36,rightMargin=36,topMargin=42,bottomMargin=42)
    styles=getSampleStyleSheet(); story=[Paragraph('LIVENZA LIFE - VACANT ROOMS STATUS',styles['Title']),Paragraph(datetime.datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%d %b %Y, %I:%M %p IST'),styles['Normal']),Spacer(1,12)]
    rows=[['City','Property','Room','Type','Tariff','Status']]
    for r in Room.query.order_by(Room.city,Room.property_name,Room.room_no):
        st=room_status(r)
        if st=='Vacant': rows.append([r.city,r.property_name,r.room_no,r.room_type,r.standard_tariff,st])
    if len(rows)==1: rows.append(['-','-','-','-','-','No vacant rooms'])
    t=Table(rows,repeatRows=1,colWidths=[65,110,45,85,70,65]); t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#5c34d6')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.4,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'TOP')]))
    story.append(t); doc.build(story); buf.seek(0); return buf

def whatsapp_upload_pdf(pdf_bytes, filename):
    token=os.getenv('WHATSAPP_CLOUD_TOKEN','').strip(); pid=os.getenv('WHATSAPP_PHONE_NUMBER_ID','').strip(); ver=os.getenv('WHATSAPP_GRAPH_VERSION','v23.0')
    if not (token and pid): return None
    r=requests.post(f'https://graph.facebook.com/{ver}/{pid}/media',headers={'Authorization':f'Bearer {token}'},data={'messaging_product':'whatsapp'},files={'file':(filename,pdf_bytes.getvalue(),'application/pdf')},timeout=30)
    return (r.json().get('id') if r.ok else None)

def whatsapp_send_document(to, media_id, filename, caption=''):
    token=os.getenv('WHATSAPP_CLOUD_TOKEN','').strip(); pid=os.getenv('WHATSAPP_PHONE_NUMBER_ID','').strip(); ver=os.getenv('WHATSAPP_GRAPH_VERSION','v23.0'); to=wa_number(to)
    if not (token and pid and to and media_id): return False
    payload={'messaging_product':'whatsapp','to':to,'type':'document','document':{'id':media_id,'filename':filename,'caption':caption[:1024]}}
    return requests.post(f'https://graph.facebook.com/{ver}/{pid}/messages',headers={'Authorization':f'Bearer {token}','Content-Type':'application/json'},json=payload,timeout=25).ok

def normalize_query_payload(source, payload):
    flat={}
    if isinstance(payload,dict):
        flat.update(payload)
        for row in payload.get('user_column_data',[]) or []:
            key=(row.get('column_name') or row.get('column_id') or '').lower(); val=row.get('string_value') or row.get('column_value') or ''
            if key: flat[key]=val
        for row in payload.get('field_data',[]) or []:
            key=(row.get('name') or '').lower(); vals=row.get('values') or []; flat[key]=(vals[0] if vals else '')
    def pick(*keys):
        for k in keys:
            v=flat.get(k)
            if v not in (None,''): return str(v)
        return ''
    return dict(source=source.title(),external_id=pick('lead_id','id','external_id','leadgen_id'),customer_name=pick('full_name','name','customer_name'),mobile=pick('phone_number','phone','mobile'),whatsapp=pick('whatsapp','phone_number','phone','mobile'),email=pick('email','email_address'),query_text=pick('query','message','comments','notes','property_query'),city=pick('city'),property_name=pick('property','property_name','listing_name'),budget=pick('budget','travel_budget'),move_in_date=pick('move_in_date','arrival_date','check_in'),stay_type=pick('stay_type','accommodation_type'))

def auto_reply_for_query(q):
    candidates=QueryTemplate.query.filter_by(active=True,auto_send=True).order_by(QueryTemplate.id).all()
    for t in candidates:
        try: sources=json.loads(t.sources_json or '[]'); statuses=json.loads(t.statuses_json or '[]')
        except Exception: sources=statuses=[]
        if sources and q.source not in sources: continue
        if statuses and q.status not in statuses: continue
        number=q.whatsapp or q.mobile
        if t.whatsapp_template_name and whatsapp_cloud_configured(): ok,msg=whatsapp_cloud_template(number,t.whatsapp_template_name)
        elif whatsapp_cloud_configured(): ok,msg=whatsapp_cloud_text(number,t.message.format(name=q.customer_name or 'Guest',property=q.property_name or 'Livenza Life',city=q.city or ''))
        else: return False,'WhatsApp Cloud API not configured; query retained for manual reply.'
        query_log(q,'Auto WhatsApp',f'{t.name}: {msg}'); return ok,msg
    return False,'No matching auto-reply template.'


def _supabase_project_ref():
    explicit=os.getenv('SUPABASE_PROJECT_REF','').strip()
    if explicit: return explicit
    m=re.search(r'postgres\.([a-z0-9]+)[:@]', raw_db or '', re.I)
    if m: return m.group(1)
    m=re.search(r'db\.([a-z0-9]+)\.supabase\.co', raw_db or '', re.I)
    return m.group(1) if m else ''

def supabase_storage_configured():
    return bool(_supabase_project_ref() and os.getenv('SUPABASE_SERVICE_ROLE_KEY','').strip())

VIDEO_WALL_BUCKET='video-wall-media'
VIDEO_WALL_MIME_BY_EXTENSION={
    '.mp4':'video/mp4','.m4v':'video/mp4','.webm':'video/webm','.mov':'video/quicktime',
    '.jpg':'image/jpeg','.jpeg':'image/jpeg','.png':'image/png','.webp':'image/webp',
}


def video_wall_upload_limit_mb():
    try: return max(10,min(512000,int(os.getenv('VIDEO_WALL_MAX_MB','2048'))))
    except Exception: return 2048


def _video_wall_storage_headers(json_content=False):
    key=os.getenv('SUPABASE_SERVICE_ROLE_KEY','').strip()
    headers={'Authorization':f'Bearer {key}','apikey':key}
    if json_content: headers['Content-Type']='application/json'
    return headers


def _ensure_video_wall_bucket():
    """Create the public playback bucket when it is missing."""
    project_ref=_supabase_project_ref()
    if not (project_ref and os.getenv('SUPABASE_SERVICE_ROLE_KEY','').strip()):
        return False,'Supabase media storage is not configured.'
    base=f'https://{project_ref}.supabase.co/storage/v1'
    headers=_video_wall_storage_headers(json_content=True)
    try:
        check=requests.get(f'{base}/bucket/{VIDEO_WALL_BUCKET}',headers=headers,timeout=20)
        if check.ok:
            info=check.json() if check.headers.get('content-type','').startswith('application/json') else {}
            if info.get('public') is False:
                return False,f'The {VIDEO_WALL_BUCKET} bucket exists but is private. Mark it Public in Supabase Storage so TV players can stream its media.'
            return True,''
        if check.status_code!=404:
            return False,f'Storage bucket check failed ({check.status_code}): {check.text[:240]}'
        created=requests.post(f'{base}/bucket',headers=headers,json={
            'id':VIDEO_WALL_BUCKET,'name':VIDEO_WALL_BUCKET,'public':True,
            'allowed_mime_types':sorted(set(VIDEO_WALL_MIME_BY_EXTENSION.values())),
        },timeout=25)
        if not created.ok and created.status_code!=409:
            return False,f'Storage bucket could not be created ({created.status_code}): {created.text[:240]}'
        return True,''
    except Exception as exc:
        return False,f'Storage bucket is unreachable: {str(exc)[:180]}'


def _video_wall_media_spec(filename,mime_type,size):
    clean_name=os.path.basename(str(filename or '')).strip()
    ext=os.path.splitext(clean_name.lower())[1]
    expected=VIDEO_WALL_MIME_BY_EXTENSION.get(ext)
    mime=(mime_type or '').lower().split(';')[0].strip()
    if not expected:
        return None,'Supported files: MP4, M4V, WebM, MOV, JPG, PNG and WebP.'
    if mime in ('','application/octet-stream','video/x-m4v'): mime=expected
    if mime not in set(VIDEO_WALL_MIME_BY_EXTENSION.values()):
        return None,'The selected file type does not match a supported video or image format.'
    try: size=int(size or 0)
    except Exception: size=0
    if size<=0: return None,'The selected media file is empty.'
    limit=video_wall_upload_limit_mb()*1024*1024
    if size>limit:
        return None,f'File exceeds the configured {video_wall_upload_limit_mb()} MB Video Wall limit.'
    return {'filename':clean_name or f'media{ext}','extension':ext,'mime':mime,'size':size,'type':('image' if mime.startswith('image/') else 'video')},''


def _video_wall_upload_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'],salt='livenza-video-wall-resumable-v1')


def create_video_wall_resumable_upload(filename,mime_type,size,title=''):
    spec,error=_video_wall_media_spec(filename,mime_type,size)
    if error: return None,error
    ready,error=_ensure_video_wall_bucket()
    if not ready: return None,error
    project_ref=_supabase_project_ref(); key=os.getenv('SUPABASE_SERVICE_ROLE_KEY','').strip()
    path=f"video-wall/{datetime.datetime.utcnow():%Y/%m}/{uuid.uuid4().hex}{spec['extension']}"
    sign_url=f"https://{project_ref}.supabase.co/storage/v1/object/upload/sign/{VIDEO_WALL_BUCKET}/{urllib.parse.quote(path,safe='/')}"
    try:
        signed=requests.post(sign_url,headers=_video_wall_storage_headers(json_content=True),json={},timeout=25)
        if not signed.ok:
            return None,f'Resumable upload could not start ({signed.status_code}): {signed.text[:260]}'
        signed_data=signed.json(); token=str(signed_data.get('token') or '')
        if not token:
            returned=str(signed_data.get('url') or signed_data.get('signedURL') or signed_data.get('signedUrl') or '')
            token=urllib.parse.parse_qs(urllib.parse.urlparse(returned).query).get('token',[''])[0]
        if not token: return None,'Storage did not return a signed upload token.'
        public=f"https://{project_ref}.supabase.co/storage/v1/object/public/{VIDEO_WALL_BUCKET}/{urllib.parse.quote(path,safe='/')}"
        reservation=_video_wall_upload_serializer().dumps({
            'path':path,'title':(title or spec['filename'])[:220],'mime':spec['mime'],'size':spec['size'],
            'type':spec['type'],'uid':current_user().id,
        })
        return {
            'endpoint':f'https://{project_ref}.storage.supabase.co/storage/v1/upload/resumable',
            'signature':token,'bucket':VIDEO_WALL_BUCKET,'object_name':path,'content_type':spec['mime'],
            'chunk_size':6*1024*1024,'reservation':reservation,'public_url':public,
        },''
    except Exception as exc:
        return None,f'Resumable upload could not start: {str(exc)[:180]}'


def finalize_video_wall_resumable_upload(reservation):
    try: data=_video_wall_upload_serializer().loads(str(reservation or ''),max_age=2*60*60)
    except Exception: return None,'The upload reservation expired. Select the file and try again.'
    if int(data.get('uid') or 0)!=int(current_user().id): return None,'This upload reservation belongs to another session.'
    path=str(data.get('path') or '')
    if not path.startswith('video-wall/'): return None,'Invalid media path.'
    project_ref=_supabase_project_ref()
    public=f"https://{project_ref}.supabase.co/storage/v1/object/public/{VIDEO_WALL_BUCKET}/{urllib.parse.quote(path,safe='/')}"
    object_url=f"https://{project_ref}.supabase.co/storage/v1/object/{VIDEO_WALL_BUCKET}/{urllib.parse.quote(path,safe='/')}"
    verified=None
    for attempt in range(3):
        try:
            verified=requests.head(object_url,headers=_video_wall_storage_headers(),timeout=25,allow_redirects=True)
            if verified.ok: break
        except Exception: verified=None
        if attempt<2: time.sleep(.25)
    if not verified or not verified.ok:
        code=verified.status_code if verified is not None else 'network'
        return None,f'Storage has not confirmed the completed media yet ({code}). Retry Finish Upload.'
    existing=VideoAsset.query.filter_by(storage_path=path).first()
    if existing: return existing,''
    actual_size=int(verified.headers.get('Content-Length') or data.get('size') or 0)
    asset=VideoAsset(title=str(data.get('title') or 'Video Wall media')[:220],media_type=data.get('type') or 'video',storage_path=path,public_url=public,mime_type=data.get('mime') or '',file_size=actual_size,uploaded_by_user_id=current_user().id)
    db.session.add(asset);db.session.commit();return asset,''

def upload_video_wall_media(file_storage):
    if not file_storage or not getattr(file_storage,'filename',''):
        return None, 'No media file selected.'
    project_ref=_supabase_project_ref(); key=os.getenv('SUPABASE_SERVICE_ROLE_KEY','').strip()
    if not (project_ref and key):
        return None, 'Supabase media upload is not configured. Add SUPABASE_SERVICE_ROLE_KEY in Render, or use an external media URL.'
    file_storage.stream.seek(0,2); size=file_storage.stream.tell(); file_storage.stream.seek(0)
    spec,error=_video_wall_media_spec(file_storage.filename,file_storage.mimetype,size)
    if error: return None,error
    # This compatibility path is for browsers without JavaScript. Large files
    # use the direct resumable browser uploader and never enter web-worker RAM.
    if size>50*1024*1024:
        return None,'Use the resumable uploader shown on this page for files larger than 50 MB.'
    ready,error=_ensure_video_wall_bucket()
    if not ready: return None,error
    mime=spec['mime'];ext=spec['extension']
    path=f"video-wall/{datetime.datetime.utcnow():%Y/%m}/{uuid.uuid4().hex}{ext}"
    base=f'https://{project_ref}.supabase.co'
    object_url=f"{base}/storage/v1/object/{VIDEO_WALL_BUCKET}/{urllib.parse.quote(path,safe='/')}"
    headers={'Authorization':f'Bearer {key}','apikey':key,'Content-Type':mime,'x-upsert':'false'}
    try:
        data=file_storage.stream.read()
        r=requests.post(object_url,headers=headers,data=data,timeout=180)
        if not r.ok:
            return None, f'Supabase Storage upload failed ({r.status_code}): {r.text[:300]}'
        public=f"{base}/storage/v1/object/public/{VIDEO_WALL_BUCKET}/{urllib.parse.quote(path,safe='/')}"
        drive_warning=''
        if setting('google_drive_auto_backup','0')=='1' and google_connected():
            _,drive_warning=google_drive_upload_bytes(data,file_storage.filename or os.path.basename(path),mime,source='video-wall',uploaded_by=current_user())
        return {'path':path,'url':public,'size':size,'mime':mime,'type':('image' if mime.startswith('image/') else 'video'),'drive_backup_error':drive_warning}, ''
    except Exception as exc:
        return None, f'Upload failed: {exc}'

def active_festive_session():
    return FestiveSession.query.filter_by(active=True).order_by(FestiveSession.id.desc()).first()

def screen_is_online(screen):
    if not screen.last_seen_at: return False
    return (datetime.datetime.utcnow()-screen.last_seen_at).total_seconds() <= 90

def screen_player_state(screen):
    festive=active_festive_session()
    assets=[]
    if festive and festive.asset_id:
        a=db.session.get(VideoAsset,festive.asset_id)
        if a and a.active: assets=[a]
    else:
        try: ids=[int(x) for x in json.loads(screen.playlist_json or '[]') if str(x).isdigit()]
        except Exception: ids=[]
        if not ids and screen.current_asset_id: ids=[screen.current_asset_id]
        for aid in ids:
            a=db.session.get(VideoAsset,aid)
            if a and a.active: assets.append(a)
    def item(a): return {'id':a.id,'title':a.title,'url':a.public_url,'type':a.media_type,'mime':a.mime_type}
    return {
        'screen':{'id':screen.id,'name':screen.name,'city':screen.city,'location':screen.location_name,'enabled':bool(screen.enabled)},
        'asset':(item(assets[0]) if assets else None),
        'playlist':[item(a) for a in assets],
        'rotation':int(screen.rotation_degrees or 0),
        'fit':screen.fit_mode or 'contain',
        'loop':bool(screen.loop_media),
        'muted':bool(screen.muted),
        'slide_duration':max(3,int(screen.slide_duration_seconds or 10)),
        'festive':({'id':festive.id,'name':festive.name} if festive else None),
    }


def login_required(fn):
    @wraps(fn)
    def wrapper(*a, **kw):
        if not session.get('uid'):
            return redirect(url_for('login', next=request.path))
        u=current_user()
        if not u or not u.active:
            session.clear()
            flash('This user account is inactive.', 'danger')
            return redirect(url_for('login'))
        return fn(*a, **kw)
    return wrapper


def current_user():
    return db.session.get(User, session.get('uid')) if session.get('uid') else None


HOST3D_INTENSITIES = {'static': 0, 'gentle': 1, 'full': 2}
HOST3D_SIZES = {'small', 'medium', 'large'}
HOST3D_POSITIONS = {'bottom-left', 'bottom-right'}


def _setting_bool(key, default='1'):
    return str(setting(key, default)).strip().lower() in ('1', 'true', 'yes', 'on')


def mascot_preferences_for(user):
    global_default = setting('host3d_default_intensity', 'full').strip().lower()
    if global_default not in HOST3D_INTENSITIES:
        global_default = 'full'
    policy_max = setting('host3d_max_intensity', 'full').strip().lower()
    if policy_max not in HOST3D_INTENSITIES:
        policy_max = 'full'
    try:
        row = MascotPreference.query.filter_by(user_id=user.id).first() if user else None
    except Exception as exc:
        try:
            db.session.rollback()
        except Exception:
            pass
        try:
            app.logger.warning('Mascot preference lookup failed: %s', str(exc)[:180])
        except Exception:
            pass
        row = None
    requested = (row.intensity if row and row.intensity in HOST3D_INTENSITIES else global_default)
    effective = requested if HOST3D_INTENSITIES[requested] <= HOST3D_INTENSITIES[policy_max] else policy_max
    size = row.size if row and row.size in HOST3D_SIZES else 'medium'
    position = row.position if row and row.position in HOST3D_POSITIONS else 'bottom-right'
    default_city = setting('host3d_default_city', setting('companion_default_city', 'Gurugram'))[:120]
    return {
        'enabled': bool(row.enabled) if row else _setting_bool('companion_enabled', '1'),
        'intensity': effective,
        'requested_intensity': requested,
        'policy_max_intensity': policy_max,
        'size': size,
        'position': position,
        'operational_updates': bool(row.operational_updates) if row else _setting_bool('host3d_operational_updates_default', setting('companion_operations_enabled', '1')),
        'motivational_messages': bool(row.motivational_messages) if row else _setting_bool('host3d_motivational_default', setting('companion_quotes_enabled', '1')),
        'weather_reactions': bool(row.weather_reactions) if row else _setting_bool('host3d_weather_default', setting('companion_weather_effects', '1')),
        'weather_city': (row.weather_city or default_city)[:120] if row else default_city,
    }


def parse_date(v):
    if not v: return None
    for f in ('%Y-%m-%d','%d-%m-%Y','%d/%m/%Y'):
        try: return datetime.datetime.strptime(v, f).date()
        except Exception: pass
    return None


def room_status(room):
    if room.status_override in ('Maintenance','Blocked'): return room.status_override
    today = datetime.date.today()
    tenants = Tenant.query.filter_by(property_name=room.property_name, room_unit_no=room.room_no).order_by(Tenant.id.desc()).all()
    for t in tenants:
        j, l = parse_date(t.joining_date), parse_date(t.leaving_date)
        if t.status == 'Notice Given' and (not l or l >= today): return 'Notice Given'
        if j and j > today: return 'Upcoming'
        if (not j or j <= today) and (not l or l >= today) and t.status not in ('Vacated','Cancelled','Terminated'):
            return 'Occupied'
    return 'Vacant'


def sync_tenant_from_agreement(ag):
    d = ag.data
    if not (d.get('tenant_name') or d.get('room_unit_no')): return
    prop, room_no = d.get('property_name','').strip(), d.get('room_unit_no','').strip()
    if room_no:
        r = Room.query.filter_by(property_name=prop, room_no=room_no).first()
        if not r:
            r=Room(city=d.get('city',''), property_name=prop, premises=d.get('premises',''), room_no=room_no, room_type=d.get('room_type',''), capacity=d.get('occupancy_limit',''), standard_tariff=d.get('monthly_rent',''))
            db.session.add(r)
        else:
            if d.get('city'): r.city=d.get('city')
            if d.get('premises'): r.premises=d.get('premises')
            if d.get('room_type'): r.room_type=d.get('room_type')
            if d.get('monthly_rent'): r.standard_tariff=d.get('monthly_rent')
    t = Tenant.query.filter_by(agreement_id=ag.id).first()
    if not t: t=Tenant(agreement_id=ag.id); db.session.add(t)
    t.tenant_name=d.get('tenant_name',''); t.tenant_mobile=d.get('tenant_mobile',''); t.tenant_whatsapp=d.get('tenant_whatsapp') or d.get('tenant_mobile','')
    t.tenant_email=d.get('tenant_email',''); t.tenant_id_type=d.get('tenant_id_type',''); t.tenant_id_no=d.get('tenant_id_no','')
    t.city=d.get('city',''); t.property_name=prop; t.premises=d.get('premises',''); t.room_unit_no=room_no; t.room_type=d.get('room_type','')
    t.tariff=d.get('monthly_rent',''); t.security_deposit=d.get('security_deposit',''); t.joining_date=d.get('start_date',''); t.leaving_date=d.get('end_date','')
    t.agreement_reference=d.get('agreement_reference','')
    j,l=parse_date(t.joining_date), parse_date(t.leaving_date); today=datetime.date.today()
    t.status='Upcoming' if j and j>today else ('Vacated' if l and l<today else 'Occupied')
    db.session.commit()


def all_form_data(preset_name=None):
    preset_name = preset_name or request.form.get('agreement_template') or 'Strong Residential - 11 Months'
    d = dict(DEFAULTS)
    d.update(PRESETS.get(preset_name, {}))
    for key, *_ in FIELDS:
        if key in request.form:
            d[key]=request.form.get(key,'').strip()
    for key in request.form:
        if key not in d:
            d[key]=request.form.get(key,'').strip()
    d['agreement_template']=preset_name
    return d


LIVENZA_APP_REGISTRY = [
    {'title':'Home','endpoint':'dashboard','permission':'','icon':'home','tone':'finder','availability':'internal'},
    {'title':'Agreement Studio','endpoint':'agreements','permission':'agreements','icon':'agreement','tone':'blue','availability':'internal'},
    {'title':'Rooms','endpoint':'rooms','permission':'rooms','icon':'room','tone':'cyan','availability':'internal'},
    {'title':'Residents','endpoint':'tenants','permission':'rooms','icon':'resident','tone':'teal','availability':'internal'},
    {'title':'Queries','endpoint':'queries','permission':'queries','icon':'queries','tone':'orange','availability':'internal'},
    {'title':'Reviews','endpoint':'reviews','permission':'reviews','icon':'review','tone':'yellow','availability':'internal'},
    {'title':'Video Wall','endpoint':'video_wall','permission':'video_wall','icon':'video','tone':'pink','availability':'internal'},
    {'title':'Food','endpoint':'food','permission':'food','icon':'food','tone':'green','availability':'internal'},
    {'title':'Billing','endpoint':'billing','permission':'rentok','icon':'billing','tone':'mint','availability':'internal'},
    {'title':'Banking','endpoint':'banking_suite','permission':'banking','icon':'banking','tone':'navy','availability':'internal'},
    {'title':'Electricity','endpoint':'electricity_studio','permission':'electricity','icon':'electricity','tone':'amber','availability':'internal'},
    {'title':'WhatsApp','endpoint':'whatsapp_workspace','permission':'whatsapp','icon':'whatsapp','tone':'green','availability':'whatsapp'},
    {'title':'Email','endpoint':'email_workspace','permission':'email','icon':'email','tone':'blue','availability':'google'},
    {'title':'Drive','endpoint':'drive_workspace','permission':'drive','icon':'drive','tone':'cyan','availability':'google'},
    {'title':'Letterhead Studio','endpoint':'letterhead_studio','permission':'letterhead','icon':'letterhead','tone':'red','availability':'internal'},
    {'title':'System Settings','endpoint':'settings_page','permission':'','icon':'settings','tone':'settings','availability':'internal'},
]


def ui_app_available(endpoint, user=None):
    """Return True only when a visible app route is real, authorized and usable."""
    item=next((row for row in LIVENZA_APP_REGISTRY if row['endpoint']==endpoint),None)
    route_names={rule.endpoint for rule in app.url_map.iter_rules()}
    if endpoint not in route_names:
        return False
    if not item:
        return True
    user=user or current_user()
    permission=item.get('permission') or ''
    if permission and not can_access(permission,user):
        return False
    availability=item.get('availability') or 'internal'
    if availability=='whatsapp':
        return whatsapp_cloud_configured()
    if availability=='google':
        return google_oauth_configured() and google_connected()
    return True


def visible_dock_apps(user=None):
    user=user or current_user()
    return [dict(item) for item in LIVENZA_APP_REGISTRY if ui_app_available(item['endpoint'],user)]


@app.context_processor
def inject_common():
    # Public authentication screens must render without any database-backed
    # Settings or companion dependency. This keeps /login available even when
    # a deployment is waiting on a settings/mascot schema migration.
    if request.endpoint in {'login','logout','diagnostics_runtime'}:
        return dict(
            current_user=None, app_version=APP_VERSION, os_name=OS_NAME, os_version=OS_VERSION, os_build=OS_BUILD,
            can_access=can_access, module_labels=MODULES, is_admin=False, masked_aadhaar=masked_aadhaar,
            kiosk_mode_enabled=False, marquee_enabled=False, companion_enabled=False,
            companion_default_city='Gurugram', companion_weather_effects=False, mascot_preferences={}, dock_apps=[], ui_app_available=lambda endpoint: False
        )
    user=current_user()
    mascot_preferences=mascot_preferences_for(user) if user else {}
    return dict(
        current_user=user, app_version=APP_VERSION, os_name=OS_NAME, os_version=OS_VERSION, os_build=OS_BUILD,
        can_access=can_access, module_labels=MODULES,
        is_admin=bool(user and (user.role or '').lower()=='admin'), masked_aadhaar=masked_aadhaar,
        kiosk_mode_enabled=setting('kiosk_mode_enabled','0')=='1', marquee_enabled=setting('marquee_enabled','1')=='1',
        companion_enabled=setting('companion_enabled','1')=='1' and bool(mascot_preferences.get('enabled',True)),
        companion_default_city=mascot_preferences.get('weather_city') or setting('companion_default_city','Gurugram'),
        companion_weather_effects=setting('companion_weather_effects','1')=='1' and bool(mascot_preferences.get('weather_reactions',True)),
        mascot_preferences=mascot_preferences, dock_apps=visible_dock_apps(user), ui_app_available=ui_app_available
    )

@app.before_request
def enforce_kiosk_pin_gate():
    """Server-side gate for every authenticated page while application lock is enabled."""
    # Diagnostics must never depend on Settings/current-user lookups before their
    # own stage-by-stage error reporting has a chance to run.
    if request.endpoint in {'health','version','diagnostics','diagnostics_authenticated','diagnostics_runtime','static'}:
        return None
    if not session.get('uid') or setting('kiosk_mode_enabled','0')!='1' or session.get('kiosk_unlocked'):
        return None
    allowed={'kiosk_lock','kiosk_unlock','logout'}
    if request.endpoint not in allowed:
        return redirect(url_for('kiosk_lock',next=request.full_path.rstrip('?')))

@app.route('/kiosk')
@login_required
def kiosk_lock():
    if setting('kiosk_mode_enabled','0')!='1' or session.get('kiosk_unlocked'):
        return redirect(url_for('dashboard'))
    return render_template('kiosk_lock.html')

@app.route('/kiosk/unlock',methods=['POST'])
@login_required
def kiosk_unlock():
    u=current_user(); secret=request.form.get('secret','')
    now=datetime.datetime.utcnow().timestamp(); attempts=int(session.get('kiosk_failed_attempts') or 0); last=float(session.get('kiosk_failed_at') or 0)
    if attempts>=5 and now-last<60:
        flash('Too many unlock attempts. Wait one minute and try again.','danger'); return redirect(url_for('kiosk_lock'))
    if now-last>=60: attempts=0
    pin_hash=setting('kiosk_pin_hash','')
    valid=bool((pin_hash and check_password_hash(pin_hash,secret)) or check_password_hash(u.password_hash,secret))
    if not valid:
        session['kiosk_failed_attempts']=attempts+1; session['kiosk_failed_at']=now
        flash('Incorrect kiosk PIN or account password.','danger'); return redirect(url_for('kiosk_lock'))
    session.pop('kiosk_failed_attempts',None); session.pop('kiosk_failed_at',None)
    session['kiosk_unlocked']=True
    target=request.form.get('next','')
    return redirect(target if target.startswith('/') and not target.startswith('//') else url_for('dashboard'))

@app.route('/kiosk/lock',methods=['POST'])
@login_required
def kiosk_relock():
    if setting('kiosk_mode_enabled','0')=='1':
        session['kiosk_unlocked']=False
        return redirect(url_for('kiosk_lock'))
    return redirect(url_for('dashboard'))

@app.route('/health')
def health(): return jsonify(status='ok', service='livenza-back-office-web', version=APP_VERSION)

@app.after_request
def livenza_cache_policy(response):
    # Versioned static assets may be cached aggressively; HTML must revalidate so
    # deploys become visible without forcing every CSS/JS/image request to reload.
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    elif response.mimetype == 'text/html':
        response.headers['Cache-Control'] = 'no-cache, max-age=0, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
    response.headers['X-Livenza-Build'] = APP_VERSION
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'publickey-credentials-get=(self)'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    return response

@app.teardown_request
def livenza_request_cleanup(exc=None):
    # Flask-SQLAlchemy removes sessions at app-context teardown. Rolling back on
    # failed requests ensures a broken transaction never poisons the next page.
    if exc is not None:
        try:
            db.session.rollback()
        except Exception:
            pass

@app.route('/diagnostics')
def diagnostics():
    route_names = {rule.endpoint for rule in app.url_map.iter_rules()}
    checks = {
        'version': APP_VERSION,
        'video_wall_route_loaded': 'video_wall' in route_names,
        'video_wall_player_loaded': 'wall_player' in route_names,
        'video_wall_template_exists': os.path.exists(os.path.join(BASE_DIR,'templates','video_wall.html')),
        'wall_player_template_exists': os.path.exists(os.path.join(BASE_DIR,'templates','wall_player.html')),
        'apple_theme_css_exists': os.path.exists(os.path.join(BASE_DIR,'static','style.css')),
        'billing_route_loaded': 'billing' in route_names,
    }
    try:
        checks['video_wall_tables_ready'] = db.session.execute(db.text("select count(*) from video_screen")).scalar() is not None
    except Exception as exc:
        checks['video_wall_tables_ready'] = False
        checks['database_error'] = str(exc)[:180]
        db.session.rollback()
    return jsonify(checks)

@app.route('/diagnostics/runtime')
def diagnostics_runtime():
    """Public fail-soft trace for request-time database/context/template dependencies."""
    result={'version':APP_VERSION,'build':OS_BUILD,'session_uid_present':bool(session.get('uid')),'stages':{}}

    def record(stage, fn):
        try:
            value=fn()
            result['stages'][stage]={'ok':True,'value':value}
            return True
        except Exception as exc:
            try: db.session.rollback()
            except Exception: pass
            result['stages'][stage]={'ok':False,'error_type':type(exc).__name__,'error':str(exc)[:900]}
            result.setdefault('failed_stage',stage)
            return False

    record('setting_table', lambda: {'rows':int(db.session.execute(db.text('select count(*) from setting')).scalar() or 0), 'columns':[c['name'] for c in inspect(db.engine).get_columns('setting')]})
    record('user_table', lambda: {'rows':int(db.session.execute(db.text('select count(*) from "user"')).scalar() or 0), 'columns':[c['name'] for c in inspect(db.engine).get_columns('user')]})
    record('mascot_preference_table', lambda: {'rows':int(db.session.execute(db.text('select count(*) from mascot_preference')).scalar() or 0), 'columns':[c['name'] for c in inspect(db.engine).get_columns('mascot_preference')]})
    record('setting_helper', lambda: {'kiosk_mode_enabled':setting('kiosk_mode_enabled','0'),'companion_enabled':setting('companion_enabled','1')})
    record('current_user', lambda: ({'id':current_user().id,'username':current_user().username,'role':current_user().role} if current_user() else None))
    record('context_processor', lambda: {k:(str(v)[:160] if not callable(v) else '<callable>') for k,v in inject_common().items() if k not in {'current_user'}})
    record('login_render', lambda: {'html_bytes':len(render_template('login.html').encode('utf-8'))})
    result['ok']='failed_stage' not in result
    return jsonify(result),200


@app.route('/diagnostics/authenticated')
def diagnostics_authenticated():
    """Trace the authenticated HTML-render path without turning a failing stage into a 500."""
    result={'session':{'ok':bool(session.get('uid'))}}
    uid=session.get('uid')
    if not uid:
        return jsonify(ok=False,failed_stage='session',stages=result),401

    try:
        user=db.session.get(User,uid)
        if not user:
            return jsonify(ok=False,failed_stage='user_load',stages={**result,'user_load':{'ok':False,'error':'Session user does not exist.'}}),409
        result['user_load']={'ok':True,'user_id':user.id,'role':user.role or ''}
    except Exception as exc:
        db.session.rollback()
        result['user_load']={'ok':False,'error_type':type(exc).__name__,'error':str(exc)[:500]}
        return jsonify(ok=False,failed_stage='user_load',stages=result),200

    try:
        result['settings']={'ok':True,'kiosk_mode_enabled':setting('kiosk_mode_enabled','0')=='1'}
    except Exception as exc:
        db.session.rollback()
        result['settings']={'ok':False,'error_type':type(exc).__name__,'error':str(exc)[:500]}
        return jsonify(ok=False,failed_stage='settings',stages=result),200

    try:
        prefs=mascot_preferences_for(user)
        result['mascot_preferences']={'ok':True,'enabled':bool(prefs.get('enabled'))}
    except Exception as exc:
        db.session.rollback()
        result['mascot_preferences']={'ok':False,'error_type':type(exc).__name__,'error':str(exc)[:500]}
        return jsonify(ok=False,failed_stage='mascot_preferences',stages=result),200

    try:
        result['permissions']={'ok':True,'settings_access':can_access('settings',user),'letterhead_access':can_access('letterhead',user)}
    except Exception as exc:
        db.session.rollback()
        result['permissions']={'ok':False,'error_type':type(exc).__name__,'error':str(exc)[:500]}
        return jsonify(ok=False,failed_stage='permissions',stages=result),200

    try:
        rendered=render_template('dashboard.html',show_login_welcome=False)
        result['dashboard_render']={'ok':True,'html_bytes':len(rendered.encode('utf-8'))}
    except Exception as exc:
        db.session.rollback()
        result['dashboard_render']={'ok':False,'error_type':type(exc).__name__,'error':str(exc)[:900]}
        return jsonify(ok=False,failed_stage='dashboard_render',stages=result),200

    return jsonify(ok=True,failed_stage=None,stages=result)


@app.route('/version')
def version():
    return jsonify(
        name=OS_NAME, version=OS_VERSION, build=OS_BUILD,
        features=[
            'tesla-os-27','macos27-clean-shell','wallpaper-picker','functional-app-registry','raf-dock-motion',
            'unified-system-settings','livenza-symbol-system','ai-logo-identity','reduced-motion',
            'agreements','rooms','queries','billing','banking','electricity','food','video-wall',
            'whatsapp','email','drive','letterhead-studio','livenza-vault','role-permissions',
            'webauthn-passkeys','pattern-login','live-companion','responsive-5k-layout'
        ]
    )

@app.route('/login', methods=['GET','POST'])
def login():
    username=request.form.get('username','').strip() if request.method=='POST' else ''
    method=request.form.get('auth_method','password') if request.method=='POST' else 'fingerprint'
    if request.method=='POST' and method not in {'password','pattern'}: method='password'
    error={}
    if request.method=='POST':
        u=User.query.filter_by(username=username).first() if username else None
        if not username:
            error={'field':'username','message':'Enter your Login ID before choosing a sign-in method.'}
        elif u and not u.active:
            error={'field':'account','message':'Sign-in is unavailable for this account. Ask an administrator to reactivate access.'}
        elif method=='pattern':
            raw_pattern=request.form.get('pattern','')
            pattern=_pattern_value(raw_pattern)
            if not pattern:
                error={'field':'pattern','message':'Connect at least four different points, then try the gesture again.'}
            elif u and u.pattern_hash and check_password_hash(u.pattern_hash,'pattern:'+pattern):
                session.clear(); session['uid']=u.id
                session['kiosk_unlocked']=setting('kiosk_mode_enabled','0')!='1'
                session['show_login_welcome']=True
                return redirect(url_for('kiosk_lock') if not session['kiosk_unlocked'] else (request.args.get('next') or url_for('dashboard')))
            else:
                error={'field':'pattern','message':'That pattern did not match. Clear the grid, redraw the saved sequence, or use another sign-in method.'}
        else:
            password=request.form.get('password','')
            if not password:
                error={'field':'password','message':'Enter your password to continue.'}
            elif u and check_password_hash(u.password_hash,password):
                session.clear(); session['uid']=u.id
                session['kiosk_unlocked']=setting('kiosk_mode_enabled','0')!='1'
                session['show_login_welcome']=True
                return redirect(url_for('kiosk_lock') if not session['kiosk_unlocked'] else (request.args.get('next') or url_for('dashboard')))
            else:
                error={'field':'password','message':'The password did not match. Check Caps Lock and try again, or use another sign-in method.'}
        return render_template('login.html',login_error=error,login_username=username,login_method=method),401
    return render_template('login.html',login_error=error,login_username=username,login_method=method)

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/account', methods=['GET','POST'])
@login_required
def account():
    if request.method=='POST':
        flash('Only administrators can change user details, passwords, photos or avatars.','danger')
    return redirect(settings_pane_url('account'))

@app.route('/account/mascot-settings', methods=['POST'])
@login_required
def account_mascot_settings():
    u = current_user()
    action = (request.form.get('mascot_action') or 'save').strip().lower()
    row = MascotPreference.query.filter_by(user_id=u.id).first()
    if action == 'reset':
        if row:
            db.session.delete(row)
            db.session.commit()
        flash('Your Livenza mascot preferences were reset to Livenza defaults.', 'success')
        return redirect(url_for('account') + '#mascot-settings')
    if not row:
        row = MascotPreference(user_id=u.id)
        db.session.add(row)
    def form_bool(name, default=True):
        values = request.form.getlist(name)
        if not values:
            return default
        return str(values[-1]).strip().lower() in ('1', 'true', 'yes', 'on')
    intensity = (request.form.get('intensity') or 'full').strip().lower()
    size = (request.form.get('size') or 'medium').strip().lower()
    position = (request.form.get('position') or 'bottom-right').strip().lower()
    row.enabled = form_bool('enabled', True)
    row.intensity = intensity if intensity in HOST3D_INTENSITIES else 'full'
    row.size = size if size in HOST3D_SIZES else 'medium'
    row.position = position if position in HOST3D_POSITIONS else 'bottom-right'
    row.operational_updates = form_bool('operational_updates', True)
    row.motivational_messages = form_bool('motivational_messages', True)
    row.weather_reactions = form_bool('weather_reactions', True)
    row.weather_city = (request.form.get('weather_city') or setting('host3d_default_city', 'Gurugram')).strip()[:120] or 'Gurugram'
    db.session.commit()
    flash('Your Livenza mascot settings were saved.', 'success')
    return redirect(url_for('account') + '#mascot-settings')


@app.route('/account/avatar',methods=['POST'])
@login_required
def account_avatar():
    wants_json=request.headers.get('X-Requested-With')=='XMLHttpRequest' or 'application/json' in request.headers.get('Accept','')
    error='Only administrators can change profile photos or live mascots.'
    if wants_json: return jsonify(ok=False,error=error),403
    flash(error,'danger'); return redirect(url_for('account'))

@app.route('/')
@login_required
def dashboard():
    # Tesla OS 27 Home is a vibrant macOS 27-style desktop workspace.
    # Operational data is still loaded only inside the suites that actually need it,
    # so a stale optional module/table cannot turn the Home route into a 500.
    show_login_welcome=bool(session.pop('show_login_welcome',False))
    return render_template('dashboard.html', show_login_welcome=show_login_welcome)

@app.route('/api/marquee')
@login_required
def marquee_status():
    try: refresh=max(30,min(600,int(setting('marquee_refresh_seconds','60') or 60)))
    except Exception: refresh=60
    return jsonify(ok=True,items=live_marquee_items(),refresh_seconds=refresh,updated_at=datetime.datetime.now(ZoneInfo('Asia/Kolkata')).isoformat())

@app.route('/api/companion/pulse')
@login_required
def companion_pulse():
    enabled=setting('companion_enabled','1')=='1'
    requested_city=request.args.get('city',setting('companion_default_city','Gurugram'))
    city=next((name for name in COMPANION_LOCATIONS if name.lower()==str(requested_city or '').strip().lower()),None) or setting('companion_default_city','Gurugram')
    if city not in COMPANION_LOCATIONS: city='Gurugram'
    weather=_companion_weather(city) if setting('companion_weather_enabled','1')=='1' else {'available':False,'city':city,'effect':'none','forecast':[]}
    try: effect_seconds=max(7,min(20,int(setting('companion_effect_seconds','11') or 11)))
    except Exception: effect_seconds=11
    operations=_companion_operations() if setting('companion_operations_enabled','1')=='1' else []
    operation_summary={str(item.get('label','Live')):str(item.get('value','—')) for item in operations}
    return jsonify(
        ok=True,enabled=enabled,weather=weather,
        weather_effects=setting('companion_weather_effects','1')=='1',effect_seconds=effect_seconds,
        operations=operations,operation_summary=operation_summary,
        quotes=COMPANION_QUOTES if setting('companion_quotes_enabled','1')=='1' else [],
        locations=list(COMPANION_LOCATIONS.keys()),refresh_seconds=120,
        updated_at=datetime.datetime.now(ZoneInfo('Asia/Kolkata')).isoformat(),
    )

@app.route('/api/presets/<path:name>')
@permission_required('agreements')
def preset_api(name):
    profile=FORMAT_PROFILES.get(name, FORMAT_PROFILES.get('Strong Residential - 11 Months',{}))
    preview_data=dict(DEFAULTS); preview_data.update(PRESETS.get(name,{})); preview_data['agreement_template']=name
    return jsonify({
        'values':PRESETS.get(name,{}),
        'required':sorted(agreement_required_fields(name,preview_data)),
        'profile':{
            'title_en':profile.get('title_en',''), 'subtitle_en':profile.get('subtitle_en',''),
            'nature_en':profile.get('nature_en','')
        }
    })

def agreement_editor_context(ag=None, data=None):
    d=dict(DEFAULTS)
    if ag: d.update(ag.data)
    if data: d.update(data)
    if not ag and not data: d.update(PRESETS.get('Strong Residential - 11 Months',{}))
    fields={x[0]:x for x in FIELDS}
    fields['city']=('city','City','entry',None)
    city_names=[c.name for c in City.query.filter_by(active=True).order_by(City.name).all()]
    preset=d.get('agreement_template') or 'Strong Residential - 11 Months'
    profile=FORMAT_PROFILES.get(preset, FORMAT_PROFILES.get('Strong Residential - 11 Months',{}))
    landlord_masters=[safe_master_summary('landlord',row) for row in LandlordMaster.query.filter_by(active=True).order_by(LandlordMaster.profile_name).all()]
    tenant_masters=[safe_master_summary('tenant',row) for row in TenantMaster.query.filter_by(active=True).order_by(TenantMaster.profile_name).all()]
    return dict(ag=ag,d=d,presets=PRESETS,field_map=fields,groups=AGREEMENT_GROUPS,
                required_fields=agreement_required_fields(preset,d),city_names=city_names,preset_profile=profile,
                landlord_masters=landlord_masters,tenant_masters=tenant_masters,
                selected_landlord_master_id=str(d.get('landlord_master_id') or ''),selected_tenant_master_id=str(d.get('tenant_master_id') or ''),
                annexure_documents=[])

def _agreement_aadhaar_extract_payload(upload):
    if not upload or not upload.filename:
        return {'ok':False,'error':'Choose an Aadhaar JPEG, PNG or PDF first.'},400
    filename=os.path.basename(upload.filename)
    ext=os.path.splitext(filename.lower())[1]
    if ext not in ('.jpg','.jpeg','.png','.pdf'):
        return {'ok':False,'error':'Supported formats: JPG, JPEG, PNG and PDF.'},400
    raw=upload.read()
    if not raw:
        return {'ok':False,'error':'The uploaded file is empty.'},400
    if len(raw)>10*1024*1024:
        return {'ok':False,'error':'Aadhaar upload must be 10 MB or smaller.'},413
    mimetype=(upload.mimetype or '').lower()
    data={}; notes=[]; local_error=''
    # Text-based PDFs can be parsed first without sending the file to any service.
    if ext=='.pdf':
        try:
            from pypdf import PdfReader
            reader=PdfReader(io.BytesIO(raw))
            pdf_text='\n'.join((page.extract_text() or '') for page in reader.pages[:4])
            if len(pdf_text.strip())>40:
                data=_parse_aadhaar_text_fallback(pdf_text)
                notes.append('Text was read directly from the PDF.')
        except Exception:
            pass
    essential=sum(bool(data.get(k)) for k in ('tenant_name','tenant_dob','tenant_address','tenant_id_no'))
    if essential<3:
        ocr_text,local_error=_aadhaar_local_ocr_text(raw,filename,mimetype or ('application/pdf' if ext=='.pdf' else 'image/jpeg'))
        if ocr_text:
            ocr_data=_parse_aadhaar_text_fallback(ocr_text)
            for key,value in ocr_data.items():
                if value and not data.get(key): data[key]=value
            notes.append('Automatic server-side OCR filled the visible identity details.')
        essential=sum(bool(data.get(k)) for k in ('tenant_name','tenant_dob','tenant_address','tenant_id_no'))
    if essential<3:
        try:
            ai_data,ai_error=_aadhaar_ai_extract(raw,filename,mimetype or ('application/pdf' if ext=='.pdf' else 'image/jpeg'))
            if ai_data:
                for k,v in ai_data.items():
                    if v: data[k]=v
                notes.append('Secure AI enhancement completed the remaining fields.')
        except Exception:
            ai_error='AI enhancement could not complete.'
    if not any(data.get(k) for k in ('tenant_name','tenant_dob','tenant_address','tenant_id_no')):
        if local_error and not os.getenv('OPENAI_API_KEY','').strip():
            message='The secure OCR reader is not ready on this deployment. Redeploy Tesla OS 27 Version 27.0.1 with its current system and Python dependencies, then try again.'
        else:
            message='No reliable Aadhaar fields were detected. Use a clear, straight photo in good light or a PDF containing both sides, then try again.'
        return {'ok':False,'error':message,'reader_status':local_error or 'No readable identity fields detected.'},422
    return {'ok':True,'fields':data,'note':' '.join(notes),'warning':'Autofill only — review the extracted details. This does not verify Aadhaar authenticity with UIDAI.'},200


@app.route('/agreements/aadhaar-extract', methods=['POST'])
@permission_required('agreements')
def agreement_aadhaar_extract():
    """Always return JSON so the form never collapses into a generic browser error."""
    try:
        payload,status=_agreement_aadhaar_extract_payload(request.files.get('aadhaar_file'))
        return jsonify(**payload),status
    except RequestEntityTooLarge:
        raise
    except Exception:
        return jsonify(ok=False,error='The secure Aadhaar reader encountered a temporary server error. The document was not saved. Try a smaller, clearer JPG/PNG or redeploy the updated OCR dependencies.'),500

@app.route('/api/agreement-party-profiles',methods=['POST'])
@permission_required('agreements')
def agreement_party_profile_save():
    body=request.get_json(silent=True) or {}
    profile_type=str(body.get('profile_type') or '').strip().lower()
    if profile_type not in AGREEMENT_PARTY_PROFILE_FIELDS:
        return jsonify(ok=False,error='Choose a landlord or tenant profile.'),400
    name=str(body.get('name') or '').strip()[:180]
    if not name: return jsonify(ok=False,error='Give this profile a recognisable name.'),400
    supplied=body.get('fields') if isinstance(body.get('fields'),dict) else {}
    clean={}
    for key in AGREEMENT_PARTY_PROFILE_FIELDS[profile_type]:
        value=str(supplied.get(key) or '').strip()
        if value: clean[key]=value[:4000]
    if not clean: return jsonify(ok=False,error=f'Fill at least one {profile_type} detail before saving the profile.'),400
    saved=AgreementPartyProfile.query.filter(
        AgreementPartyProfile.profile_type==profile_type,
        func.lower(AgreementPartyProfile.name)==name.lower(),
    ).first()
    if not saved:
        saved=AgreementPartyProfile(profile_type=profile_type,name=name,created_by_user_id=current_user().id)
        db.session.add(saved)
    saved.name=name
    saved.data_ciphertext=_integration_cipher().encrypt(json.dumps(clean,ensure_ascii=False).encode('utf-8')).decode('ascii')
    db.session.commit()
    return jsonify(ok=True,profile={'id':saved.id,'name':saved.name,'profile_type':profile_type,'fields':clean},message=f'{profile_type.title()} profile saved securely.')

@app.route('/api/agreement-party-profiles/<int:profile_id>',methods=['DELETE'])
@permission_required('agreements')
def agreement_party_profile_delete(profile_id):
    saved=db.session.get(AgreementPartyProfile,profile_id)
    if not saved: return jsonify(ok=False,error='Saved profile not found.'),404
    db.session.delete(saved);db.session.commit()
    return jsonify(ok=True,message='Saved party profile removed.')

# ===== Tesla OS 27 • Separate Landlord / Tenant Masters =====
def _master_model(kind):
    if kind=='landlord': return LandlordMaster
    if kind=='tenant': return TenantMaster
    abort(404)

def _master_list_query(Model):
    q=(request.args.get('q') or '').strip()
    status=(request.args.get('status') or 'active').strip().lower()
    party_type=(request.args.get('party_type') or '').strip().lower()
    city=(request.args.get('city') or '').strip()
    query=Model.query
    if status=='active': query=query.filter_by(active=True)
    elif status=='archived': query=query.filter_by(active=False)
    if party_type: query=query.filter(func.lower(Model.party_type)==party_type)
    if city: query=query.filter(func.lower(Model.city)==city.lower())
    if q:
        like=f'%{q.lower()}%'
        query=query.filter(or_(func.lower(Model.profile_name).like(like),func.lower(Model.legal_name).like(like),func.lower(Model.primary_mobile).like(like),func.lower(Model.email).like(like),func.lower(Model.city).like(like),func.lower(Model.tags).like(like)))
    return query.order_by(Model.active.desc(),Model.updated_at.desc())

def _master_rows_with_admin_exact_identifier(Model,kind):
    rows=_master_list_query(Model).all()
    q=(request.args.get('q') or '').strip()
    user=current_user()
    if not q or not user or (user.role or '').lower()!='admin': return rows
    try: wanted=identifier_lookup_hash(q,_master_key())
    except Exception: return rows
    if not wanted: return rows
    seen={row.id for row in rows}
    for row in Model.query.order_by(Model.active.desc(),Model.updated_at.desc()).all():
        if row.id in seen: continue
        try: hashes=json.loads(row.identifier_lookup_json or '[]')
        except Exception: hashes=[]
        if wanted in hashes:
            rows.append(row);seen.add(row.id)
    return rows

def _master_form_payload(kind,row=None):
    supplied={key:request.form.get(key,'') for key in MASTER_FIELD_SET[kind]}
    old=_master_payload(row) if row else {}
    normalized=normalize_master_payload(kind,supplied)
    for key in SENSITIVE_FIELDS[kind]:
        if not str(supplied.get(key) or '').strip() and old.get(key): normalized[key]=old[key]
    # Re-normalize derived search text after preservation without ever indexing sensitive values.
    normalized=normalize_master_payload(kind,normalized)
    return normalized

def _master_agreement_usage(kind,mid):
    if kind not in ('landlord','tenant'): return []
    key='landlord_master_id' if kind=='landlord' else 'tenant_master_id'
    wanted=str(mid)
    return [ag for ag in Agreement.query.order_by(Agreement.updated_at.desc()).all() if str(ag.data.get(key) or '')==wanted]

def _safe_audit_note(value):
    note=str(value or '').strip()[:180]
    if not note: return ''
    lowered=note.lower()
    blocked=('aadhaar','passport','bank account','ifsc','upi','password','secret','otp','cvv','card pin','upi pin')
    if any(term in lowered for term in blocked): return ''
    return note

def _safe_master_audit_history(kind,mid,limit=80):
    target_type=f'{kind}_master'
    rows=AuditEvent.query.filter_by(module='agreement_master',target_type=target_type,target_id=mid).order_by(AuditEvent.id.desc()).limit(limit).all()
    actor_ids={row.actor_user_id for row in rows if row.actor_user_id}
    actors={u.id:(u.full_name or u.username) for u in User.query.filter(User.id.in_(actor_ids)).all()} if actor_ids else {}
    return [{'action':row.action,'status':row.status,'created_at':row.created_at,'actor':actors.get(row.actor_user_id,'System'),'note':_safe_audit_note(row.note)} for row in rows]

def _master_template_context(kind,row=None,read_only=False):
    payload=_master_payload(row) if row else normalize_master_payload(kind,{})
    display_payload=master_display_payload(kind,payload)
    documents=[];usage=[];audit_history=[]
    if row:
        documents=MasterDocument.query.filter_by(**({'landlord_master_id':row.id} if kind=='landlord' else {'tenant_master_id':row.id})).order_by(MasterDocument.uploaded_at.desc()).all()
        usage=_master_agreement_usage(kind,row.id)
        audit_history=_safe_master_audit_history(kind,row.id)
    return dict(master=row,kind=kind,display_payload=display_payload,field_names=MASTER_FIELD_SET[kind],sensitive_fields=SENSITIVE_FIELDS[kind],document_categories=DOCUMENT_CATEGORIES,documents=documents,usage=usage,audit_history=audit_history,read_only=read_only)

@app.route('/agreements/landlords')
@permission_required('agreements')
def landlord_masters():
    rows=_master_rows_with_admin_exact_identifier(LandlordMaster,'landlord');return render_template('landlord_masters.html',items=rows,usage_counts={row.id:len(_master_agreement_usage('landlord',row.id)) for row in rows},q=(request.args.get('q') or ''),status=(request.args.get('status') or 'active'),is_master_admin=(current_user().role or '').lower()=='admin')

@app.route('/agreements/tenants/master')
@permission_required('agreements')
def tenant_masters():
    rows=_master_rows_with_admin_exact_identifier(TenantMaster,'tenant');return render_template('tenant_masters.html',items=rows,usage_counts={row.id:len(_master_agreement_usage('tenant',row.id)) for row in rows},q=(request.args.get('q') or ''),status=(request.args.get('status') or 'active'),is_master_admin=(current_user().role or '').lower()=='admin')

@app.route('/agreements/landlords/<int:mid>')
@permission_required('agreements')
def landlord_master_safe_view(mid):
    row=db.session.get(LandlordMaster,mid) or abort(404)
    payload=master_display_payload('landlord',_master_payload(row))
    mapped=[{'field':agreement_key,'label':agreement_key.replace('_',' ').title(),'value':payload.get(master_key,'')} for master_key,agreement_key in LANDLORD_AGREEMENT_MAP.items() if str(payload.get(master_key,'') or '').strip()]
    return render_template('master_safe_view.html',kind='landlord',master=row,summary=safe_master_summary('landlord',row),mapped=mapped,usage_count=len(_master_agreement_usage('landlord',row.id)))

@app.route('/agreements/tenants/master/<int:mid>')
@permission_required('agreements')
def tenant_master_safe_view(mid):
    row=db.session.get(TenantMaster,mid) or abort(404)
    payload=master_display_payload('tenant',_master_payload(row))
    mapped=[{'field':agreement_key,'label':agreement_key.replace('_',' ').title(),'value':payload.get(master_key,'')} for master_key,agreement_key in TENANT_AGREEMENT_MAP.items() if str(payload.get(master_key,'') or '').strip()]
    return render_template('master_safe_view.html',kind='tenant',master=row,summary=safe_master_summary('tenant',row),mapped=mapped,usage_count=len(_master_agreement_usage('tenant',row.id)))

@app.route('/agreements/landlords/new',methods=['GET','POST'])
@app.route('/agreements/landlords/<int:mid>/edit',methods=['GET','POST'])
@admin_required
def landlord_master_edit(mid=None):
    row=db.session.get(LandlordMaster,mid) if mid else None
    if mid and not row: abort(404)
    if request.method=='POST':
        try:
            payload=_master_form_payload('landlord',row)
            if not payload.get('profile_name') and not payload.get('full_legal_name'):
                flash('Profile name or legal name is required.','danger');return render_template('landlord_master_edit.html',**_master_template_context('landlord',row))
            created=row is None
            if not row:
                row=LandlordMaster(master_code=_new_master_code('landlord'),profile_name=payload.get('profile_name') or payload.get('full_legal_name') or 'Landlord',created_by_user_id=current_user().id)
                db.session.add(row);db.session.flush()
            _set_master_payload(row,payload);row.updated_by_user_id=current_user().id
            record_audit('landlord_master_created' if created else 'landlord_master_updated','landlord_master',row.id,module='agreement_master',meta={'party_type':row.party_type,'city':row.city,'verification_status':row.verification_status})
            db.session.commit();flash('Landlord Master saved securely.','success');return redirect(url_for('landlord_master_edit',mid=row.id))
        except Exception as exc:
            db.session.rollback();flash(str(exc),'danger')
    return render_template('landlord_master_edit.html',**_master_template_context('landlord',row))

@app.route('/agreements/tenants/master/new',methods=['GET','POST'])
@app.route('/agreements/tenants/master/<int:mid>/edit',methods=['GET','POST'])
@admin_required
def tenant_master_edit(mid=None):
    row=db.session.get(TenantMaster,mid) if mid else None
    if mid and not row: abort(404)
    if request.method=='POST':
        try:
            payload=_master_form_payload('tenant',row)
            if not payload.get('profile_name') and not payload.get('full_legal_name'):
                flash('Profile name or legal name is required.','danger');return render_template('tenant_master_edit.html',**_master_template_context('tenant',row))
            created=row is None
            if not row:
                row=TenantMaster(master_code=_new_master_code('tenant'),profile_name=payload.get('profile_name') or payload.get('full_legal_name') or 'Tenant',created_by_user_id=current_user().id)
                db.session.add(row);db.session.flush()
            _set_master_payload(row,payload);row.updated_by_user_id=current_user().id
            record_audit('tenant_master_created' if created else 'tenant_master_updated','tenant_master',row.id,module='agreement_master',meta={'party_type':row.party_type,'city':row.city,'verification_status':row.verification_status})
            db.session.commit();flash('Tenant Master saved securely.','success');return redirect(url_for('tenant_master_edit',mid=row.id))
        except Exception as exc:
            db.session.rollback();flash(str(exc),'danger')
    return render_template('tenant_master_edit.html',**_master_template_context('tenant',row))

@app.route('/agreements/landlords/<int:mid>/duplicate',methods=['POST'])
@admin_required
def landlord_master_duplicate(mid):
    source=db.session.get(LandlordMaster,mid) or abort(404);payload=_master_payload(source)
    payload['profile_name']=((payload.get('profile_name') or source.profile_name)+' Copy')[:180];payload['verification_status']='unverified'
    row=LandlordMaster(master_code=_new_master_code('landlord'),profile_name=payload['profile_name'],created_by_user_id=current_user().id,updated_by_user_id=current_user().id)
    _set_master_payload(row,payload);db.session.add(row);db.session.flush();record_audit('landlord_master_duplicated','landlord_master',row.id,module='agreement_master',meta={'source_master_id':source.id});db.session.commit()
    flash('Landlord Master duplicated without copying protected documents.','success');return redirect(url_for('landlord_master_edit',mid=row.id))

@app.route('/agreements/tenants/master/<int:mid>/duplicate',methods=['POST'])
@admin_required
def tenant_master_duplicate(mid):
    source=db.session.get(TenantMaster,mid) or abort(404);payload=_master_payload(source)
    payload['profile_name']=((payload.get('profile_name') or source.profile_name)+' Copy')[:180];payload['verification_status']='unverified'
    row=TenantMaster(master_code=_new_master_code('tenant'),profile_name=payload['profile_name'],created_by_user_id=current_user().id,updated_by_user_id=current_user().id)
    _set_master_payload(row,payload);db.session.add(row);db.session.flush();record_audit('tenant_master_duplicated','tenant_master',row.id,module='agreement_master',meta={'source_master_id':source.id});db.session.commit()
    flash('Tenant Master duplicated without copying protected documents.','success');return redirect(url_for('tenant_master_edit',mid=row.id))

@app.route('/agreements/landlords/<int:mid>/archive',methods=['POST'])
@admin_required
def landlord_master_archive(mid):
    row=db.session.get(LandlordMaster,mid) or abort(404);row.active=not bool(row.active);record_audit('landlord_master_reactivated' if row.active else 'landlord_master_archived','landlord_master',row.id,module='agreement_master');db.session.commit();flash('Landlord Master reactivated.' if row.active else 'Landlord Master archived.','success');return redirect(url_for('landlord_masters'))

@app.route('/agreements/tenants/master/<int:mid>/archive',methods=['POST'])
@admin_required
def tenant_master_archive(mid):
    row=db.session.get(TenantMaster,mid) or abort(404);row.active=not bool(row.active);record_audit('tenant_master_reactivated' if row.active else 'tenant_master_archived','tenant_master',row.id,module='agreement_master');db.session.commit();flash('Tenant Master reactivated.' if row.active else 'Tenant Master archived.','success');return redirect(url_for('tenant_masters'))

def _agreement_master_reverse_payload(kind,fields,profile_name=''):
    if kind not in ('landlord','tenant'): abort(404)
    mapping=LANDLORD_AGREEMENT_MAP if kind=='landlord' else TENANT_AGREEMENT_MAP
    reverse={agreement_key:master_key for master_key,agreement_key in mapping.items()}
    payload={}
    for agreement_key,master_key in reverse.items():
        value=str((fields or {}).get(agreement_key) or '').strip()
        if value: payload[master_key]=value
    if profile_name: payload['profile_name']=str(profile_name).strip()[:180]
    elif payload.get('full_legal_name'): payload['profile_name']=payload['full_legal_name']
    return payload

@app.route('/api/agreement-masters/<kind>/<int:mid>/apply')
@permission_required('agreements')
def agreement_master_apply(kind,mid):
    if kind not in ('landlord','tenant'): abort(404)
    row=db.session.get(_master_model(kind),mid) or abort(404)
    if not row.active: return jsonify(ok=False,error='This master profile is archived.'),409
    payload=_master_payload(row);fields,_=apply_master_mapping(kind,payload,{},replace=True)
    record_audit('master_applied_to_agreement',f'{kind}_master',row.id,module='agreement_master',meta={'master_type':kind});db.session.commit()
    return jsonify(ok=True,master={'id':row.id,'name':row.profile_name},fields=fields)

@app.route('/api/agreement-masters/<kind>/from-agreement',methods=['POST'])
@admin_required
def agreement_master_from_agreement(kind):
    if kind not in ('landlord','tenant'): abort(404)
    body=request.get_json(silent=True) or {};fields=body.get('fields') if isinstance(body.get('fields'),dict) else {}
    payload=_agreement_master_reverse_payload(kind,fields,body.get('profile_name') or '')
    if not payload.get('full_legal_name'): return jsonify(ok=False,error='Fill the party name before creating a master.'),400
    Model=_master_model(kind);row=Model(master_code=_new_master_code(kind),profile_name=payload.get('profile_name') or payload['full_legal_name'],created_by_user_id=current_user().id,updated_by_user_id=current_user().id)
    _set_master_payload(row,payload);db.session.add(row);db.session.flush();record_audit(f'{kind}_master_created_from_agreement',f'{kind}_master',row.id,module='agreement_master');db.session.commit()
    return jsonify(ok=True,master=safe_master_summary(kind,row),message=f'{kind.title()} Master created from the current agreement details.')

@app.route('/api/agreement-masters/<kind>/<int:mid>/update-from-agreement',methods=['POST'])
@admin_required
def agreement_master_update_from_agreement(kind,mid):
    if kind not in ('landlord','tenant'): abort(404)
    row=db.session.get(_master_model(kind),mid) or abort(404);body=request.get_json(silent=True) or {};fields=body.get('fields') if isinstance(body.get('fields'),dict) else {}
    overlay=_agreement_master_reverse_payload(kind,fields);payload=_master_payload(row)
    for key,value in overlay.items():
        if str(value or '').strip(): payload[key]=value
    _set_master_payload(row,payload);row.updated_by_user_id=current_user().id;record_audit(f'{kind}_master_updated_from_agreement',f'{kind}_master',row.id,module='agreement_master');db.session.commit()
    return jsonify(ok=True,master=safe_master_summary(kind,row),message=f'{kind.title()} Master updated explicitly from this agreement.')

def _require_admin_password_from_form():
    admin=current_user();supplied=request.form.get('admin_password','')
    if not admin or (admin.role or '').lower()!='admin' or not check_password_hash(admin.password_hash,supplied):
        try:
            record_audit('admin_reauth_failed','agreement_master',None,status='failed',module='agreement_master');db.session.commit()
        except Exception: db.session.rollback()
        abort(403)
    return admin

def _master_document_owner(doc):
    if not doc: return None
    if doc.owner_type=='landlord' and doc.landlord_master_id: return db.session.get(LandlordMaster,doc.landlord_master_id)
    if doc.owner_type=='tenant' and doc.tenant_master_id: return db.session.get(TenantMaster,doc.tenant_master_id)
    return None

def _master_document_redirect(doc):
    return url_for('landlord_master_edit',mid=doc.landlord_master_id) if doc.owner_type=='landlord' else url_for('tenant_master_edit',mid=doc.tenant_master_id)

def _read_master_document_upload(upload):
    if not upload or not upload.filename: raise ValueError('Choose a document to upload.')
    raw=upload.stream.read(20*1024*1024+1)
    if len(raw)>20*1024*1024: raise ValueError('Master documents must be 20 MB or smaller.')
    ext,mime=validate_master_document(upload.filename,upload.mimetype or '',len(raw))
    if not raw: raise ValueError('The uploaded document is empty.')
    return raw,ext,mime

def _new_master_document(owner_type,mid,upload,source_doc=None):
    raw,ext,mime=_read_master_document_upload(upload)
    category=(request.form.get('category') or (source_doc.category if source_doc else 'miscellaneous')).strip()
    if category not in DOCUMENT_CATEGORIES: category='miscellaneous'
    label=(request.form.get('display_label') or DOCUMENT_CATEGORIES.get(category) or 'Supporting document').strip()[:180]
    ciphertext,nonce=encrypt_blob(raw,_master_key())
    doc=MasterDocument(owner_type=owner_type,category=category,display_label=label,storage_id=secrets.token_hex(24),extension=ext,mime_type=mime,ciphertext=ciphertext,nonce=nonce,issue_date=parse_date(request.form.get('issue_date')),expiry_date=parse_date(request.form.get('expiry_date')),verification_status=(request.form.get('verification_status') or 'unverified')[:40],replaced_document_id=(source_doc.id if source_doc else None),uploaded_by_user_id=current_user().id)
    if owner_type=='landlord': doc.landlord_master_id=mid
    else: doc.tenant_master_id=mid
    return doc

@app.route('/agreements/landlords/<int:mid>/documents',methods=['POST'])
@app.route('/agreements/tenants/master/<int:mid>/documents',methods=['POST'])
@admin_required
def master_document_upload(mid):
    owner_type='landlord' if request.path.startswith('/agreements/landlords/') else 'tenant'
    owner=db.session.get(_master_model(owner_type),mid) or abort(404)
    try:
        doc=_new_master_document(owner_type,owner.id,request.files.get('document'))
        db.session.add(doc);db.session.flush();record_audit('master_document_uploaded','master_document',doc.id,module='agreement_master',meta={'owner_type':owner_type,'owner_id':owner.id,'category':doc.category});db.session.commit();flash('Document encrypted and stored.','success')
    except Exception as exc:
        db.session.rollback();flash(str(exc),'danger')
    return redirect(url_for('landlord_master_edit',mid=mid) if owner_type=='landlord' else url_for('tenant_master_edit',mid=mid))

@app.route('/agreement-master-documents/<int:did>/download',methods=['POST'])
@admin_required
def master_document_download(did):
    doc=db.session.get(MasterDocument,did) or abort(404);_require_admin_password_from_form()
    try:
        raw=decrypt_blob(doc.ciphertext,doc.nonce,_master_key())
    except Exception:
        record_audit('master_document_download_failed','master_document',doc.id,status='failed',module='agreement_master');db.session.commit();abort(500)
    record_audit('master_document_downloaded','master_document',doc.id,module='agreement_master',meta={'owner_type':doc.owner_type});db.session.commit()
    safe_name=f"{doc.owner_type}-{doc.category}-{doc.id}{doc.extension}"
    response=send_file(io.BytesIO(raw),mimetype=doc.mime_type,as_attachment=True,download_name=safe_name)
    response.headers['Cache-Control']='no-store, private';response.headers['Pragma']='no-cache'
    return response

@app.route('/agreement-master-documents/<int:did>/replace',methods=['POST'])
@admin_required
def master_document_replace(did):
    old=db.session.get(MasterDocument,did) or abort(404);_require_admin_password_from_form();owner=_master_document_owner(old) or abort(404)
    try:
        new=_new_master_document(old.owner_type,owner.id,request.files.get('document'),source_doc=old)
        old.active=False;db.session.add(new);db.session.flush();record_audit('master_document_replaced','master_document',new.id,module='agreement_master',meta={'old_document_id':old.id,'owner_type':old.owner_type,'owner_id':owner.id});db.session.commit();flash('New encrypted document version stored; the historical version was retained.','success')
    except Exception as exc:
        db.session.rollback();flash(str(exc),'danger')
    return redirect(_master_document_redirect(old))

@app.route('/agreement-master-documents/<int:did>/delete',methods=['POST'])
@admin_required
def master_document_delete(did):
    doc=db.session.get(MasterDocument,did) or abort(404);doc.active=False;record_audit('master_document_deactivated','master_document',doc.id,module='agreement_master',meta={'owner_type':doc.owner_type});db.session.commit();flash('Document deactivated. Historical encrypted bytes were retained for saved agreement references.','success');return redirect(_master_document_redirect(doc))

@app.route('/agreement-masters/<kind>/<int:mid>/reveal',methods=['POST'])
@admin_required
def master_sensitive_reveal(kind,mid):
    if kind not in ('landlord','tenant'): abort(404)
    row=db.session.get(_master_model(kind),mid) or abort(404)
    body=request.get_json(silent=True) or {}
    admin=current_user();supplied=str(body.get('admin_password') or '')
    if not check_password_hash(admin.password_hash,supplied):
        record_audit('master_sensitive_reveal_failed',f'{kind}_master',mid,status='failed',module='agreement_master');db.session.commit();abort(403)
    requested=body.get('fields') if isinstance(body.get('fields'),list) else []
    allowed_requested_fields=[str(field) for field in requested if str(field) in SENSITIVE_FIELDS[kind]]
    payload=_master_payload(row)
    values={field:payload.get(field,'') for field in allowed_requested_fields}
    record_audit('master_sensitive_fields_revealed',f'{kind}_master',mid,module='agreement_master',meta={'fields':allowed_requested_fields});db.session.commit()
    response=jsonify(ok=True,fields=values)
    response.headers['Cache-Control']='no-store, private';response.headers['Pragma']='no-cache'
    return response


@app.route('/api/agreement-masters/<kind>/<int:mid>/documents-for-annexure')
@admin_required
def master_documents_for_annexure(kind,mid):
    if kind not in ('landlord','tenant'): abort(404)
    owner=db.session.get(_master_model(kind),mid) or abort(404)
    filters={'owner_type':kind,'active':True}
    filters['landlord_master_id' if kind=='landlord' else 'tenant_master_id']=owner.id
    docs=MasterDocument.query.filter_by(**filters).order_by(MasterDocument.category,MasterDocument.uploaded_at.desc()).all()
    return jsonify(ok=True,documents=[{
        'id':doc.id,'category':doc.category,'display_label':doc.display_label,
        'extension':doc.extension,'verification_status':doc.verification_status,'active':bool(doc.active),
        'embeddable':doc.extension.lower() in {'.pdf','.jpg','.jpeg','.png','.webp'},
    } for doc in docs])

@app.route('/agreements')
@permission_required('agreements')
def agreements(): return render_template('agreements.html', items=Agreement.query.order_by(Agreement.updated_at.desc()).all())


def _validated_annexure_ids_for_data(data):
    wanted=parse_annexure_ids((data or {}).get('annexure_document_ids',''))
    landlord_id=str((data or {}).get('landlord_master_id') or '').strip()
    tenant_id=str((data or {}).get('tenant_master_id') or '').strip()
    try: landlord_id=int(landlord_id) if landlord_id else 0
    except ValueError: landlord_id=0
    try: tenant_id=int(tenant_id) if tenant_id else 0
    except ValueError: tenant_id=0
    valid=[]
    for did in wanted:
        doc=db.session.get(MasterDocument,did)
        if not doc: continue
        if doc.owner_type=='landlord' and landlord_id and doc.landlord_master_id==landlord_id: valid.append(did)
        elif doc.owner_type=='tenant' and tenant_id and doc.tenant_master_id==tenant_id: valid.append(did)
    return valid

@app.route('/agreements/new', methods=['GET','POST'])
@app.route('/agreements/<int:aid>/edit', methods=['GET','POST'])
@permission_required('agreements')
def agreement_edit(aid=None):
    ag=db.session.get(Agreement,aid) if aid else None
    if request.method=='POST':
        preset=request.form.get('agreement_template') or 'Strong Residential - 11 Months'
        before_annexures=parse_annexure_ids((ag.data if ag else {}).get('annexure_document_ids',''))
        d=all_form_data(preset)
        valid_annexures=_validated_annexure_ids_for_data(d)
        d['annexure_document_ids']=','.join(str(item) for item in valid_annexures)
        if not ag:
            ag=Agreement(name='Agreement',preset=preset,data_json='{}'); db.session.add(ag); db.session.flush()
        ag.preset=preset; ag.data_json=json.dumps(d,ensure_ascii=False)
        ag.name=f"{d.get('tenant_name') or 'Agreement'} - {d.get('room_unit_no') or d.get('property_name') or ag.id}"
        if before_annexures!=valid_annexures:
            record_audit('agreement_annexures_updated','agreement',ag.id,module='agreement_master',meta={'document_ids':valid_annexures})
        db.session.commit(); sync_tenant_from_agreement(ag)
        flash(f'{preset} agreement saved. Preset-specific format and clauses have been applied.','success')
        return redirect(url_for('agreement_preview',aid=ag.id,lang=request.form.get('save_lang','en')))
    return render_template('agreement_edit.html', **agreement_editor_context(ag))

@app.route('/agreements/<int:aid>/preview')
@permission_required('agreements')
def agreement_preview(aid):
    ag=db.session.get(Agreement,aid) or abort(404); lang=request.args.get('lang','en')
    text=build_agreement_text_hindi(ag.data,[]) if lang=='hi' else build_agreement_text(ag.data,[])
    return render_template('agreement_preview.html',ag=ag,text=text,lang=lang)

def build_agreement_pdf_bytes(ag):
    from reportlab.lib.pagesizes import A4, legal, letter
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
    d=ag.data
    ps={'A4':A4,'Legal':legal,'Letter':letter}.get(d.get('paper_size'),legal)
    def num(key,default):
        try:return float(d.get(key) or default)
        except:return float(default)
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=ps,leftMargin=num('margin_left_mm',20)*mm,rightMargin=num('margin_right_mm',18)*mm,topMargin=num('margin_top_mm',16)*mm,bottomMargin=num('margin_bottom_mm',16)*mm)
    styles=getSampleStyleSheet()
    body=ParagraphStyle('AgreementBody',parent=styles['BodyText'],fontName='Times-Roman',fontSize=9.5,leading=14,spaceAfter=4)
    head=ParagraphStyle('AgreementHead',parent=styles['Heading3'],fontName='Times-Bold',fontSize=11,leading=14,spaceBefore=8,spaceAfter=5)
    title=ParagraphStyle('AgreementTitle',parent=styles['Title'],fontName='Times-Bold',fontSize=15,leading=18,alignment=TA_CENTER,spaceAfter=8)
    story=[]
    logo=os.path.join(BASE_DIR,'static','livenza_logo.png')
    if os.path.exists(logo):
        story += [Image(logo,width=38*mm,height=24*mm),Spacer(1,3*mm)]
    text=build_agreement_text(d,[])
    for line in text.splitlines():
        line=line.strip()
        if not line:
            story.append(Spacer(1,2.5*mm)); continue
        escaped=html.escape(line)
        if line.isupper() and len(line)<150:
            story.append(Paragraph(escaped, title if len(story)<8 else head))
        else:
            story.append(Paragraph(escaped,body))
    doc.build(story); buf.seek(0); return buf


def _annexure_reference_page(reference_lines):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer
    from reportlab.lib.units import mm
    buf=io.BytesIO();styles=getSampleStyleSheet()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=18*mm,rightMargin=18*mm,topMargin=18*mm,bottomMargin=18*mm)
    story=[Paragraph('ANNEXURE REFERENCES',styles['Heading2']),Spacer(1,5*mm),Paragraph('The following selected master documents are retained as protected references but were not embedded because their format could not be converted safely.',styles['BodyText']),Spacer(1,5*mm)]
    for line in reference_lines:
        story.append(Paragraph(html.escape(str(line)),styles['BodyText']));story.append(Spacer(1,2*mm))
    doc.build(story);buf.seek(0);return buf

def _image_bytes_as_pdf(raw):
    image=PILImage.open(io.BytesIO(raw));image=ImageOps.exif_transpose(image)
    if image.mode!='RGB': image=image.convert('RGB')
    out=io.BytesIO();image.save(out,format='PDF',resolution=150.0);out.seek(0);return out

def _append_master_annexures(base_pdf,agreement_data):
    from pypdf import PdfReader,PdfWriter
    data=dict(agreement_data or {})
    selected=_validated_annexure_ids_for_data(data)
    if not selected:
        base_pdf.seek(0);return base_pdf
    writer=PdfWriter();base_pdf.seek(0)
    for page in PdfReader(base_pdf).pages: writer.add_page(page)
    references=[];embedded=[]
    for did in selected:
        doc=db.session.get(MasterDocument,did)
        if not doc: continue
        label=f'{doc.display_label or DOCUMENT_CATEGORIES.get(doc.category,"Supporting document")} ({doc.extension.lower()})'
        ext=(doc.extension or '').lower()
        if ext not in {'.pdf','.jpg','.jpeg','.png','.webp'}:
            references.append(label+' — reference only; format not embeddable.')
            continue
        try:
            raw=decrypt_blob(doc.ciphertext,doc.nonce,_master_key())
            annexure_stream=io.BytesIO(raw) if ext=='.pdf' else _image_bytes_as_pdf(raw)
            annexure_stream.seek(0)
            annexure_reader=PdfReader(annexure_stream)
            if not annexure_reader.pages: raise ValueError('No printable pages')
            for page in annexure_reader.pages: writer.add_page(page)
            embedded.append(did)
        except Exception:
            references.append(label+' — reference retained; printable conversion unavailable.')
    if references:
        ref_stream=_annexure_reference_page(references)
        for page in PdfReader(ref_stream).pages: writer.add_page(page)
    output=io.BytesIO();writer.write(output);output.seek(0)
    return output

@app.route('/agreements/<int:aid>/pdf-with-annexures',methods=['POST'])
@admin_required
def agreement_pdf_with_annexures(aid):
    ag=db.session.get(Agreement,aid) or abort(404)
    _require_admin_password_from_form()
    try:
        output=_append_master_annexures(build_agreement_pdf_bytes(ag),ag.data)
    except Exception:
        record_audit('agreement_annexure_pdf_failed','agreement',ag.id,status='failed',module='agreement_master');db.session.commit();abort(500)
    ids=_validated_annexure_ids_for_data(ag.data)
    record_audit('agreement_annexure_pdf_downloaded','agreement',ag.id,module='agreement_master',meta={'document_ids':ids});db.session.commit()
    response=send_file(output,mimetype='application/pdf',as_attachment=True,download_name=f'Livenza_Agreement_{ag.id}_with_annexures.pdf')
    response.headers['Cache-Control']='no-store, private';response.headers['Pragma']='no-cache'
    return response

@app.route('/agreements/<int:aid>/pdf')
@permission_required('agreements')
def agreement_pdf(aid):
    ag=db.session.get(Agreement,aid) or abort(404)
    return send_file(build_agreement_pdf_bytes(ag),mimetype='application/pdf',as_attachment=True,download_name=f"Livenza_Agreement_{ag.id}.pdf")

@app.route('/share/agreement/<token>.pdf')
def shared_agreement_pdf(token):
    try:
        payload=share_serializer().loads(token,max_age=7*24*3600)
        aid=int(payload.get('aid'))
    except (BadSignature,SignatureExpired,ValueError,TypeError):
        abort(404)
    ag=db.session.get(Agreement,aid) or abort(404)
    return send_file(build_agreement_pdf_bytes(ag),mimetype='application/pdf',as_attachment=True,download_name=f"Livenza_Agreement_{ag.id}.pdf")

@app.route('/agreements/<int:aid>/whatsapp')
@permission_required('agreements')
def agreement_whatsapp(aid):
    ag=db.session.get(Agreement,aid) or abort(404)
    d=ag.data; number=normalize_whatsapp_number(d.get('tenant_whatsapp') or d.get('tenant_mobile'))
    if not number:
        flash('A valid tenant WhatsApp number with country code is required before sharing the agreement.','danger')
        return redirect(url_for('agreement_edit',aid=aid))
    token=share_serializer().dumps({'aid':aid})
    share_url=url_for('shared_agreement_pdf',token=token,_external=True,_scheme='https')
    tenant=d.get('tenant_name') or 'Customer'
    message=f"Dear {tenant}, your Livenza agreement is ready. Download the PDF here: {share_url}\n\nThis secure link is valid for 7 days."
    return redirect(f"https://wa.me/{number}?text={urllib.parse.quote(message)}")

@app.route('/agreements/<int:aid>/delete', methods=['POST'])
@permission_required('agreements')
def agreement_delete(aid):
    ag=db.session.get(Agreement,aid) or abort(404)
    Tenant.query.filter_by(agreement_id=aid).update({'agreement_id':None}); db.session.delete(ag); db.session.commit(); flash('Agreement deleted.','success')
    return redirect(url_for('agreements'))

@app.route('/rooms', methods=['GET','POST'])
@permission_required('rooms')
def rooms():
    if request.method=='POST':
        rid=request.form.get('id'); r=db.session.get(Room,int(rid)) if rid else Room()
        if not rid: db.session.add(r)
        for k in ('city','property_name','premises','room_no','room_type','capacity','standard_tariff','status_override','notes'):
            setattr(r,k,request.form.get(k,'').strip())
        db.session.commit(); flash('Room saved.','success'); return redirect(url_for('rooms'))
    allrooms=Room.query.order_by(Room.city,Room.property_name,Room.room_no).all()
    room_rows=[]
    for r in allrooms:
        st=room_status(r); tenant=None
        if st!='Vacant': tenant=Tenant.query.filter_by(property_name=r.property_name,room_unit_no=r.room_no).order_by(Tenant.id.desc()).first()
        room_rows.append((r,st,tenant))
    return render_template('rooms.html', rows=room_rows, cities=City.query.filter_by(active=True).order_by(City.name).all())

@app.route('/tenants', methods=['GET','POST'])
@permission_required('rooms')
def tenants():
    if request.method=='POST':
        tid=request.form.get('id'); t=db.session.get(Tenant,int(tid)) if tid else Tenant()
        if not tid: db.session.add(t)
        for k in ('tenant_name','tenant_mobile','tenant_whatsapp','tenant_email','tenant_id_type','tenant_id_no','city','property_name','premises','room_unit_no','room_type','tariff','security_deposit','joining_date','leaving_date','status','agreement_reference','notes'):
            setattr(t,k,request.form.get(k,'').strip())
        if t.room_unit_no and not Room.query.filter_by(property_name=t.property_name,room_no=t.room_unit_no).first():
            db.session.add(Room(city=t.city,property_name=t.property_name,premises=t.premises,room_no=t.room_unit_no,room_type=t.room_type,standard_tariff=t.tariff))
        db.session.commit(); flash('Tenant saved.','success'); return redirect(url_for('tenants'))
    return render_template('tenants.html',items=Tenant.query.order_by(Tenant.updated_at.desc()).all(),cities=City.query.filter_by(active=True).order_by(City.name).all())

@app.route('/date-calculator', methods=['POST'])
@permission_required('agreements')
def date_calc():
    start=parse_date(request.form.get('start_date')); months=int(request.form.get('months') or 0); days=int(request.form.get('days') or 0)
    if not start: return jsonify(error='Invalid date'),400
    end=start+relativedelta(months=months)+datetime.timedelta(days=days)-datetime.timedelta(days=1)
    return jsonify(end_date=end.isoformat(), start_weekday=start.strftime('%A'), end_weekday=end.strftime('%A'), total_days=(end-start).days+1)

@app.route('/rooms/empty-report.pdf')
@permission_required('rooms')
def empty_rooms_pdf():
    return send_file(generate_empty_rooms_pdf_bytes(),mimetype='application/pdf',as_attachment=True,download_name=f"Livenza_Empty_Rooms_{datetime.date.today().isoformat()}.pdf")

@app.route('/jobs/vacant-room-report', methods=['GET','POST'])
def vacant_room_job():
    token=request.headers.get('X-Livenza-Job-Token') or request.args.get('token','')
    if not os.getenv('VACANT_REPORT_JOB_TOKEN') or token!=os.getenv('VACANT_REPORT_JOB_TOKEN'): abort(401)
    if setting('vacant_report_enabled','0')!='1': return jsonify(ok=True,skipped='disabled')
    now=datetime.datetime.now(ZoneInfo('Asia/Kolkata')); configured=setting('vacant_report_time','08:30')
    try:
        hh,mm=[int(x) for x in configured.split(':')[:2]]
    except Exception:
        hh,mm=8,30
    now_minutes=now.hour*60+now.minute; target_minutes=hh*60+mm
    if abs(now_minutes-target_minutes)>10:
        return jsonify(ok=True,skipped='not scheduled window',now=now.strftime('%H:%M'),scheduled=configured)
    if setting('vacant_report_last_sent','')==now.date().isoformat(): return jsonify(ok=True,skipped='already sent today')
    recipients=[x.strip() for x in setting('vacant_report_recipients','').split(',') if x.strip()]
    if not recipients: return jsonify(ok=False,error='No recipients configured'),400
    pdf=generate_empty_rooms_pdf_bytes(); filename=f'Livenza_Vacant_Rooms_{now.date().isoformat()}.pdf'; media_id=whatsapp_upload_pdf(pdf,filename)
    if not media_id: return jsonify(ok=False,error='WhatsApp Cloud API not configured or media upload failed'),503
    sent=[]
    for num in recipients:
        if whatsapp_send_document(num,media_id,filename,'Livenza Life daily vacant rooms status'): sent.append(num)
    if sent: set_setting('vacant_report_last_sent',now.date().isoformat())
    return jsonify(ok=True,sent=sent,total=len(recipients))


def offline_review(business, business_type, experience, tone, language):
    base=experience.strip() or f"The experience with {business or business_type or 'the business'} was smooth and professionally managed."
    if language=='Hindi':
        return [f"{business or 'यह स्थान'} में मेरा अनुभव अच्छा रहा। {base} सेवा सुव्यवस्थित थी और पूरी प्रक्रिया सहज लगी।", f"{business or 'यह व्यवसाय'} के साथ अनुभव संतोषजनक रहा। {base} टीम का व्यवहार सहयोगी और पेशेवर लगा।", f"कुल मिलाकर {business or 'यहाँ'} का अनुभव सकारात्मक रहा। {base} मैं वास्तविक अनुभव के आधार पर इसे अच्छा विकल्प मानता/मानती हूँ।"]
    return [f"I had a positive experience with {business or business_type}. {base} The process felt organized and professional throughout.", f"A smooth and dependable experience with {business or business_type}. {base} Communication was clear and the overall service was well managed.", f"Overall, my experience with {business or business_type} was very good. {base} I appreciated the professional approach and would consider using the service again."]


def normalize_google_review_url(value):
    value=(value or '').strip()
    if not value:
        return ''
    if not value.lower().startswith(('http://','https://')):
        value='https://' + value
    try:
        u=urllib.parse.urlparse(value)
    except Exception:
        return ''
    host=(u.hostname or '').lower()
    allowed=(
        host == 'g.page' or host.endswith('.g.page') or
        host == 'google.com' or host.endswith('.google.com') or
        host == 'google.co.in' or host.endswith('.google.co.in') or
        host == 'maps.app.goo.gl' or host.endswith('.maps.app.goo.gl')
    )
    if u.scheme not in ('http','https') or not host or not allowed:
        return ''
    return urllib.parse.urlunparse((u.scheme,u.netloc,u.path,u.params,u.query,u.fragment))

@app.route('/reviews/qr.png')
@permission_required('reviews')
def review_qr():
    review_url=normalize_google_review_url(request.args.get('url',''))
    if not review_url:
        abort(400, 'A valid Google Review URL is required.')
    qr=qrcode.QRCode(version=None,error_correction=qrcode.constants.ERROR_CORRECT_M,box_size=10,border=3)
    qr.add_data(review_url); qr.make(fit=True)
    img=qr.make_image(fill_color='black',back_color='white')
    buf=io.BytesIO(); img.save(buf,format='PNG'); buf.seek(0)
    return send_file(buf,mimetype='image/png',download_name='Livenza_Google_Review_QR.png')

def online_reviews(form):
    key=os.getenv('OPENAI_API_KEY','').strip()
    if not key: return None
    from openai import OpenAI
    client=OpenAI(api_key=key)
    prompt=f"""Create exactly 3 distinct Google review drafts based ONLY on the genuine experience details below. Do not invent amenities, staff names, prices, dates, outcomes, or experiences not provided. Make them natural, useful, non-spammy, and different from each other. Business: {form.get('business_name','')}. Industry: {form.get('business_type','')}. Genuine experience: {form.get('experience','')}. Rating context: {form.get('rating','5')}/5. Tone: {form.get('tone','Professional')}. Vocabulary: {form.get('vocabulary','Natural')}. Language: {form.get('language','English')}. Length: {form.get('length','Medium')}. Return only the three reviews separated by a line containing exactly ---REVIEW---."""
    resp=client.responses.create(model=os.getenv('OPENAI_REVIEW_MODEL','gpt-5.6-luna'),input=prompt)
    parts=[p.strip() for p in resp.output_text.split('---REVIEW---') if p.strip()]
    return parts[:3] if parts else None

@app.route('/reviews', methods=['GET','POST'])
@permission_required('reviews')
def reviews():
    generated=[]
    review_url=setting('default_google_review_url','')
    form_data={}
    if request.method=='POST':
        form=request.form.to_dict(); form_data=form
        submitted_url=(form.get('google_review_url') or '').strip()
        review_url=normalize_google_review_url(submitted_url)
        if not review_url:
            flash('Google Business Review Link is mandatory. Paste the direct review-request link before generating reviews or QR codes.','danger')
            hist=Review.query.order_by(Review.created_at.desc()).limit(20).all()
            return render_template('reviews.html',generated=[],history=hist,review_url='',form_data=form_data),400
        if form.get('save_review_link')=='1':
            set_setting('default_google_review_url',review_url)
        try:
            generated=online_reviews(form) or offline_review(form.get('business_name',''),form.get('business_type',''),form.get('experience',''),form.get('tone',''),form.get('language','English'))
        except Exception as e:
            flash(f'Online AI unavailable; offline drafting used. ({e})','warning')
            generated=offline_review(form.get('business_name',''),form.get('business_type',''),form.get('experience',''),form.get('tone',''),form.get('language','English'))
        for out in generated:
            db.session.add(Review(business_name=form.get('business_name',''),business_type=form.get('business_type',''),experience=form.get('experience',''),output=out,language=form.get('language','English'),google_review_url=review_url))
        db.session.commit()
    hist=Review.query.order_by(Review.created_at.desc()).limit(20).all()
    return render_template('reviews.html',generated=generated,history=hist,review_url=review_url,form_data=form_data)

@app.route('/food', methods=['GET','POST'])
@permission_required('food')
def food():
    ensure_default_food_integrations()
    if request.method=='POST':
        f=request.form
        gross=_food_float(f.get('gross')); commission=_food_float(f.get('commission')); fees=_food_float(f.get('fees')); taxes=_food_float(f.get('taxes'))
        net=_food_float(f.get('net')) or (gross-commission-fees-taxes)
        db.session.add(FoodOrder(platform=f.get('platform','Direct'),order_id=f.get('order_id',''),outlet=f.get('outlet',''),customer=f.get('customer',''),order_time=f.get('order_time',''),status=f.get('status','New'),payment_mode=f.get('payment_mode',''),gross=gross,commission=commission,fees=fees,taxes=taxes,net=net,settlement_status=f.get('settlement_status','Pending')))
        db.session.commit(); flash('Order saved.','success'); return redirect(url_for('food'))
    items=FoodOrder.query.order_by(FoodOrder.created_at.desc()).all(); sums=db.session.query(func.sum(FoodOrder.gross),func.sum(FoodOrder.net)).first()
    integrations=FoodIntegration.query.filter_by(active=True).order_by(FoodIntegration.platform,FoodIntegration.display_name).all()
    return render_template('food.html',items=items,total_gross=sums[0] or 0,total_net=sums[1] or 0,integrations=integrations)

@app.route('/food/integrations')
@permission_required('food')
def food_integrations():
    return redirect(url_for('integrations_center',category='food',workflow='food_connections'))
    rows=FoodIntegration.query.order_by(FoodIntegration.platform,FoodIntegration.display_name,FoodIntegration.id).all()
    return render_template('food_integrations.html',integrations=rows,webhook_token_configured=bool(setting('food_webhook_token','')),official=OFFICIAL_FOOD_PORTALS)

@app.route('/food/integrations/save',methods=['POST'])
@admin_required
def food_integration_save():
    iid=request.form.get('id','').strip(); row=db.session.get(FoodIntegration,int(iid)) if iid.isdigit() else FoodIntegration()
    if not iid: db.session.add(row)
    row.platform=(request.form.get('platform') or 'Other').strip()[:60]
    row.display_name=(request.form.get('display_name') or f'{row.platform} Integration').strip()[:160]
    row.outlet_id=(request.form.get('outlet_id') or '').strip()[:180]
    row.account_identifier=(request.form.get('account_identifier') or '').strip()[:180]
    row.portal_url=(request.form.get('portal_url') or '').strip()
    row.developer_url=(request.form.get('developer_url') or '').strip()
    row.api_base_url=(request.form.get('api_base_url') or '').strip()
    row.api_token_env=(request.form.get('api_token_env') or '').strip()[:120]
    row.api_key_env=(request.form.get('api_key_env') or '').strip()[:120]
    row.webhook_enabled=request.form.get('webhook_enabled')=='1'; row.api_enabled=request.form.get('api_enabled')=='1'; row.active=request.form.get('active')=='1'
    for attr in ('portal_url','developer_url','api_base_url'):
        val=getattr(row,attr) or ''
        if val and not re.match(r'^https?://',val,re.I):
            flash(f'{attr.replace("_"," ").title()} must start with https:// or http://','danger');return redirect(url_for('integrations_center',category='food',workflow='food_connections'))
    db.session.commit();flash('Food partner integration saved.','success');return redirect(url_for('integrations_center',category='food',workflow='food_connections'))

@app.route('/food/integrations/<int:iid>/delete',methods=['POST'])
@admin_required
def food_integration_delete(iid):
    row=db.session.get(FoodIntegration,iid) or abort(404);db.session.delete(row);db.session.commit();flash('Integration removed.','success');return redirect(url_for('integrations_center',category='food',workflow='food_connections'))

@app.route('/food/integrations/<int:iid>/sync',methods=['POST'])
@permission_required('food')
def food_integration_sync(iid):
    row=db.session.get(FoodIntegration,iid) or abort(404)
    if not row.active or not row.api_enabled or not (row.api_base_url or '').strip():
        flash('Enable API Sync and add the official/API endpoint supplied by the platform first.','warning');return redirect(url_for('integrations_center',category='food',workflow='food_connections'))
    headers={'Accept':'application/json','User-Agent':'LivenzaLife-OperationsCloud/1.5.13'}
    bearer=os.getenv((row.api_token_env or '').strip(),'').strip() if row.api_token_env else ''
    api_key=os.getenv((row.api_key_env or '').strip(),'').strip() if row.api_key_env else ''
    if bearer: headers['Authorization']=f'Bearer {bearer}'
    if api_key: headers['X-API-Key']=api_key
    try:
        resp=requests.get(row.api_base_url,headers=headers,timeout=35);resp.raise_for_status();payload=resp.json();count=_ingest_food_payload(row.platform,payload,default_outlet=row.display_name or row.outlet_id)
        row.last_sync_at=datetime.datetime.utcnow();row.last_sync_count=count;row.last_sync_status=f'OK • {count} record(s) received';db.session.commit();flash(f'{row.platform} sync completed: {count} order record(s).','success')
    except Exception as e:
        db.session.rollback();row=db.session.get(FoodIntegration,iid);row.last_sync_at=datetime.datetime.utcnow();row.last_sync_count=0;row.last_sync_status=f'ERROR • {str(e)[:300]}';db.session.commit();flash(f'{row.platform} API sync failed. Check the endpoint, partner access and Render environment credentials.','danger')
    return redirect(url_for('integrations_center',category='food',workflow='food_connections'))

@app.route('/food/portals')
@permission_required('food')
def food_portals():
    return redirect(url_for('integrations_center',category='food',workflow='food_portals'))
    selected_id=request.args.get('id','');selected=None
    if selected_id.isdigit(): selected=db.session.get(FoodIntegration,int(selected_id))
    if not selected and rows: selected=rows[0]
    return render_template('food_portals.html',integrations=rows,selected=selected)

@app.route('/food/import', methods=['POST'])
@permission_required('food')
def food_import():
    file=request.files.get('file')
    if not file: flash('Choose a CSV file.','danger'); return redirect(url_for('food'))
    text=io.TextIOWrapper(file.stream,encoding='utf-8-sig'); reader=csv.DictReader(text); n=0
    for row in reader:
        _upsert_food_order(row.get('platform','Direct'),row,default_outlet=row.get('outlet','')); n+=1
    db.session.commit(); flash(f'Imported {n} orders.','success'); return redirect(url_for('food'))

@app.route('/webhooks/food/<platform>', methods=['POST'])
def food_webhook(platform):
    token=request.headers.get('X-Livenza-Webhook-Token','')
    configured=setting('food_webhook_token','')
    if not configured or not secrets.compare_digest(token,configured): abort(401)
    payload=request.get_json(silent=True)
    if payload is None: return jsonify(ok=False,error='JSON payload required'),400
    integration=FoodIntegration.query.filter(func.lower(FoodIntegration.platform)==platform.lower(),FoodIntegration.active.is_(True)).first()
    if integration and not integration.webhook_enabled: abort(403)
    display=(integration.platform if integration else platform.replace('-',' ').title())
    count=_ingest_food_payload(display,payload,default_outlet=(integration.display_name if integration else ''))
    if integration:
        integration.last_sync_at=datetime.datetime.utcnow();integration.last_sync_count=count;integration.last_sync_status=f'WEBHOOK • {count} record(s)';db.session.commit()
    return jsonify(ok=True,platform=display,records=count)

@app.route('/queries', methods=['GET','POST'])
@permission_required('queries')
def queries():
    if request.method=='POST':
        f=request.form; q=QueryLead(source=f.get('source','Manual'),city=f.get('city',''),property_name=f.get('property_name',''),customer_name=f.get('customer_name',''),mobile=f.get('mobile',''),whatsapp=f.get('whatsapp','') or f.get('mobile',''),email=f.get('email',''),query_text=f.get('query_text',''),budget=f.get('budget',''),move_in_date=f.get('move_in_date',''),stay_type=f.get('stay_type',''),status=f.get('status','Live'),heat=f.get('heat','Warm'),score=int(f.get('score') or 50),next_follow_up=f.get('next_follow_up',''),notes=f.get('notes',''))
        au=f.get('assigned_user_id'); q.assigned_user_id=int(au) if au and au.isdigit() else None
        db.session.add(q); db.session.flush(); query_log(q,'Created','Manual query created',current_user()); db.session.commit(); flash('Query added to live queue.','success'); return redirect(url_for('queries'))
    status=request.args.get('status',''); heat=request.args.get('heat',''); source=request.args.get('source',''); term=request.args.get('q','').strip()
    qry=QueryLead.query
    if status: qry=qry.filter_by(status=status)
    if heat: qry=qry.filter_by(heat=heat)
    if source: qry=qry.filter_by(source=source)
    if term: qry=qry.filter(or_(QueryLead.customer_name.ilike(f'%{term}%'),QueryLead.mobile.ilike(f'%{term}%'),QueryLead.query_text.ilike(f'%{term}%'),QueryLead.property_name.ilike(f'%{term}%')))
    items=qry.order_by(QueryLead.updated_at.desc()).all()
    return render_template('queries.html',items=items,templates=QueryTemplate.query.filter_by(active=True).order_by(QueryTemplate.name).all(),users=User.query.filter_by(active=True).order_by(User.full_name,User.username).all(),cities=City.query.filter_by(active=True).order_by(City.name).all(),cloud_whatsapp=whatsapp_cloud_configured())


@app.route('/queries/sheet')
@permission_required('queries')
def query_sheet():
    term=request.args.get('q','').strip()
    qry=QueryLead.query
    if term:
        qry=qry.filter(or_(
            QueryLead.customer_name.ilike(f'%{term}%'),
            QueryLead.mobile.ilike(f'%{term}%'),
            QueryLead.whatsapp.ilike(f'%{term}%'),
            QueryLead.city.ilike(f'%{term}%'),
            QueryLead.property_name.ilike(f'%{term}%'),
            QueryLead.query_text.ilike(f'%{term}%')
        ))
    items=qry.order_by(QueryLead.updated_at.desc()).limit(750).all()
    return render_template('query_sheet.html',items=items,blank_rows=30,cities=City.query.filter_by(active=True).order_by(City.name).all())


def _query_sheet_apply(q, payload):
    allowed=('source','city','property_name','customer_name','mobile','whatsapp','email','query_text','budget','move_in_date','stay_type','status','heat','next_follow_up','notes')
    for k in allowed:
        if k in payload:
            setattr(q,k,str(payload.get(k) or '').strip())
    if 'score' in payload:
        try:q.score=max(0,min(100,int(payload.get('score') or 0)))
        except Exception:pass
    q.updated_at=datetime.datetime.utcnow()
    return q

@app.route('/api/queries',methods=['POST'])
@permission_required('queries')
def query_sheet_create():
    payload=request.get_json(silent=True) or {}
    q=QueryLead(source='Manual',status='Live',heat='Warm',score=50)
    _query_sheet_apply(q,payload)
    if not q.whatsapp:q.whatsapp=q.mobile
    db.session.add(q);db.session.flush();query_log(q,'Created','Created from spreadsheet view',current_user());db.session.commit()
    return jsonify(ok=True,id=q.id,updated_at=q.updated_at.isoformat() if q.updated_at else '')

@app.route('/api/queries/<int:qid>',methods=['PATCH'])
@permission_required('queries')
def query_sheet_patch(qid):
    q=db.session.get(QueryLead,qid) or abort(404)
    payload=request.get_json(silent=True) or {}
    _query_sheet_apply(q,payload)
    query_log(q,'Spreadsheet update','Updated from spreadsheet view',current_user());db.session.commit()
    return jsonify(ok=True,id=q.id,updated_at=q.updated_at.isoformat() if q.updated_at else '')

@app.route('/api/queries/batch',methods=['POST'])
@permission_required('queries')
def query_sheet_batch_save():
    body=request.get_json(silent=True) or {}; rows=body.get('rows') if isinstance(body.get('rows'),list) else []
    if not rows: return jsonify(ok=False,error='No query rows were supplied.'),400
    if len(rows)>500: return jsonify(ok=False,error='Save up to 500 rows in one batch.'),400
    saved=[]
    for index,payload in enumerate(rows):
        if not isinstance(payload,dict): continue
        row_id=str(payload.get('id') or '').strip(); q=db.session.get(QueryLead,int(row_id)) if row_id.isdigit() else None
        if row_id and not q: continue
        if not q:
            meaningful=any(str(payload.get(k) or '').strip() for k in ('customer_name','mobile','whatsapp','email','city','property_name','budget','move_in_date','stay_type','query_text','notes'))
            if not meaningful: continue
            q=QueryLead(source='Manual',status='Live',heat='Warm',score=50);db.session.add(q);db.session.flush()
            action='Created from editable query sheet'
        else: action='Batch-updated from editable query sheet'
        _query_sheet_apply(q,payload)
        if not q.whatsapp:q.whatsapp=q.mobile
        query_log(q,'Spreadsheet batch save',action,current_user())
        saved.append({'client_ref':str(payload.get('client_ref') or index),'id':q.id,'updated_at':q.updated_at.isoformat() if q.updated_at else ''})
    db.session.commit()
    return jsonify(ok=True,saved=saved,count=len(saved))


QUERY_IMPORT_ALIASES={
    'source':'source','channel':'source','customer':'customer_name','customer name':'customer_name','name':'customer_name',
    'mobile':'mobile','phone':'mobile','phone number':'mobile','whatsapp':'whatsapp','whatsapp number':'whatsapp',
    'email':'email','email address':'email','city':'city','property':'property_name','property name':'property_name',
    'budget':'budget','move in':'move_in_date','move in date':'move_in_date','move-in':'move_in_date','move-in date':'move_in_date',
    'stay type':'stay_type','status':'status','heat':'heat','lead heat':'heat','score':'score',
    'next follow up':'next_follow_up','next follow-up':'next_follow_up','follow up':'next_follow_up',
    'requirement':'query_text','query':'query_text','requirement query':'query_text','notes':'notes',
}

def _query_import_header(value):
    return re.sub(r'\s+',' ',re.sub(r'[^a-z0-9]+',' ',str(value or '').strip().lower())).strip()

def _query_import_cell(value):
    if value is None:return ''
    if isinstance(value,(datetime.datetime,datetime.date)):return value.isoformat()
    if isinstance(value,float) and value.is_integer():return str(int(value))
    return str(value).strip()

@app.route('/queries/sheet/import',methods=['POST'])
@permission_required('queries')
def query_sheet_import():
    upload=request.files.get('query_file')
    if not upload or not upload.filename:
        flash('Choose an Excel .xlsx or CSV file first.','danger');return redirect(url_for('query_sheet'))
    name=os.path.basename(upload.filename);ext=os.path.splitext(name.lower())[1]
    if ext not in ('.xlsx','.csv'):
        flash('Query import supports .xlsx and .csv files. Save older .xls files as .xlsx first.','danger');return redirect(url_for('query_sheet'))
    try: upload.stream.seek(0,2);upload_size=upload.stream.tell();upload.stream.seek(0)
    except Exception: upload_size=0
    if upload_size>15*1024*1024:
        flash('Query spreadsheet imports must be 15 MB or smaller.','danger');return redirect(url_for('query_sheet'))
    rows=[]
    try:
        if ext=='.csv':
            reader=csv.reader(io.TextIOWrapper(upload.stream,encoding='utf-8-sig',errors='replace'))
            rows=[list(row)[:100] for _,row in zip(range(1002),reader)]
        else:
            from openpyxl import load_workbook
            book=load_workbook(upload.stream,read_only=True,data_only=True)
            sheet=book.active
            rows=[list(row) for row in sheet.iter_rows(min_row=1,max_row=1001,max_col=100,values_only=True)]
            book.close()
    except Exception as exc:
        flash(f'The spreadsheet could not be read: {str(exc)[:180]}','danger');return redirect(url_for('query_sheet'))
    if not rows:
        flash('The spreadsheet is empty.','warning');return redirect(url_for('query_sheet'))
    mapped=[QUERY_IMPORT_ALIASES.get(_query_import_header(value),'') for value in rows[0]]
    if not any(mapped):
        flash('No recognised query columns were found. Use headers such as Customer Name, Mobile, City, Property, Status and Query.','danger');return redirect(url_for('query_sheet'))
    created=0
    for raw in rows[1:1001]:
        payload={field:_query_import_cell(raw[index]) for index,field in enumerate(mapped) if field and index<len(raw)}
        if not any(payload.values()):continue
        q=QueryLead(source='Manual',status='Live',heat='Warm',score=50);_query_sheet_apply(q,payload)
        if not q.whatsapp:q.whatsapp=q.mobile
        db.session.add(q);db.session.flush();query_log(q,'Imported',f'Imported from {name[:180]}',current_user());created+=1
    db.session.commit();flash(f'Imported and saved {created} query row(s) from {name}.','success');return redirect(url_for('query_sheet'))

@app.route('/queries/<int:qid>/update',methods=['POST'])
@permission_required('queries')
def query_update(qid):
    q=db.session.get(QueryLead,qid) or abort(404); f=request.form
    for k in ('source','city','property_name','customer_name','mobile','whatsapp','email','query_text','budget','move_in_date','stay_type','status','heat','next_follow_up','notes'):
        if k in f: setattr(q,k,f.get(k,'').strip())
    try:q.score=max(0,min(100,int(f.get('score') or q.score or 0)))
    except Exception:pass
    au=f.get('assigned_user_id'); q.assigned_user_id=int(au) if au and au.isdigit() else None
    query_log(q,'Updated',f'Status {q.status}; heat {q.heat}',current_user()); db.session.commit(); flash('Query updated.','success'); return redirect(request.referrer or url_for('queries'))

@app.route('/queries/<int:qid>/whatsapp')
@permission_required('queries')
def query_whatsapp(qid):
    q=db.session.get(QueryLead,qid) or abort(404); tid=request.args.get('template'); t=db.session.get(QueryTemplate,int(tid)) if tid and tid.isdigit() else None
    msg=(t.message if t else setting('query_default_message','Dear {name}, thank you for contacting Livenza Life. Our team is reviewing your requirement and will assist you shortly.')).format(name=q.customer_name or 'Guest',property=q.property_name or 'Livenza Life',city=q.city or '')
    number=wa_number(q.whatsapp or q.mobile)
    if not number: flash('Query has no valid WhatsApp number.','danger'); return redirect(url_for('queries'))
    query_log(q,'WhatsApp opened',t.name if t else 'Default message',current_user()); db.session.commit()
    return redirect(f'https://wa.me/{number}?text={urllib.parse.quote(msg)}')

@app.route('/queries/<int:qid>/send-template/<int:tid>',methods=['POST'])
@permission_required('queries')
def query_send_template(qid,tid):
    q=db.session.get(QueryLead,qid) or abort(404); t=db.session.get(QueryTemplate,tid) or abort(404); number=q.whatsapp or q.mobile
    if t.whatsapp_template_name: ok,msg=whatsapp_cloud_template(number,t.whatsapp_template_name)
    else: ok,msg=whatsapp_cloud_text(number,t.message.format(name=q.customer_name or 'Guest',property=q.property_name or 'Livenza Life',city=q.city or ''))
    query_log(q,'WhatsApp template',f'{t.name}: {msg}',current_user()); db.session.commit(); flash(('Message sent.' if ok else msg),('success' if ok else 'warning')); return redirect(url_for('queries'))

@app.route('/queries/export.csv')
@permission_required('queries')
def query_export_csv():
    view=request.args.get('view','all').lower(); qry=QueryLead.query
    if view=='hot': qry=qry.filter_by(heat='Hot')
    elif view=='live': qry=qry.filter(QueryLead.status.in_(['New','Live','Follow-up']))
    rows=qry.order_by(QueryLead.updated_at.desc()).all(); buf=io.StringIO(); w=csv.writer(buf); w.writerow(['ID','Source','City','Property','Customer','Mobile','WhatsApp','Email','Status','Heat','Score','Budget','Move-in','Next Follow-up','Query','Notes'])
    for q in rows:w.writerow([q.id,q.source,q.city,q.property_name,q.customer_name,q.mobile,q.whatsapp,q.email,q.status,q.heat,q.score,q.budget,q.move_in_date,q.next_follow_up,q.query_text,q.notes])
    b=io.BytesIO(buf.getvalue().encode('utf-8-sig')); return send_file(b,mimetype='text/csv',as_attachment=True,download_name=f'Livenza_{view.title()}_Queries_{datetime.date.today()}.csv')

@app.route('/queries/report.pdf')
@permission_required('queries')
def query_report_pdf():
    from reportlab.lib.pagesizes import A4,landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle
    view=request.args.get('view','live').lower(); qry=QueryLead.query
    if view=='hot': qry=qry.filter_by(heat='Hot')
    elif view=='live': qry=qry.filter(QueryLead.status.in_(['New','Live','Follow-up']))
    items=qry.order_by(QueryLead.updated_at.desc()).all(); buf=io.BytesIO(); doc=SimpleDocTemplate(buf,pagesize=landscape(A4),leftMargin=24,rightMargin=24,topMargin=28,bottomMargin=28); styles=getSampleStyleSheet(); story=[Paragraph(f'LIVENZA LIFE - {view.upper()} QUERIES',styles['Title']),Paragraph(datetime.datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%d %b %Y %I:%M %p IST'),styles['Normal']),Spacer(1,10)]; rows=[['ID','Source','City','Customer','Mobile','Property','Status','Heat','Score','Follow-up']]
    for q in items:rows.append([q.id,q.source,q.city,q.customer_name,q.mobile,q.property_name,q.status,q.heat,q.score,q.next_follow_up])
    t=Table(rows,repeatRows=1,colWidths=[28,60,60,95,75,100,55,45,36,85]);t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6d5dfc')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.35,colors.grey),('FONTSIZE',(0,0),(-1,-1),7.5),('VALIGN',(0,0),(-1,-1),'TOP')]));story.append(t);doc.build(story);buf.seek(0);return send_file(buf,mimetype='application/pdf',as_attachment=True,download_name=f'Livenza_{view.title()}_Queries.pdf')

def create_query_from_payload(source,payload):
    d=normalize_query_payload(source,payload); ext=d.get('external_id',''); existing=QueryLead.query.filter_by(source=d['source'],external_id=ext).first() if ext else None
    if existing:return existing,False
    q=QueryLead(**d,raw_json=json.dumps(payload,ensure_ascii=False)[:30000],status='Live',heat='Warm',score=55);db.session.add(q);db.session.flush();query_log(q,'Webhook received',source);auto_reply_for_query(q);db.session.commit();return q,True

@app.route('/webhooks/queries/<source>',methods=['POST'])
def query_webhook(source):
    expected=setting('query_webhook_token','') or os.getenv('QUERY_WEBHOOK_TOKEN',''); supplied=request.headers.get('X-Livenza-Webhook-Token') or request.args.get('token','')
    if not expected or supplied!=expected:abort(401)
    q,created=create_query_from_payload(source,request.get_json(silent=True) or {});return jsonify(ok=True,id=q.id,created=created)

@app.route('/webhooks/google/leads',methods=['POST'])
def google_lead_webhook():
    payload=request.get_json(silent=True) or {}; secret=os.getenv('GOOGLE_LEAD_WEBHOOK_SECRET',''); supplied=str(payload.get('google_key') or request.headers.get('X-Google-Key') or '')
    if secret and supplied!=secret:abort(401)
    q,created=create_query_from_payload('Google',payload);return jsonify(ok=True,id=q.id,created=created)

@app.route('/webhooks/meta/leads',methods=['GET','POST'])
def meta_lead_webhook():
    if request.method=='GET':
        if request.args.get('hub.verify_token')==os.getenv('META_VERIFY_TOKEN',''):return request.args.get('hub.challenge','')
        abort(403)
    payload=request.get_json(silent=True) or {}; token=os.getenv('META_PAGE_ACCESS_TOKEN',''); lead_payload=payload
    try:
        change=payload['entry'][0]['changes'][0]['value']; lead_id=change.get('leadgen_id')
        if lead_id and token:
            ver=os.getenv('META_GRAPH_VERSION','v23.0'); r=requests.get(f'https://graph.facebook.com/{ver}/{lead_id}',params={'access_token':token},timeout=20)
            if r.ok:lead_payload=r.json();lead_payload['leadgen_id']=lead_id
    except Exception:pass
    q,created=create_query_from_payload('Meta',lead_payload);return jsonify(ok=True,id=q.id,created=created)

@app.route('/admin/query-templates/save',methods=['POST'])
@admin_required
def query_template_save():
    tid=request.form.get('id');t=db.session.get(QueryTemplate,int(tid)) if tid else QueryTemplate();
    if not tid:db.session.add(t)
    t.name=request.form.get('name','').strip();t.category=request.form.get('category','General').strip();t.message=request.form.get('message','').strip();t.whatsapp_template_name=request.form.get('whatsapp_template_name','').strip();t.sources_json=json.dumps(request.form.getlist('sources'));t.statuses_json=json.dumps(request.form.getlist('statuses'));t.auto_send=request.form.get('auto_send')=='1';t.active=request.form.get('active')=='1';db.session.commit();flash('Query message template saved.','success');return redirect(url_for('admin_panel')+'#query-templates')

@app.route('/admin/query-templates/<int:tid>/delete',methods=['POST'])
@admin_required
def query_template_delete(tid):
    t=db.session.get(QueryTemplate,tid) or abort(404);db.session.delete(t);db.session.commit();flash('Template deleted.','success');return redirect(url_for('admin_panel')+'#query-templates')

@app.route('/admin/users/<int:uid>/aadhaar-verify',methods=['POST'])
@admin_required
def aadhaar_verify(uid):
    u=db.session.get(User,uid) or abort(404); endpoint=os.getenv('AADHAAR_AUTH_URL','').strip(); token=os.getenv('AADHAAR_AUTH_TOKEN','').strip()
    if not endpoint:
        flash('UIDAI/AUA-KUA authentication provider is not configured. Use UIDAI Paperless Offline e-KYC for offline verification or configure an authorized provider in Render.','warning');return redirect(url_for('admin_panel'))
    aadhaar_or_vid=request.form.get('aadhaar_or_vid','').strip(); otp=request.form.get('otp','').strip()
    try:
        r=requests.post(endpoint,headers={'Authorization':f'Bearer {token}','Content-Type':'application/json'},json={'aadhaar_or_vid':aadhaar_or_vid,'otp':otp,'reference':str(u.id)},timeout=30); data=r.json() if r.headers.get('content-type','').startswith('application/json') else {}; verified=bool(data.get('verified') or data.get('success'))
        u.aadhaar_verification_status='Verified' if verified else 'Verification failed';u.aadhaar_verification_method='Authorized AUA/KUA provider';u.aadhaar_verification_ref=str(data.get('reference') or data.get('txn_id') or '')[:180];u.aadhaar_verified_at=datetime.datetime.utcnow() if verified else None;db.session.commit();flash(('Aadhaar verification confirmed by configured provider.' if verified else 'Provider did not confirm Aadhaar verification.'),('success' if verified else 'danger'))
    except Exception as e:flash(f'Aadhaar provider error: {e}','danger')
    return redirect(url_for('admin_panel'))


@app.route('/video-wall')
@permission_required('video_wall')
def video_wall():
    screens=VideoScreen.query.order_by(VideoScreen.city,VideoScreen.location_name,VideoScreen.name).all()
    assets=VideoAsset.query.filter_by(active=True).order_by(VideoAsset.id.desc()).all()
    festive=active_festive_session()
    rows=[]
    for sc in screens:
        
        try: playlist_ids=[int(x) for x in json.loads(sc.playlist_json or '[]')]
        except Exception: playlist_ids=[]
        rows.append({'screen':sc,'online':screen_is_online(sc),'asset':db.session.get(VideoAsset,sc.current_asset_id) if sc.current_asset_id else None,'playlist_ids':playlist_ids,'player_url':url_for('wall_player',token=sc.player_token,_external=True)})
    return render_template('video_wall.html',screens=rows,assets=assets,festive=festive,cities=City.query.filter_by(active=True).order_by(City.name).all(),storage_ready=supabase_storage_configured(),video_upload_limit_mb=video_wall_upload_limit_mb())

@app.route('/video-wall/assets/resumable/start',methods=['POST'])
@permission_required('video_wall')
def video_wall_resumable_start():
    body=request.get_json(silent=True) or {}
    info,error=create_video_wall_resumable_upload(
        body.get('filename'),body.get('content_type'),body.get('size'),body.get('title'),
    )
    if error: return jsonify(ok=False,error=error),400
    return jsonify(ok=True,upload=info)

@app.route('/video-wall/assets/resumable/finish',methods=['POST'])
@permission_required('video_wall')
def video_wall_resumable_finish():
    body=request.get_json(silent=True) or {}
    asset,error=finalize_video_wall_resumable_upload(body.get('reservation'))
    if error: return jsonify(ok=False,error=error),409
    return jsonify(ok=True,asset={
        'id':asset.id,'title':asset.title,'media_type':asset.media_type,
        'public_url':asset.public_url,'file_size':asset.file_size,
    })

@app.route('/video-wall/assets/upload',methods=['POST'])
@permission_required('video_wall')
def video_wall_asset_upload():
    title=request.form.get('title','').strip()
    external=request.form.get('external_url','').strip()
    if external:
        if not re.match(r'^https?://',external,re.I):
            flash('External media URL must start with http:// or https://','danger'); return redirect(url_for('video_wall'))
        lower=external.lower().split('?')[0]
        mtype='image' if lower.endswith(('.jpg','.jpeg','.png','.webp')) else 'video'
        asset=VideoAsset(title=title or os.path.basename(urllib.parse.urlparse(external).path) or 'External media',media_type=mtype,public_url=external,mime_type='',uploaded_by_user_id=current_user().id)
        db.session.add(asset); db.session.commit(); flash('External media added to the library.','success'); return redirect(url_for('video_wall'))
    f=request.files.get('media_file')
    info,err=upload_video_wall_media(f)
    if err:
        flash(err,'danger'); return redirect(url_for('video_wall'))
    asset=VideoAsset(title=title or f.filename,media_type=info['type'],storage_path=info['path'],public_url=info['url'],mime_type=info['mime'],file_size=info['size'],uploaded_by_user_id=current_user().id)
    db.session.add(asset); db.session.commit(); flash('Media uploaded and ready for screens.','success')
    if info.get('drive_backup_error'): flash('Google Drive mirror warning: '+info['drive_backup_error'],'warning')
    return redirect(url_for('video_wall'))

@app.route('/video-wall/assets/<int:asset_id>/toggle',methods=['POST'])
@permission_required('video_wall')
def video_wall_asset_toggle(asset_id):
    a=db.session.get(VideoAsset,asset_id) or abort(404); a.active=not a.active; db.session.commit(); flash('Media library updated.','success'); return redirect(url_for('video_wall'))

@app.route('/video-wall/screens/save',methods=['POST'])
@permission_required('video_wall')
def video_wall_screen_save():
    sid=request.form.get('id','').strip()
    sc=db.session.get(VideoScreen,int(sid)) if sid.isdigit() else None
    if not sc:
        sc=VideoScreen(name='Screen',player_token=secrets.token_urlsafe(28)); db.session.add(sc)
    sc.name=request.form.get('name','').strip() or sc.name or 'Screen'
    sc.city=request.form.get('city','').strip(); sc.location_name=request.form.get('location_name','').strip(); sc.device_label=request.form.get('device_label','').strip()
    aid=request.form.get('current_asset_id','').strip(); sc.current_asset_id=int(aid) if aid.isdigit() else None
    playlist=[]
    for raw in request.form.getlist('playlist_asset_ids'):
        if str(raw).isdigit() and int(raw) not in playlist: playlist.append(int(raw))
    if not playlist and sc.current_asset_id: playlist=[sc.current_asset_id]
    sc.playlist_json=json.dumps(playlist)
    if playlist: sc.current_asset_id=playlist[0]
    try: sc.rotation_degrees=int(float(request.form.get('rotation_degrees','0') or 0))%360
    except Exception: sc.rotation_degrees=0
    sc.fit_mode=request.form.get('fit_mode','contain') if request.form.get('fit_mode') in ('contain','cover','fill') else 'contain'
    sc.loop_media=request.form.get('loop_media')=='1'; sc.muted=request.form.get('muted')=='1'; sc.enabled=request.form.get('enabled')=='1'
    try: sc.slide_duration_seconds=max(3,min(3600,int(request.form.get('slide_duration_seconds','10') or 10)))
    except Exception: sc.slide_duration_seconds=10
    db.session.commit(); flash('TV / screen configuration saved.','success'); return redirect(url_for('video_wall'))

@app.route('/video-wall/screens/<int:sid>/delete',methods=['POST'])
@permission_required('video_wall')
def video_wall_screen_delete(sid):
    sc=db.session.get(VideoScreen,sid) or abort(404); db.session.delete(sc); db.session.commit(); flash('Screen removed.','success'); return redirect(url_for('video_wall'))

@app.route('/video-wall/festive/start',methods=['POST'])
@permission_required('video_wall')
def video_wall_festive_start():
    aid=request.form.get('asset_id','').strip()
    asset=db.session.get(VideoAsset,int(aid)) if aid.isdigit() else None
    if not asset:
        flash('Choose a festive commercial first.','danger'); return redirect(url_for('video_wall'))
    FestiveSession.query.filter_by(active=True).update({'active':False,'ended_at':datetime.datetime.utcnow()})
    fs=FestiveSession(name=request.form.get('name','').strip() or 'Festive Takeover',asset_id=asset.id,active=True,started_by_user_id=current_user().id,started_at=datetime.datetime.utcnow(),notes=request.form.get('notes','').strip())
    db.session.add(fs); db.session.commit(); flash('Festive takeover is LIVE on every enabled TV.','success'); return redirect(url_for('video_wall'))

@app.route('/video-wall/festive/stop',methods=['POST'])
@permission_required('video_wall')
def video_wall_festive_stop():
    for fs in FestiveSession.query.filter_by(active=True).all(): fs.active=False; fs.ended_at=datetime.datetime.utcnow()
    db.session.commit(); flash('Festive takeover stopped. Screens returned to their individual media.','success'); return redirect(url_for('video_wall'))

@app.route('/wall/<token>')
def wall_player(token):
    sc=VideoScreen.query.filter_by(player_token=token).first() or abort(404)
    return render_template('wall_player.html',screen=sc)

@app.route('/api/wall/<token>/state')
def wall_state(token):
    sc=VideoScreen.query.filter_by(player_token=token).first() or abort(404)
    return jsonify(screen_player_state(sc))

@app.route('/api/wall/<token>/heartbeat',methods=['POST'])
def wall_heartbeat(token):
    sc=VideoScreen.query.filter_by(player_token=token).first() or abort(404)
    sc.last_seen_at=datetime.datetime.utcnow(); sc.last_ip=(request.headers.get('X-Forwarded-For') or request.remote_addr or '')[:120]; db.session.commit()
    return jsonify(ok=True)


# ===== Tesla OS 27 • Landlord / Tenant Master secure storage =====
def _master_key():
    value=os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()
    if not value:
        raise RuntimeError('LIVENZA_VAULT_MASTER_KEY is required for master-profile encryption.')
    return value

def _master_kind_for_row(row):
    return 'landlord' if isinstance(row,LandlordMaster) else 'tenant'

def _master_payload(row):
    if not row or not row.encrypted_payload:
        return {}
    try:
        value=json.loads(decrypt_secret(row.encrypted_payload,row.encrypted_nonce,_master_key()))
        return value if isinstance(value,dict) else {}
    except Exception:
        return {}

def _set_master_payload(row,payload):
    kind=_master_kind_for_row(row)
    normalized=normalize_master_payload(kind,payload)
    row.encrypted_payload,row.encrypted_nonce=encrypt_secret(json.dumps(normalized,ensure_ascii=False),_master_key())
    row.profile_name=(normalized.get('profile_name') or normalized.get('full_legal_name') or 'Unnamed profile')[:180]
    row.party_type=(normalized.get('party_type') or 'individual')[:40]
    row.legal_name=(normalized.get('full_legal_name') or normalized.get('entity_legal_name') or normalized.get('corporate_legal_name') or '')[:220]
    row.primary_mobile=(normalized.get('primary_mobile') or '')[:40]
    row.email=(normalized.get('email') or '')[:220]
    row.city=(normalized.get('city') or '')[:120]
    row.state=(normalized.get('state') or '')[:120]
    row.country=(normalized.get('country') or 'India')[:120]
    row.verification_status=(normalized.get('verification_status') or 'unverified')[:40]
    row.tags=(normalized.get('tags') or '')[:500]
    row.search_text=(normalized.get('search_text') or '')[:12000]
    row.identifier_lookup_json=json.dumps(identifier_lookup_hashes(kind,normalized,_master_key()))
    return normalized

def _new_master_code(kind):
    prefix='LM' if kind=='landlord' else 'TM'
    while True:
        code=f"{prefix}-{datetime.datetime.utcnow():%y%m}-{secrets.token_hex(3).upper()}"
        Model=LandlordMaster if kind=='landlord' else TenantMaster
        if not Model.query.filter_by(master_code=code).first():
            return code

def migrate_legacy_party_profiles():
    counts={'created':0,'updated':0,'failed':0,'skipped':0}
    try:
        _master_key()
    except Exception:
        print('Livenza master migration skipped: LIVENZA_VAULT_MASTER_KEY is not configured.')
        counts['skipped']=AgreementPartyProfile.query.count()
        return counts
    for saved in AgreementPartyProfile.query.order_by(AgreementPartyProfile.id).all():
        if saved.profile_type not in ('landlord','tenant'):
            counts['skipped']+=1
            continue
        Model=LandlordMaster if saved.profile_type=='landlord' else TenantMaster
        try:
            converted=legacy_profile_to_master(saved.profile_type,saved.name,saved.data)
            row=Model.query.filter_by(legacy_profile_id=saved.id).first()
            created=row is None
            if not row:
                row=Model(legacy_profile_id=saved.id,master_code=f"{'LM' if saved.profile_type=='landlord' else 'TM'}-{saved.id:06d}",profile_name=saved.name or 'Legacy profile')
                db.session.add(row)
            _set_master_payload(row,converted)
            row.updated_by_user_id=saved.created_by_user_id
            if created:
                row.created_by_user_id=saved.created_by_user_id
                counts['created']+=1
            else:
                counts['updated']+=1
        except Exception as exc:
            db.session.rollback()
            counts['failed']+=1
            try:
                record_audit('legacy_party_profile_migration_failed','agreement_party_profile',saved.id,status='failed',module='agreement_master',meta={'profile_type':saved.profile_type})
                db.session.commit()
            except Exception:
                db.session.rollback()
            print(f'Livenza master migration skipped legacy profile {saved.id}: {type(exc).__name__}')
            continue
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    return counts

# ===== Tesla OS 27 • Electricity Bill Studio / Livenza Vault =====
ELECTRICITY_MAX_FILE_BYTES=16*1024*1024
ELECTRICITY_ALLOWED_EXTENSIONS={'.pdf','.jpg','.jpeg','.png','.webp','.heic','.heif','.tif','.tiff','.csv','.xlsx','.xls'}

def _audit_safe_meta(meta):
    if not isinstance(meta,dict): return {}
    blocked=('password','secret','token','otp','pin','cvv','captcha','cookie','credential')
    return {str(k)[:80]:v for k,v in meta.items() if not any(x in str(k).lower() for x in blocked)}

def record_audit(action,target_type='',target_id=None,status='success',note='',meta=None,module='electricity'):
    try:
        actor=current_user()
        event=AuditEvent(actor_user_id=actor.id if actor else None,module=module,action=(action or '')[:120],target_type=(target_type or '')[:80],target_id=target_id,status=(status or 'success')[:32],note=(note or '')[:1000],meta_json=json.dumps(_audit_safe_meta(meta or {}),ensure_ascii=False))
        db.session.add(event)
        return event
    except Exception:
        return None

def _mask_connection_identifier(value):
    raw=str(value or '').strip()
    if not raw: return ''
    return ('•'*max(4,len(raw)-4))+raw[-4:]

def _electricity_decimal(value,scale='0.00'):
    try: return decimal.Decimal(str(value or '0').replace(',','')).quantize(decimal.Decimal(scale))
    except Exception: return decimal.Decimal('0.00')

def _electricity_num_or_none(value):
    if value in (None,''): return None
    try: return decimal.Decimal(str(value).replace(',',''))
    except Exception: return None

def _electricity_provider_rows(include_inactive=False):
    query=ElectricityProvider.query
    if not include_inactive: query=query.filter_by(active=True)
    return query.order_by(ElectricityProvider.state,ElectricityProvider.city,ElectricityProvider.name).all()

def _electricity_page_context(bill_draft=None):
    refresh_electricity_reminders()
    connections=ElectricityConnection.query.order_by(ElectricityConnection.property_name,ElectricityConnection.connection_name).all()
    providers=_electricity_provider_rows()
    bills=ElectricityBill.query.order_by(ElectricityBill.due_date.asc().nullslast(),ElectricityBill.id.desc()).limit(300).all()
    cities=City.query.filter_by(active=True).order_by(City.name).all()
    vault_entries=VaultSecret.query.order_by(VaultSecret.label).all() if current_user() and (current_user().role or '').lower()=='admin' else []
    due_count=sum(1 for bill in bills if bill.status in ('due_soon','due_today','overdue','payment_pending_confirmation'))
    payment_by_bill={}
    for payment in ElectricityPayment.query.order_by(ElectricityPayment.id.desc()).all():
        payment_by_bill.setdefault(payment.bill_id,payment)
    return dict(connections=connections,providers=providers,bills=bills,cities=cities,vault_entries=vault_entries,bill_draft=bill_draft or {},is_admin=(current_user().role or '').lower()=='admin',due_count=due_count,payment_by_bill=payment_by_bill)

def _electricity_extract_text(raw,filename,mimetype):
    ext=Path(filename or '').suffix.lower()
    if ext=='.pdf':
        try:
            try: import pymupdf as fitz
            except ImportError: import fitz
            with fitz.open(stream=raw,filetype='pdf') as doc:
                text='\n'.join((doc.load_page(i).get_text('text') or '') for i in range(min(doc.page_count,20)))
            if len(text.strip())>=40: return text,''
        except Exception: pass
    if ext=='.csv':
        for enc in ('utf-8-sig','utf-8','cp1252','latin-1'):
            try: return raw.decode(enc),''
            except Exception: pass
    if ext=='.xlsx':
        try:
            from openpyxl import load_workbook
            wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
            return '\n'.join(' | '.join(str(v or '') for v in row) for row in ws.iter_rows(values_only=True) if any(v not in (None,'') for v in row)),''
        except Exception: pass
    if ext=='.xls':
        try:
            import xlrd
            wb=xlrd.open_workbook(file_contents=raw,on_demand=True); ws=wb.sheet_by_index(0)
            return '\n'.join(' | '.join(str(v or '') for v in ws.row_values(i)) for i in range(ws.nrows)),''
        except Exception: pass
    return _aadhaar_local_ocr_text(raw,filename,mimetype)

def _electricity_value_after_label(text,labels,max_len=120):
    label='|'.join(re.escape(x) for x in labels)
    m=re.search(rf'(?:{label})\s*[:#\-]?\s*([^\n|]{{1,{max_len}}})',text,re.I)
    return (m.group(1).strip() if m else '')

def _electricity_parse_bill_text(text):
    txt=(text or '').replace('\r','\n')
    payload={}
    payload['identifier_primary']=_electricity_value_after_label(txt,['K No','K Number','KNO','Consumer No','Consumer Number','CA No','CA Number','Account No','Account Number','Service No','Unique Service No'])
    payload['bill_number']=_electricity_value_after_label(txt,['Bill No','Bill Number','Bill ID'])
    payload['meter_number']=_electricity_value_after_label(txt,['Meter No','Meter Number'])
    payload['consumer_name']=_electricity_value_after_label(txt,['Consumer Name','Name of Consumer','Customer Name'])
    payload['bill_month']=_electricity_value_after_label(txt,['Bill Month','Billing Month','Bill Period','Billing Period'])
    def date_for(labels):
        raw=_electricity_value_after_label(txt,labels,60)
        if raw:
            m=re.search(r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[- ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[- ]\d{2,4})\b',raw,re.I)
            return m.group(0) if m else raw[:30]
        label='|'.join(re.escape(x) for x in labels)
        m=re.search(rf'(?:{label}).{{0,22}}?(\d{{1,2}}[/-]\d{{1,2}}[/-]\d{{2,4}}|\d{{4}}-\d{{1,2}}-\d{{1,2}})',txt,re.I)
        return m.group(1) if m else ''
    payload['bill_date']=date_for(['Bill Date','Date of Bill'])
    payload['due_date']=date_for(['Due Date','Payment Due Date','Pay By'])
    def money_for(labels):
        label='|'.join(re.escape(x) for x in labels)
        matches=list(re.finditer(rf'(?:{label})[^\n₹\d]{{0,30}}(?:₹|Rs\.?|INR)?\s*([\d,]+(?:\.\d{{1,2}})?)',txt,re.I))
        return matches[-1].group(1) if matches else ''
    payload['total_due_amount']=money_for(['Total Amount','Amount Due','Net Payable','Total Due','Payable Amount','Bill Amount','Current Amount'])
    payload['current_charges']=money_for(['Current Charges','Current Demand','Energy Charges'])
    payload['arrears_amount']=money_for(['Arrears','Previous Dues','Outstanding'])
    payload['late_fee_amount']=money_for(['Late Fee','LPSC','Late Payment Surcharge','Surcharge'])
    def number_for(labels):
        label='|'.join(re.escape(x) for x in labels)
        m=re.search(rf'(?:{label})[^\n\d]{{0,30}}([\d,]+(?:\.\d+)?)',txt,re.I)
        return m.group(1) if m else ''
    payload['units_consumed']=number_for(['Units Consumed','Consumption','Units'])
    payload['previous_reading']=number_for(['Previous Reading','Prev Reading'])
    payload['current_reading']=number_for(['Current Reading'])
    normalized=normalize_bill_payload(payload)
    normalized['raw_excerpt']=re.sub(r'\s+',' ',txt)[:1200]
    return normalized

def _electricity_bill_complete(draft):
    return bool(draft.get('total_due_amount') not in ('','0.00') and draft.get('due_date'))

def _electricity_bill_to_row(bill):
    c=bill.connection; p=bill.provider
    return {
        'City':(c.city_ref.name if c and c.city_ref else (p.city if p else '')),
        'Property':c.property_name if c else '',
        'Provider':p.name if p else '',
        'Connection':c.connection_name if c else '',
        'Consumer Name':bill.consumer_name or (c.consumer_name if c else ''),
        'Identifier':_mask_connection_identifier(c.identifier_primary if c else ''),
        'Meter No':bill.meter_number or (c.meter_number if c else ''),
        'Bill Month':bill.bill_month or '',
        'Bill No':bill.bill_number or '',
        'Bill Date':bill.bill_date.isoformat() if bill.bill_date else '',
        'Due Date':bill.due_date.isoformat() if bill.due_date else '',
        'Units':str(bill.units_consumed or ''),
        'Previous Reading':str(bill.previous_reading or ''),
        'Current Reading':str(bill.current_reading or ''),
        'Current Charges':str(bill.current_charges or 0),
        'Arrears':str(bill.arrears_amount or 0),
        'Late Fee':str(bill.late_fee_amount or 0),
        'Total Due':str(bill.total_due_amount or 0),
        'Status':bill.status,
        'Source':bill.source_type,
        'Receipt':('Yes' if bill.receipt_file_path_or_token else 'No'),
    }

def sync_bill_reminder(bill,connection=None):
    connection=connection or bill.connection
    pending=ElectricityPayment.query.filter_by(bill_id=bill.id).filter(ElectricityPayment.status.in_(['initiated','pending','manual_confirmation_required'])).first() is not None
    status,severity=reminder_status(bill.due_date,bill.status=='paid',pending,days_before=(connection.reminder_days_before if connection else 5))
    bill.status=status if status!='paid' else 'paid'
    reminder=ReminderItem.query.filter_by(module='electricity',entity_id=bill.id).first()
    if not reminder:
        reminder=ReminderItem(module='electricity',entity_id=bill.id,title='Electricity bill reminder'); db.session.add(reminder)
    provider=bill.provider
    payload={'bill_id':bill.id,'connection_id':bill.connection_id,'property':connection.property_name if connection else '','provider':provider.name if provider else '','identifier':_mask_connection_identifier(connection.identifier_primary if connection else ''),'amount':str(bill.total_due_amount or 0),'due_date':bill.due_date.isoformat() if bill.due_date else '','bill_status':bill.status}
    reminder.title=f"{payload['property'] or payload['provider']} electricity bill"
    existing_payload={}
    try: existing_payload=json.loads(reminder.payload_json or '{}')
    except Exception: existing_payload={}
    snoozed_until=parse_date(existing_payload.get('snoozed_until',''))
    reminder.severity=severity; reminder.due_at=datetime.datetime.combine(bill.due_date,datetime.time(9,0)) if bill.due_date else None
    if snoozed_until and snoozed_until>datetime.date.today() and bill.status!='paid':
        payload['snoozed_until']=snoozed_until.isoformat(); reminder.status='snoozed'; reminder.due_at=datetime.datetime.combine(snoozed_until,datetime.time(9,0))
    else:
        reminder.status='resolved' if bill.status=='paid' else ('active' if status in ('due_soon','due_today','overdue','payment_pending_confirmation') else 'resolved')
    reminder.payload_json=json.dumps(payload,ensure_ascii=False)
    return reminder

def upsert_electricity_bill(connection,payload,source_type='manual_entry',raw_file=None,file_name='',mime_type=''):
    normalized=normalize_bill_payload(payload or {})
    bill_month=(normalized.get('bill_month') or (parse_date(normalized.get('bill_date')) or datetime.date.today()).strftime('%Y-%m'))[:24]
    key=bill_dedupe_key(connection.provider_id,connection.id,normalized.get('bill_number',''),bill_month)
    bill=ElectricityBill.query.filter_by(dedupe_key=key).first()
    if not bill:
        bill=ElectricityBill(connection_id=connection.id,provider_id=connection.provider_id,dedupe_key=key); db.session.add(bill)
    bill.bill_month=bill_month; bill.bill_number=normalized.get('bill_number','')[:140]; bill.bill_date=parse_date(normalized.get('bill_date')); bill.due_date=parse_date(normalized.get('due_date'))
    bill.consumer_name=(normalized.get('consumer_name') or connection.consumer_name or '')[:180]; bill.meter_number=(normalized.get('meter_number') or connection.meter_number or '')[:120]
    bill.units_consumed=_electricity_num_or_none(normalized.get('units_consumed')); bill.previous_reading=_electricity_num_or_none(normalized.get('previous_reading')); bill.current_reading=_electricity_num_or_none(normalized.get('current_reading'))
    bill.current_charges=_electricity_decimal(normalized.get('current_charges')); bill.arrears_amount=_electricity_decimal(normalized.get('arrears_amount')); bill.late_fee_amount=_electricity_decimal(normalized.get('late_fee_amount')); bill.total_due_amount=_electricity_decimal(normalized.get('total_due_amount')); bill.net_amount=bill.total_due_amount
    bill.source_type=(source_type or 'manual_entry')[:48]; bill.raw_source_meta_json=json.dumps({'identifier_from_bill':normalized.get('identifier_primary',''),'excerpt':normalized.get('raw_excerpt','')},ensure_ascii=False)
    if raw_file:
        bill.encrypted_bill_blob=_bank_encrypt_bytes(raw_file); bill.bill_file_name=secure_filename(file_name or 'electricity-bill')[:255]; bill.bill_mime_type=(mime_type or 'application/octet-stream')[:120]; bill.bill_file_path_or_token='database-encrypted'
    db.session.flush(); sync_bill_reminder(bill,connection); return bill

def _electricity_payment_configured(provider):
    return bool(provider and provider.supports_bbps_payment and os.getenv('ELECTRICITY_PAYMENT_PROVIDER_URL','').strip())

def refresh_electricity_reminders():
    bills=ElectricityBill.query.filter(ElectricityBill.status!='paid').all()
    changed=False
    for bill in bills:
        sync_bill_reminder(bill); changed=True
    if changed: db.session.commit()
    return len(bills)

def _electricity_current_reminders(limit=8):
    items=ReminderItem.query.filter_by(module='electricity',status='active').all()
    rank={'danger':0,'warning':1,'info':2,'success':3}
    items.sort(key=lambda r:(rank.get(r.severity,9),r.due_at or datetime.datetime.max))
    out=[]
    for r in items[:limit]:
        try: payload=json.loads(r.payload_json or '{}')
        except Exception: payload={}
        out.append({'reminder':r,'payload':payload})
    return out

def ensure_electricity_provider_seed():
    try:
        if ElectricityProvider.query.count()>0: return 0
        rows=load_seed_providers(os.path.join(BASE_DIR,'data','electricity_providers_india.json'))
        return seed_electricity_providers(db.session,ElectricityProvider,rows)
    except Exception as exc:
        app.logger.warning('Electricity provider seed skipped: %s',str(exc)[:180]); return 0

@app.route('/electricity')
@permission_required('electricity')
def electricity_studio():
    return render_template('electricity.html',**_electricity_page_context())

@app.route('/electricity/connections/save',methods=['POST'])
@admin_required
def electricity_connection_save():
    cid=(request.form.get('id') or '').strip(); connection=db.session.get(ElectricityConnection,int(cid)) if cid.isdigit() else None
    provider_id=(request.form.get('provider_id') or '').strip(); provider=db.session.get(ElectricityProvider,int(provider_id)) if provider_id.isdigit() else None
    identifier=(request.form.get('identifier_primary') or '').strip(); property_name=(request.form.get('property_name') or '').strip()
    if not provider or not identifier or not property_name:
        flash('Provider, property and primary K/CA/Consumer/Account number are required.','danger'); return redirect(url_for('electricity_studio'))
    if not connection:
        connection=ElectricityConnection(provider_id=provider.id,identifier_primary=identifier,created_by_user_id=current_user().id); db.session.add(connection)
    connection.provider_id=provider.id; connection.property_name=property_name[:180]; connection.connection_name=(request.form.get('connection_name') or property_name)[:180]; connection.consumer_name=(request.form.get('consumer_name') or '')[:180]
    connection.identifier_primary=identifier[:180]; connection.identifier_primary_type=(request.form.get('identifier_primary_type') or 'CONSUMER_NO')[:40]; connection.identifier_secondary=(request.form.get('identifier_secondary') or '')[:180]; connection.identifier_secondary_type=(request.form.get('identifier_secondary_type') or '')[:40]; connection.meter_number=(request.form.get('meter_number') or '')[:120]
    city_id=(request.form.get('city_id') or '').strip(); connection.city_id=int(city_id) if city_id.isdigit() else None
    vault_id=(request.form.get('vault_credential_id') or '').strip(); connection.vault_credential_id=int(vault_id) if vault_id.isdigit() else None
    try: connection.reminder_days_before=max(0,min(30,int(request.form.get('reminder_days_before') or 5)))
    except Exception: connection.reminder_days_before=5
    connection.status='active'; db.session.flush(); record_audit('connection_saved','electricity_connection',connection.id,meta={'provider_id':provider.id,'property':property_name}); db.session.commit(); flash('Electricity connection saved.','success'); return redirect(url_for('electricity_studio'))

@app.route('/electricity/connections/<int:cid>/delete',methods=['POST'])
@admin_required
def electricity_connection_delete(cid):
    connection=db.session.get(ElectricityConnection,cid) or abort(404)
    bills=ElectricityBill.query.filter_by(connection_id=cid).all()
    for bill in bills:
        ReminderItem.query.filter_by(module='electricity',entity_id=bill.id).delete(); ElectricityPayment.query.filter_by(bill_id=bill.id).delete(); db.session.delete(bill)
    VaultSecret.query.filter_by(linked_connection_id=cid).update({'linked_connection_id':None})
    record_audit('connection_deleted','electricity_connection',connection.id,meta={'property':connection.property_name}); db.session.delete(connection); db.session.commit(); flash('Electricity connection removed.','success'); return redirect(url_for('electricity_studio'))

@app.route('/electricity/providers/<int:provider_id>/portal')
@permission_required('electricity')
def electricity_provider_portal(provider_id):
    provider=db.session.get(ElectricityProvider,provider_id) or abort(404)
    target=safe_official_url(provider.official_login_url or provider.official_website_url)
    if not target: flash('This provider does not have a verified official portal URL saved yet.','warning'); return redirect(url_for('electricity_studio'))
    connection_id=(request.args.get('connection_id') or '').strip(); connection=db.session.get(ElectricityConnection,int(connection_id)) if connection_id.isdigit() else None
    if connection and connection.provider_id!=provider.id: connection=None
    return render_template('electricity_portal.html',provider=provider,portal_url=target,connection=connection,inline_allowed=(provider.embedding_mode=='inline'),vault_linked=bool(connection and connection.vault_credential_id))

@app.route('/electricity/connections/<int:cid>/fetch',methods=['POST'])
@permission_required('electricity')
def electricity_bill_fetch(cid):
    connection=db.session.get(ElectricityConnection,cid) or abort(404); provider=connection.provider
    config={'base_url':os.getenv('BBPS_PROVIDER_BASE_URL',''),'client_id':os.getenv('BBPS_PROVIDER_CLIENT_ID',''),'client_secret':os.getenv('BBPS_PROVIDER_CLIENT_SECRET',''),'fetch_path':os.getenv('BBPS_PROVIDER_FETCH_PATH','/bill-fetch')}
    result=fetch_bill_from_provider(connection,provider,config)
    connection.last_fetch_at=datetime.datetime.utcnow(); connection.last_fetch_status=result.get('status','failed'); connection.status='active' if result.get('ok') else 'needs_attention'
    if result.get('ok'):
        bill=upsert_electricity_bill(connection,result.get('bill') or {},'bbps'); record_audit('bill_fetched','electricity_bill',bill.id,meta={'provider_id':provider.id}); db.session.commit(); flash(result.get('message') or 'Current bill fetched.','success')
    else:
        record_audit('bill_fetch_failed','electricity_connection',connection.id,status='failed',note=result.get('message','')); db.session.commit(); flash(result.get('message') or 'Automatic bill fetch needs manual action. Use the official portal or upload the bill.','warning')
    return redirect(url_for('electricity_studio'))

@app.route('/electricity/bills/upload',methods=['POST'])
@permission_required('electricity')
def electricity_bill_upload():
    connection_id=(request.form.get('connection_id') or '').strip(); connection=db.session.get(ElectricityConnection,int(connection_id)) if connection_id.isdigit() else None
    if not connection: flash('Choose the electricity connection for this bill.','danger'); return redirect(url_for('electricity_studio'))
    if request.form.get('confirm_extracted')=='1':
        draft={k:request.form.get(k,'') for k in ('identifier_primary','bill_number','bill_date','due_date','bill_month','total_due_amount','current_charges','arrears_amount','late_fee_amount','units_consumed','meter_number','previous_reading','current_reading','consumer_name')}
        bill=upsert_electricity_bill(connection,draft,'manual_entry'); record_audit('bill_import_confirmed','electricity_bill',bill.id); db.session.commit(); flash('Electricity bill saved to the register.','success'); return redirect(url_for('electricity_studio'))
    upload=request.files.get('bill_file')
    if not upload or not upload.filename: flash('Choose an electricity bill PDF, image or spreadsheet.','danger'); return redirect(url_for('electricity_studio'))
    ext=Path(upload.filename).suffix.lower()
    if ext not in ELECTRICITY_ALLOWED_EXTENSIONS: flash('Use PDF, JPG, PNG, WebP, HEIC/HEIF, TIFF, CSV, XLSX or XLS.','danger'); return redirect(url_for('electricity_studio'))
    raw=upload.read(ELECTRICITY_MAX_FILE_BYTES+1)
    if not raw or len(raw)>ELECTRICITY_MAX_FILE_BYTES: flash('Electricity bill files must be smaller than 16 MB.','danger'); return redirect(url_for('electricity_studio'))
    text,error=_electricity_extract_text(raw,upload.filename,upload.mimetype or '')
    if error and not text: flash(error,'danger'); return redirect(url_for('electricity_studio'))
    draft=_electricity_parse_bill_text(text); draft['connection_id']=connection.id; draft['file_name']=secure_filename(upload.filename); draft['parse_notice']='Please review the extracted fields before saving.'
    if _electricity_bill_complete(draft):
        bill=upsert_electricity_bill(connection,draft,'manual_upload',raw,upload.filename,upload.mimetype or 'application/octet-stream'); record_audit('bill_imported','electricity_bill',bill.id,meta={'file_name':secure_filename(upload.filename)}); db.session.commit(); flash('Bill extracted, saved securely and added to the register.','success'); return redirect(url_for('electricity_studio'))
    record_audit('bill_import_needs_review','electricity_connection',connection.id,status='review',meta={'file_name':secure_filename(upload.filename)}); db.session.commit(); flash('Some bill fields need confirmation before saving.','warning'); return render_template('electricity.html',**_electricity_page_context(draft))

@app.route('/electricity/bills/<int:bill_id>/download')
@permission_required('electricity')
def electricity_bill_download(bill_id):
    bill=db.session.get(ElectricityBill,bill_id) or abort(404)
    if not bill.encrypted_bill_blob: abort(404)
    raw=_bank_decrypt_bytes(bill.encrypted_bill_blob)
    if not raw: abort(410)
    return send_file(io.BytesIO(raw),as_attachment=True,download_name=bill.bill_file_name or 'electricity-bill',mimetype=bill.bill_mime_type or 'application/octet-stream')

@app.route('/electricity/register')
@permission_required('electricity')
def electricity_register():
    bills=ElectricityBill.query.order_by(ElectricityBill.due_date.desc().nullslast(),ElectricityBill.id.desc()).all(); return render_template('electricity_register.html',bills=bills,rows=[_electricity_bill_to_row(b) for b in bills])

@app.route('/electricity/register.csv')
@permission_required('electricity')
def electricity_register_csv():
    rows=[_electricity_bill_to_row(b) for b in ElectricityBill.query.order_by(ElectricityBill.id.desc()).all()]; raw=build_electricity_csv(rows); return send_file(io.BytesIO(raw),as_attachment=True,download_name='Livenza_Electricity_Bill_Register.csv',mimetype='text/csv; charset=utf-8')

@app.route('/electricity/register.xlsx')
@permission_required('electricity')
def electricity_register_xlsx():
    rows=[_electricity_bill_to_row(b) for b in ElectricityBill.query.order_by(ElectricityBill.id.desc()).all()]; raw=build_electricity_xlsx(rows); return send_file(io.BytesIO(raw),as_attachment=True,download_name='Livenza_Electricity_Bill_Register.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/electricity/bills/<int:bill_id>/pay',methods=['POST'])
@admin_required
def electricity_payment_start(bill_id):
    bill=db.session.get(ElectricityBill,bill_id) or abort(404); provider=bill.provider
    if bill.status=='paid': flash('This bill is already marked paid.','success'); return redirect(url_for('electricity_studio'))
    endpoint=os.getenv('ELECTRICITY_PAYMENT_PROVIDER_URL','').strip()
    if not _electricity_payment_configured(provider):
        target=safe_official_url(provider.official_payment_url or provider.official_website_url)
        flash('Authorized in-Livenza payment is not configured for this provider yet. Opening the official payment page; upload or confirm the receipt afterward.','warning')
        return redirect(target or url_for('electricity_studio'))
    payment=ElectricityPayment(bill_id=bill.id,connection_id=bill.connection_id,provider_id=bill.provider_id,payment_provider=os.getenv('ELECTRICITY_PAYMENT_PROVIDER','Bharat Connect / BBPS')[:120],paid_amount=bill.total_due_amount,status='initiated'); db.session.add(payment); db.session.flush()
    payload={'biller_id':provider.bbps_biller_id,'consumer_identifier':bill.connection.identifier_primary,'bill_number':bill.bill_number,'amount':str(bill.total_due_amount or 0),'callback_url':url_for('electricity_payment_callback',payment_id=payment.id,_external=True)}
    try:
        client_id=os.getenv('BBPS_PROVIDER_CLIENT_ID','').strip(); secret=os.getenv('BBPS_PROVIDER_CLIENT_SECRET','').strip(); headers={'X-Client-Id':client_id,'Authorization':f'Bearer {secret}','Content-Type':'application/json'}
        response=requests.post(endpoint,json=payload,headers=headers,timeout=30)
        data=response.json() if response.headers.get('content-type','').startswith('application/json') else {}
        if not response.ok: raise RuntimeError(f'Payment provider returned HTTP {response.status_code}')
        provider_status=str(data.get('status') or 'pending').lower(); event='provider_confirmed' if provider_status in ('confirmed','success','paid') else 'provider_pending'
        payment.status=transition_payment_status(payment.status,event); payment.provider_txn_id=str(data.get('transaction_id') or data.get('txn_id') or '')[:180]; payment.payment_reference=str(data.get('reference') or '')[:180]
        if payment.status=='confirmed': payment.confirmed_at=datetime.datetime.utcnow(); bill.status='paid'
        sync_bill_reminder(bill); record_audit('payment_initiated','electricity_payment',payment.id,meta={'bill_id':bill.id,'provider_id':provider.id}); db.session.commit()
        pay_url=safe_official_url(str(data.get('payment_url') or ''))
        if pay_url: return redirect(pay_url)
        flash('Payment request created. Livenza is waiting for provider confirmation.','success'); return redirect(url_for('electricity_studio'))
    except Exception as exc:
        payment.status='failed'; record_audit('payment_failed','electricity_payment',payment.id,status='failed',note=str(exc)[:400]); db.session.commit(); flash(f'Payment could not be started: {str(exc)[:180]}','danger'); return redirect(url_for('electricity_studio'))

@app.route('/electricity/payments/<int:payment_id>/callback',methods=['POST'])
def electricity_payment_callback(payment_id):
    payment=db.session.get(ElectricityPayment,payment_id) or abort(404); raw=request.get_data() or b''; callback_secret=os.getenv('ELECTRICITY_PAYMENT_CALLBACK_SECRET','').encode('utf-8')
    if callback_secret:
        signature=(request.headers.get('X-Livenza-Signature') or '').strip(); expected=hmac.new(callback_secret,raw,hashlib.sha256).hexdigest()
        if not signature or not hmac.compare_digest(signature,expected): abort(403)
    data=request.get_json(silent=True) or {}; state=str(data.get('status') or '').lower(); event='provider_confirmed' if state in ('confirmed','success','paid') else ('provider_failed' if state in ('failed','declined','error') else 'provider_pending')
    try: payment.status=transition_payment_status(payment.status,event)
    except ValueError: return jsonify(ok=False,error='Invalid payment state transition.'),409
    bill=payment.bill
    if payment.status=='confirmed': payment.confirmed_at=datetime.datetime.utcnow(); bill.status='paid'; payment.payment_reference=str(data.get('reference') or payment.payment_reference)[:180]; payment.provider_txn_id=str(data.get('transaction_id') or data.get('txn_id') or payment.provider_txn_id)[:180]
    sync_bill_reminder(bill); record_audit('payment_callback','electricity_payment',payment.id,status=payment.status,meta={'bill_id':bill.id}); db.session.commit(); return jsonify(ok=True,status=payment.status)

@app.route('/electricity/payments/<int:payment_id>/confirm',methods=['POST'])
@admin_required
def electricity_payment_confirm_manual(payment_id):
    payment=db.session.get(ElectricityPayment,payment_id) or abort(404); admin=current_user()
    if not check_password_hash(admin.password_hash,request.form.get('admin_password','')): flash('Administrator password is required to confirm payment.','danger'); return redirect(url_for('electricity_studio'))
    if payment.status not in ('pending','manual_confirmation_required','initiated'):
        flash('This payment is not awaiting confirmation.','warning'); return redirect(url_for('electricity_studio'))
    payment.status='confirmed'; payment.confirmed_at=datetime.datetime.utcnow(); payment.payment_reference=(request.form.get('payment_reference') or payment.payment_reference)[:180]; payment.bill.status='paid'; sync_bill_reminder(payment.bill); record_audit('payment_confirmed_manual','electricity_payment',payment.id,meta={'bill_id':payment.bill_id}); db.session.commit(); flash('Electricity payment confirmed and reminder cleared.','success'); return redirect(url_for('electricity_studio'))

@app.route('/electricity/reminders/<int:rid>/snooze',methods=['POST'])
@admin_required
def electricity_reminder_snooze(rid):
    reminder=db.session.get(ReminderItem,rid) or abort(404)
    if reminder.module!='electricity': abort(404)
    try: days=max(1,min(30,int(request.form.get('days') or 1)))
    except Exception: days=1
    until=datetime.date.today()+datetime.timedelta(days=days)
    try: payload=json.loads(reminder.payload_json or '{}')
    except Exception: payload={}
    payload['snoozed_until']=until.isoformat(); reminder.payload_json=json.dumps(payload,ensure_ascii=False); reminder.status='snoozed'; reminder.due_at=datetime.datetime.combine(until,datetime.time(9,0)); record_audit('reminder_snoozed','reminder_item',reminder.id,meta={'days':days,'bill_id':reminder.entity_id}); db.session.commit(); flash(f'Reminder snoozed for {days} day(s).','success'); return redirect(url_for('dashboard')+'#live-reminders')

@app.route('/admin/vault')
@admin_required
def vault_page():
    values={k:v for k,v in {'edit_secret':request.args.get('edit_secret'),'edit_provider':request.args.get('edit_provider')}.items() if v}
    return redirect(settings_pane_url('organisation', **values))

@app.route('/admin/vault/save',methods=['POST'])
@admin_required
def vault_secret_save():
    master=os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()
    if not master: flash('LIVENZA_VAULT_MASTER_KEY is not configured on the server.','danger'); return redirect(url_for('vault_page'))
    sid=(request.form.get('id') or '').strip(); entry=db.session.get(VaultSecret,int(sid)) if sid.isdigit() else None
    try: secret_type=validate_secret_type(request.form.get('secret_type',''))
    except ValueError as exc: flash(str(exc),'danger'); return redirect(url_for('vault_page'))
    username=(request.form.get('username') or '').strip(); secret=request.form.get('secret_value') or ''
    if entry and not username and entry.ciphertext:
        try: username=str(json.loads(decrypt_secret(entry.ciphertext,entry.nonce,master)).get('username') or '')
        except Exception: username=''
    if not secret and not entry: flash('Enter the secret value.','danger'); return redirect(url_for('vault_page'))
    if not entry: entry=VaultSecret(secret_type=secret_type,label=(request.form.get('label') or 'Vault Entry')[:180],ciphertext='',nonce='',created_by_user_id=current_user().id); db.session.add(entry)
    entry.secret_type=secret_type; entry.label=(request.form.get('label') or entry.label or 'Vault Entry')[:180]
    if username: entry.username_masked=mask_secret(username)
    if secret:
        payload=json.dumps({'username':username,'secret':secret},ensure_ascii=False); entry.ciphertext,entry.nonce=encrypt_secret(payload,master)
    provider_id=(request.form.get('linked_provider_id') or '').strip(); connection_id=(request.form.get('linked_connection_id') or '').strip(); entry.linked_provider_id=int(provider_id) if provider_id.isdigit() else None; entry.linked_connection_id=int(connection_id) if connection_id.isdigit() else None; entry.updated_by_user_id=current_user().id
    db.session.flush()
    if entry.linked_connection_id:
        linked=db.session.get(ElectricityConnection,entry.linked_connection_id)
        if linked: linked.vault_credential_id=entry.id
    record_audit('vault_secret_saved','vault_secret',entry.id,module='vault',meta={'secret_type':entry.secret_type,'linked_provider_id':entry.linked_provider_id,'linked_connection_id':entry.linked_connection_id}); db.session.commit(); flash('Vault entry encrypted and saved.','success'); return redirect(url_for('vault_page'))

@app.route('/admin/vault/<int:sid>/reveal',methods=['POST'])
@admin_required
def vault_secret_reveal(sid):
    entry=db.session.get(VaultSecret,sid) or abort(404); admin=current_user()
    if not check_password_hash(admin.password_hash,request.form.get('admin_password','')): return jsonify(ok=False,error='Administrator password did not match.'),403
    master=os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()
    try: payload=json.loads(decrypt_secret(entry.ciphertext,entry.nonce,master))
    except Exception: return jsonify(ok=False,error='Vault secret could not be decrypted.'),422
    record_audit('vault_secret_revealed','vault_secret',entry.id,module='vault',meta={'secret_type':entry.secret_type}); db.session.commit(); response=jsonify(ok=True,username=payload.get('username',''),secret=payload.get('secret','')); response.headers['Cache-Control']='no-store, private'; return response

@app.route('/admin/vault/<int:sid>/delete',methods=['POST'])
@admin_required
def vault_secret_delete(sid):
    entry=db.session.get(VaultSecret,sid) or abort(404); ElectricityConnection.query.filter_by(vault_credential_id=sid).update({'vault_credential_id':None}); record_audit('vault_secret_deleted','vault_secret',entry.id,module='vault',meta={'secret_type':entry.secret_type}); db.session.delete(entry); db.session.commit(); flash('Vault entry deleted.','success'); return redirect(url_for('vault_page'))

@app.route('/admin/electricity/providers/save',methods=['POST'])
@admin_required
def electricity_provider_save():
    pid=(request.form.get('id') or '').strip(); provider=db.session.get(ElectricityProvider,int(pid)) if pid.isdigit() else None
    if not provider: provider=ElectricityProvider(name='Provider',state='',city=''); db.session.add(provider)
    provider.name=(request.form.get('name') or '').strip()[:180]; provider.state=(request.form.get('state') or '').strip()[:120]; provider.city=(request.form.get('city') or '').strip()[:120]
    if not provider.name or not provider.state: flash('Provider name and state are required.','danger'); return redirect(url_for('vault_page')+'#providers')
    provider.official_website_url=safe_official_url(request.form.get('official_website_url','')); provider.official_payment_url=safe_official_url(request.form.get('official_payment_url','')); provider.official_login_url=safe_official_url(request.form.get('official_login_url',''))
    provider.identifier_types_json=json.dumps([x.strip().upper().replace(' ','_') for x in (request.form.get('identifier_types') or 'CONSUMER_NO').split(',') if x.strip()]); provider.bbps_biller_id=(request.form.get('bbps_biller_id') or '')[:120]; provider.supports_bbps_fetch=request.form.get('supports_bbps_fetch')=='1'; provider.supports_bbps_payment=request.form.get('supports_bbps_payment')=='1'; provider.embedding_mode=request.form.get('embedding_mode') if request.form.get('embedding_mode') in ('inline','external','none') else 'external'; provider.workflow_mode=request.form.get('workflow_mode') if request.form.get('workflow_mode') in ('bbps','portal','upload_only','hybrid') else 'hybrid'; provider.active=request.form.get('active')=='1'; provider.notes=(request.form.get('notes') or '')[:1000]
    db.session.flush(); record_audit('provider_saved','electricity_provider',provider.id,meta={'state':provider.state,'city':provider.city}); db.session.commit(); flash('Electricity provider saved.','success'); return redirect(url_for('vault_page')+'#providers')

@app.route('/admin/electricity/providers/<int:pid>/delete',methods=['POST'])
@admin_required
def electricity_provider_delete(pid):
    provider=db.session.get(ElectricityProvider,pid) or abort(404)
    if ElectricityConnection.query.filter_by(provider_id=pid).first() or VaultSecret.query.filter_by(linked_provider_id=pid).first(): flash('This provider is linked to saved connections or Vault entries and cannot be deleted. Disable it instead.','warning'); return redirect(url_for('vault_page')+'#providers')
    record_audit('provider_deleted','electricity_provider',provider.id,meta={'state':provider.state}); db.session.delete(provider); db.session.commit(); flash('Electricity provider removed.','success'); return redirect(url_for('vault_page')+'#providers')

@app.route('/banking')
@permission_required('banking')
def banking_suite():
    user=current_user(); selected_key=(request.args.get('bank') or 'sbi').strip().lower()
    selected=next((b for b in BANK_PORTALS if b['key']==selected_key),BANK_PORTALS[0])
    statements=BankDocument.query.filter_by(user_id=user.id,document_type='statement').order_by(BankDocument.id.desc()).limit(100).all()
    templates=BankDocument.query.filter_by(user_id=user.id,document_type='template').order_by(BankDocument.id.desc()).limit(100).all()
    runs=BankReconciliationRun.query.filter_by(user_id=user.id).order_by(BankReconciliationRun.id.desc()).limit(30).all()
    run_cards=[]
    for run in runs:
        try: summary=json.loads(run.summary_json or '{}')
        except Exception: summary={}
        run_cards.append({'run':run,'summary':summary,'statement':db.session.get(BankDocument,run.statement_id),'template':db.session.get(BankDocument,run.template_id)})
    categories=[]
    for row in BANK_PORTALS:
        if row['category'] not in categories: categories.append(row['category'])
    return render_template('banking.html',banks=BANK_PORTALS,categories=categories,selected=selected,statements=statements,templates=templates,runs=run_cards)


@app.route('/banking/documents/upload',methods=['POST'])
@permission_required('banking')
def banking_document_upload():
    user=current_user(); kind=(request.form.get('document_type') or 'statement').strip().lower()
    if kind not in ('statement','template'): kind='statement'
    upload=request.files.get('bank_file')
    if not upload or not upload.filename:
        flash('Choose a bank statement or reconciliation template file.','danger'); return redirect(url_for('banking_suite')+'#vault')
    name=_bank_clean_name(upload.filename); ext=Path(name).suffix.lower()
    if ext not in BANK_ALLOWED_EXTENSIONS:
        flash('Use CSV, XLSX, XLS or PDF files for banking reconciliation.','danger'); return redirect(url_for('banking_suite')+'#vault')
    raw=upload.read(BANK_MAX_FILE_BYTES+1)
    if not raw or len(raw)>BANK_MAX_FILE_BYTES:
        flash('Bank files must be smaller than 16 MB.','danger'); return redirect(url_for('banking_suite')+'#vault')
    rows=_bank_parse_rows(raw,name)
    status=(f'{len(rows)} entries extracted' if rows else ('Saved securely; PDF/table extraction needs review' if ext=='.pdf' else 'Saved securely; no transaction rows detected'))
    doc=BankDocument(user_id=user.id,document_type=kind,bank_name=(request.form.get('bank_name') or '').strip()[:160],account_label=(request.form.get('account_label') or '').strip()[:160],title=(request.form.get('title') or '').strip()[:180],file_name=name,mime_type=(upload.mimetype or 'application/octet-stream')[:120],encrypted_blob=_bank_encrypt_bytes(raw),parsed_ciphertext=_bank_encrypt_json(rows),row_count=len(rows),parse_status=status)
    db.session.add(doc); db.session.commit()
    template_id=request.form.get('template_id','').strip()
    if kind=='statement' and template_id.isdigit() and rows:
        template=db.session.get(BankDocument,int(template_id))
        if template and template.user_id==user.id and template.document_type=='template':
            template_rows=_bank_decrypt_json(template.parsed_ciphertext)
            if template_rows:
                result=_bank_reconcile(rows,template_rows); summary=result['summary']
                run=BankReconciliationRun(user_id=user.id,statement_id=doc.id,template_id=template.id,summary_json=json.dumps(summary,separators=(',',':')),result_ciphertext=_bank_encrypt_json(result))
                db.session.add(run); db.session.commit()
                flash(f"Statement saved and reconciled: {summary['matched']} matched, {summary['missing']} missing, {summary['review']} to review.",'success')
                return redirect(url_for('banking_reconciliation',run_id=run.id))
    flash(('Statement' if kind=='statement' else 'Template')+' saved securely. '+status+'.','success')
    return redirect(url_for('banking_suite')+'#vault')


@app.route('/banking/reconcile',methods=['POST'])
@permission_required('banking')
def banking_reconcile():
    user=current_user(); sid=request.form.get('statement_id',''); tid=request.form.get('template_id','')
    if not (sid.isdigit() and tid.isdigit()):
        flash('Select both a bank statement and a reconciliation template.','danger'); return redirect(url_for('banking_suite')+'#reconcile')
    statement=db.session.get(BankDocument,int(sid)); template=db.session.get(BankDocument,int(tid))
    if not statement or not template or statement.user_id!=user.id or template.user_id!=user.id or statement.document_type!='statement' or template.document_type!='template': abort(404)
    srows=_bank_decrypt_json(statement.parsed_ciphertext); trows=_bank_decrypt_json(template.parsed_ciphertext)
    if not srows or not trows:
        flash('Both files need readable transaction rows. CSV/XLSX gives the most reliable reconciliation.','danger'); return redirect(url_for('banking_suite')+'#reconcile')
    result=_bank_reconcile(srows,trows); summary=result['summary']
    run=BankReconciliationRun(user_id=user.id,statement_id=statement.id,template_id=template.id,summary_json=json.dumps(summary,separators=(',',':')),result_ciphertext=_bank_encrypt_json(result))
    db.session.add(run); db.session.commit()
    return redirect(url_for('banking_reconciliation',run_id=run.id))


@app.route('/banking/reconciliation/<int:run_id>')
@permission_required('banking')
def banking_reconciliation(run_id):
    user=current_user(); run=db.session.get(BankReconciliationRun,run_id) or abort(404)
    if run.user_id!=user.id: abort(404)
    statement=db.session.get(BankDocument,run.statement_id); template=db.session.get(BankDocument,run.template_id)
    result=_bank_decrypt_json(run.result_ciphertext)
    return render_template('bank_reconciliation.html',run=run,result=result,statement=statement,template=template)


@app.route('/banking/reconciliation/<int:run_id>/csv')
@permission_required('banking')
def banking_reconciliation_csv(run_id):
    user=current_user(); run=db.session.get(BankReconciliationRun,run_id) or abort(404)
    if run.user_id!=user.id: abort(404)
    result=_bank_decrypt_json(run.result_ciphertext); output=io.StringIO(); writer=csv.writer(output)
    writer.writerow(['Status','Score','Expected Date','Expected Amount','Expected Reference','Expected Description','Statement Date','Statement Amount','Statement Reference','Statement Description'])
    for item in result.get('matches',[]):
        e=item.get('expected') or {}; a=item.get('actual') or {}
        writer.writerow([item.get('status',''),item.get('score',''),e.get('date',''),e.get('amount',''),e.get('reference',''),e.get('description',''),a.get('date',''),a.get('amount',''),a.get('reference',''),a.get('description','')])
    for a in result.get('extras',[]): writer.writerow(['extra','', '', '', '', '',a.get('date',''),a.get('amount',''),a.get('reference',''),a.get('description','')])
    raw=output.getvalue().encode('utf-8-sig')
    return send_file(io.BytesIO(raw),mimetype='text/csv',as_attachment=True,download_name=f'livenza-reconciliation-{run.id}.csv')


@app.route('/banking/documents/<int:doc_id>/download')
@permission_required('banking')
def banking_document_download(doc_id):
    user=current_user(); doc=db.session.get(BankDocument,doc_id) or abort(404)
    if doc.user_id!=user.id: abort(404)
    raw=_bank_decrypt_bytes(doc.encrypted_blob)
    if not raw: abort(404)
    return send_file(io.BytesIO(raw),mimetype=doc.mime_type or 'application/octet-stream',as_attachment=True,download_name=doc.file_name)


@app.route('/banking/documents/<int:doc_id>/delete',methods=['POST'])
@permission_required('banking')
def banking_document_delete(doc_id):
    user=current_user(); doc=db.session.get(BankDocument,doc_id) or abort(404)
    if doc.user_id!=user.id: abort(404)
    BankReconciliationRun.query.filter((BankReconciliationRun.statement_id==doc.id)|(BankReconciliationRun.template_id==doc.id)).delete(synchronize_session=False)
    db.session.delete(doc); db.session.commit(); flash('Banking document removed from the encrypted vault.','success')
    return redirect(url_for('banking_suite')+'#vault')


@app.route('/billing')
@permission_required('rentok')
def billing(): return render_template('rentok.html',url='https://manager.rentok.com/')

@app.route('/rentok')
def rentok_legacy():
    return redirect(url_for('billing'))


HELP_FEATURES = {
    'agreement': 'Open Operations → Agreements. Create a new agreement, choose a preset, optionally upload Aadhaar to auto-fill tenant details, complete any fields you need, then save and preview in English or Hindi. All agreement fields are optional.',
    'room': 'Open Operations → Rooms to maintain room inventory, occupancy and vacant-room reporting. Vacancy reports can also be scheduled to pre-fed WhatsApp recipients when the WhatsApp Cloud configuration is available.',
    'tenant': 'Open Operations → Tenants to maintain resident records, room allocation, contact information, tariff, security deposit, joining/leaving dates and agreement references.',
    'review': 'Open Reviews. Paste the direct Google Business review link, enter the genuine customer experience, generate review drafts, then copy the review and open the Google review page or scan/download its QR code.',
    'query': 'Open Queries for the live lead manager. Use card view for follow-up workflow or Spreadsheet View for direct Excel-style editing. Queries can come from manual entry, Google/Meta webhooks and OTA integrations.',
    'spreadsheet': 'In Queries, click Spreadsheet View. Edit cells directly like a sheet. Existing rows auto-save when a cell changes; use + New Row to create a fresh enquiry line.',
    'video': 'Open Video Wall. Add media, create TV/screen endpoints, assign a different playlist to each TV, set rotation/fit/loop options, or start a Festive Takeover to run one commercial across all enabled screens.',
    'billing': 'Open Billing for the Livenza Billing Suite. It embeds the configured billing manager when the external service allows embedding and otherwise provides a direct-open fallback.',
    'electricity': 'Open Applications → Electricity Bill Studio to manage saved utility connections, fetch or upload electricity bills, track K/CA/Consumer numbers, export the Bill Register, view due reminders and use official payment/provider flows. Connection identifiers, provider setup and payment confirmation are Admin-controlled.',
    'vault': 'Admins can open the profile menu → Livenza Vault to store approved electricity utility logins and operational API secrets encrypted at rest. Vault reveal requires the current Admin password. Banking passwords, UPI/card PINs, CVVs, OTPs, CAPTCHA answers and banking session cookies are not allowed.',
    'food': 'Open Food for orders and settlements. Use Integrations to configure Swiggy, Zomato, Toing or another partner using webhook/API details, and Live Partner Websites to open their official restaurant portals inside Operations Cloud when embedding is allowed.',
    'whatsapp': 'Open WhatsApp to send Cloud API messages and view the incoming message feed. Admin must configure the Meta token, phone-number ID and webhook verification secrets.',
    'email': 'Open Email to view the latest Gmail inbox metadata and compose messages without leaving Livenza. An admin must connect Google once from the Admin panel.',
    'drive': 'Open Drive to upload, list, open and download Livenza files in the configured Google Drive folder. Admin can also mirror Video Wall uploads automatically.',
    'security': 'Admins can configure pattern login and allow fingerprint/passkey enrollment per user. The kiosk PIN gate and Windows startup downloads are in Admin → Kiosk & Main Screen.',
    'fullscreen': 'Open View → Full Screen. While fullscreen is active, top navigation uses in-place page switching so moving between modules does not exit fullscreen.',
    'rotate': 'Open View and choose Auto, Portrait, Landscape, 90°, 180° or 270°. Portrait/Landscape can use device orientation on supported fullscreen mobile/tablet browsers; desktop custom angles rotate the application viewport.',
    'user': 'Admins can open the profile menu → Admin to create user IDs, passwords, profile photos and module-by-module access permissions.',
    'city': 'Admins can manage operating cities from the Admin panel. City data is then available across the dashboard, rooms, tenants, agreements and queries.',
    'letterhead': 'Open Applications → Livenza Letterhead Studio to create a residence certificate or other official document manually or with Ask Livenza AI. Review source facts and suggested attachments, complete Final Review, then finalize the PDF. Finalized PDFs are stored in Document Vault and can be sent by Email or WhatsApp when those providers are configured in Integrations Center.'
}

def _local_help_answer(question):
    q=(question or '').strip().lower()
    if not q:return 'Ask me about any Livenza Operations Cloud feature or workflow.'
    aliases=[
        (('agreement','rent agreement','lease','stamp'), 'agreement'),
        (('vacant','vacancy','room','occupancy'), 'room'),
        (('tenant','resident'), 'tenant'),
        (('review','google review','qr'), 'review'),
        (('spreadsheet','excel','sheet','grid'), 'spreadsheet'),
        (('query','lead','enquiry','inquiry','meta','facebook','airbnb','ota'), 'query'),
        (('video wall','tv','screen','festive','playlist'), 'video'),
        (('food','swiggy','zomato','toing','restaurant partner','delivery order'), 'food'),
        (('billing','rentok','rent ok'), 'billing'),
        (('electricity','electric bill','electricity bill','k no','consumer no','ca no','meter bill','discom','power bill','utility bill'), 'electricity'),
        (('vault','credential','utility password','secret'), 'vault'),
        (('whatsapp','message','chat'), 'whatsapp'),
        (('email','gmail','mail','inbox'), 'email'),
        (('drive','google drive','cloud file','upload'), 'drive'),
        (('fingerprint','passkey','windows hello','pattern','kiosk','pin','lock'), 'security'),
        (('fullscreen','full screen','f11'), 'fullscreen'),
        (('rotate','portrait','landscape','orientation'), 'rotate'),
        (('user','permission','login','password','access'), 'user'),
        (('letterhead','residence certificate','document vault','ask livenza document','send pdf by email','send pdf on whatsapp','official letter'), 'letterhead'),
        (('city','location'), 'city'),
    ]
    for words,key in aliases:
        if any(w in q for w in words):return HELP_FEATURES[key]
    return 'I can guide you through Agreements, Rooms, Tenants, Reviews, Queries and Spreadsheet View, Video Wall, Billing, Banking & Reconciliation, Electricity Bill Studio, Livenza Vault, Letterhead Studio and Document Vault, Live Reminders, Fullscreen/Rotate, users/permissions and city setup. Ask a specific question about the feature you want to use.'

@app.route('/api/help',methods=['POST'])
@login_required
def help_assistant():
    payload=request.get_json(silent=True) or {}
    question=str(payload.get('message') or '')[:1200].strip()
    if not question:return jsonify(ok=False,error='Type a question first.'),400
    fallback=_local_help_answer(question)
    key=os.getenv('OPENAI_API_KEY','').strip()
    if not key:return jsonify(ok=True,answer=fallback,mode='local')
    try:
        from openai import OpenAI
        client=OpenAI(api_key=key)
        feature_text='\n'.join(f'- {k}: {v}' for k,v in HELP_FEATURES.items())
        prompt=(
            'You are the embedded help assistant for Livenza Life Operations Cloud. '
            'Answer only questions about how to use this web application. Be concise, practical and never invent a capability. '
            'If the question is outside the application, say you can only help with Operations Cloud features. '
            'Available feature guide:\n'+feature_text+'\n\nUser question: '+question
        )
        resp=client.responses.create(model=os.getenv('OPENAI_HELP_MODEL',os.getenv('OPENAI_REVIEW_MODEL','gpt-5.6-luna')),input=prompt)
        answer=(getattr(resp,'output_text','') or '').strip() or fallback
        return jsonify(ok=True,answer=answer[:4000],mode='ai')
    except Exception:
        return jsonify(ok=True,answer=fallback,mode='local')

@app.route('/api/webauthn/register/options',methods=['POST'])
@login_required
def webauthn_register_options():
    u=current_user()
    if not u.webauthn_enabled: return jsonify(ok=False,error='Fingerprint/passkey enrollment is not enabled for this user.'),403
    try:
        from webauthn import generate_registration_options, options_to_json
        from webauthn.helpers.structs import AuthenticatorSelectionCriteria, PublicKeyCredentialDescriptor, ResidentKeyRequirement, UserVerificationRequirement
        rp_id,_=_webauthn_context()
        existing=WebAuthnCredential.query.filter_by(user_id=u.id).all()
        options=generate_registration_options(
            rp_id=rp_id,rp_name='Livenza Back Office',user_id=str(u.id).encode(),user_name=u.username,
            user_display_name=u.full_name or u.username,
            exclude_credentials=[PublicKeyCredentialDescriptor(id=c.credential_id) for c in existing],
            authenticator_selection=AuthenticatorSelectionCriteria(
                resident_key=ResidentKeyRequirement.PREFERRED,
                user_verification=UserVerificationRequirement.REQUIRED,
            ),
        )
        session['webauthn_register_challenge']=base64.urlsafe_b64encode(options.challenge).decode('ascii')
        return app.response_class(options_to_json(options),mimetype='application/json')
    except Exception as exc:
        return jsonify(ok=False,error=f'Passkey service is unavailable: {exc}'),503

@app.route('/api/webauthn/register/verify',methods=['POST'])
@login_required
def webauthn_register_verify():
    u=current_user(); challenge=session.pop('webauthn_register_challenge','')
    if not (u.webauthn_enabled and challenge): return jsonify(ok=False,error='Enrollment request expired.'),400
    try:
        from webauthn import verify_registration_response
        rp_id,origin=_webauthn_context(); payload=request.get_json(force=True)
        verified=verify_registration_response(
            credential=payload,expected_challenge=base64.urlsafe_b64decode(challenge.encode('ascii')),
            expected_rp_id=rp_id,expected_origin=origin,require_user_verification=True,
        )
        row=WebAuthnCredential.query.filter_by(credential_id=verified.credential_id).first()
        if not row:
            row=WebAuthnCredential(user_id=u.id,credential_id=verified.credential_id,public_key=verified.credential_public_key)
            db.session.add(row)
        row.sign_count=verified.sign_count; row.device_name=str(payload.get('device_name') or 'Windows Hello / fingerprint')[:180]
        row.transports=json.dumps(((payload.get('response') or {}).get('transports') or []))
        u.webauthn_enrolled_at=datetime.datetime.utcnow(); db.session.commit()
        return jsonify(ok=True,message='Fingerprint/passkey enrolled on this device.')
    except Exception as exc:
        db.session.rollback(); return jsonify(ok=False,error=f'Enrollment could not be verified: {exc}'),400

@app.route('/api/webauthn/auth/options',methods=['POST'])
def webauthn_auth_options():
    payload=request.get_json(silent=True) or {}; username=str(payload.get('username') or '').strip()
    u=User.query.filter_by(username=username,active=True).first()
    credentials=WebAuthnCredential.query.filter_by(user_id=u.id).all() if u and u.webauthn_enabled else []
    if not credentials: return jsonify(ok=False,error='No fingerprint/passkey is enrolled for this login ID.'),404
    try:
        from webauthn import generate_authentication_options, options_to_json
        from webauthn.helpers.structs import PublicKeyCredentialDescriptor, UserVerificationRequirement
        rp_id,_=_webauthn_context()
        options=generate_authentication_options(
            rp_id=rp_id,allow_credentials=[PublicKeyCredentialDescriptor(id=c.credential_id) for c in credentials],
            user_verification=UserVerificationRequirement.REQUIRED,
        )
        session['webauthn_auth_challenge']=base64.urlsafe_b64encode(options.challenge).decode('ascii'); session['webauthn_auth_user']=u.id
        return app.response_class(options_to_json(options),mimetype='application/json')
    except Exception as exc:
        return jsonify(ok=False,error=f'Passkey service is unavailable: {exc}'),503

@app.route('/api/webauthn/auth/verify',methods=['POST'])
def webauthn_auth_verify():
    challenge=session.pop('webauthn_auth_challenge',''); uid=session.pop('webauthn_auth_user',None)
    u=db.session.get(User,uid) if uid else None; payload=request.get_json(force=True)
    if not (u and u.active and u.webauthn_enabled and challenge): return jsonify(ok=False,error='Authentication request expired.'),400
    try:
        from webauthn import verify_authentication_response
        credential_id=_b64url_decode(payload.get('id',''))
        row=WebAuthnCredential.query.filter_by(user_id=u.id,credential_id=credential_id).first()
        if not row: return jsonify(ok=False,error='This passkey is not registered.'),404
        rp_id,origin=_webauthn_context()
        verified=verify_authentication_response(
            credential=payload,expected_challenge=base64.urlsafe_b64decode(challenge.encode('ascii')),
            expected_rp_id=rp_id,expected_origin=origin,credential_public_key=row.public_key,
            credential_current_sign_count=row.sign_count,require_user_verification=True,
        )
        row.sign_count=verified.new_sign_count; row.last_used_at=datetime.datetime.utcnow(); db.session.commit()
        session.clear(); session['uid']=u.id; session['kiosk_unlocked']=setting('kiosk_mode_enabled','0')!='1'; session['show_login_welcome']=True
        return jsonify(ok=True,redirect=(url_for('kiosk_lock') if not session['kiosk_unlocked'] else url_for('dashboard')))
    except Exception as exc:
        db.session.rollback(); return jsonify(ok=False,error=f'Fingerprint/passkey verification failed: {exc}'),400

@app.route('/admin/webauthn/<int:credential_id>/delete',methods=['POST'])
@admin_required
def webauthn_credential_delete(credential_id):
    row=db.session.get(WebAuthnCredential,credential_id) or abort(404); uid=row.user_id
    db.session.delete(row); db.session.commit()
    if WebAuthnCredential.query.filter_by(user_id=uid).count()==0:
        u=db.session.get(User,uid); u.webauthn_enrolled_at=None; db.session.commit()
    flash('Registered fingerprint/passkey removed.','success'); return redirect(url_for('admin_panel')+'#user-'+str(uid))

@app.route('/whatsapp',methods=['GET','POST'])
@permission_required('whatsapp')
def whatsapp_workspace():
    if request.method=='GET': return redirect(url_for('integrations_center',category='whatsapp',workflow='whatsapp'))
    to=wa_number(request.form.get('to','')); body=request.form.get('body','').strip()
    if not (to and body): flash('Enter a valid WhatsApp number and message.','danger'); return redirect(url_for('integrations_center',category='whatsapp',workflow='whatsapp'))
    ok,result=whatsapp_cloud_text(to,body)
    if ok:
        mid=result if result.startswith('wamid.') else None; db.session.add(WhatsAppMessage(direction='outbound',wa_id=to,message_id=mid,body=body,status='sent',raw_json=json.dumps({'api_result':result}))); db.session.commit(); flash('WhatsApp message sent.','success')
    else: flash('WhatsApp send failed: '+result,'danger')
    return redirect(url_for('integrations_center',category='whatsapp',workflow='whatsapp'))

@app.route('/webhooks/whatsapp/messages',methods=['GET','POST'])
def whatsapp_messages_webhook():
    if request.method=='GET':
        verify=os.getenv('WHATSAPP_VERIFY_TOKEN',os.getenv('META_VERIFY_TOKEN','')).strip()
        if request.args.get('hub.mode')=='subscribe' and verify and hmac.compare_digest(request.args.get('hub.verify_token',''),verify):
            return request.args.get('hub.challenge','')
        return 'Verification failed',403
    app_secret=os.getenv('META_APP_SECRET','').strip(); raw=request.get_data(cache=True)
    signature=request.headers.get('X-Hub-Signature-256','')
    if app_secret:
        expected='sha256='+hmac.new(app_secret.encode(),raw,hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature,expected): return jsonify(ok=False),403
    payload=request.get_json(silent=True) or {}; changed=0
    for entry in payload.get('entry',[]) or []:
        for change in entry.get('changes',[]) or []:
            value=change.get('value') or {}; contacts={x.get('wa_id'):((x.get('profile') or {}).get('name') or '') for x in value.get('contacts',[]) or []}
            for msg in value.get('messages',[]) or []:
                mid=str(msg.get('id') or '') or None
                row=WhatsAppMessage.query.filter_by(message_id=mid).first() if mid else None
                if not row:
                    mtype=str(msg.get('type') or 'unknown'); content=msg.get('text',{}).get('body','') if mtype=='text' else json.dumps(msg.get(mtype) or {})
                    row=WhatsAppMessage(direction='inbound',contact_name=contacts.get(msg.get('from'),'')[:180],wa_id=str(msg.get('from') or '')[:40],message_id=mid,message_type=mtype,body=content,status='received',raw_json=json.dumps(msg))
                    db.session.add(row); changed+=1
            for status in value.get('statuses',[]) or []:
                row=WhatsAppMessage.query.filter_by(message_id=str(status.get('id') or '')).first()
                if row: row.status=str(status.get('status') or row.status)[:40]; row.raw_json=json.dumps(status); changed+=1
    if changed: db.session.commit()
    return jsonify(ok=True)

@app.route('/integrations/google/connect')
@admin_required
def google_connect():
    if not google_oauth_configured():
        flash('Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET first.','danger'); return redirect(url_for('integrations_center',category='google_email',workflow='email'))
    state=secrets.token_urlsafe(30); session['google_oauth_state']=state
    redirect_uri=url_for('google_callback',_external=True,_scheme='https' if os.getenv('FORCE_HTTPS','1')=='1' else request.scheme)
    params={'client_id':os.getenv('GOOGLE_CLIENT_ID'),'redirect_uri':redirect_uri,'response_type':'code','scope':' '.join(GOOGLE_SCOPES),'access_type':'offline','include_granted_scopes':'true','prompt':'consent','state':state}
    return redirect('https://accounts.google.com/o/oauth2/v2/auth?'+urllib.parse.urlencode(params))

@app.route('/integrations/google/callback')
@admin_required
def google_callback():
    if not hmac.compare_digest(str(session.pop('google_oauth_state','')),str(request.args.get('state',''))):
        flash('Google connection state did not match. Please try again.','danger'); return redirect(url_for('integrations_center',category='google_email',workflow='email'))
    redirect_uri=url_for('google_callback',_external=True,_scheme='https' if os.getenv('FORCE_HTTPS','1')=='1' else request.scheme)
    try:
        r=requests.post('https://oauth2.googleapis.com/token',data={'code':request.args.get('code',''),'client_id':os.getenv('GOOGLE_CLIENT_ID'),'client_secret':os.getenv('GOOGLE_CLIENT_SECRET'),'redirect_uri':redirect_uri,'grant_type':'authorization_code'},timeout=25)
        if not r.ok: raise RuntimeError(r.text[:500])
        data=r.json(); data['expires_at']=datetime.datetime.utcnow().timestamp()+int(data.get('expires_in') or 3600)
        _encrypted_setting_set('google_oauth_token',json.dumps(data)); ensure_google_drive_folder(); flash('Google Drive and Gmail connected.','success')
    except Exception as exc: flash(f'Google connection failed: {exc}','danger')
    return redirect(url_for('integrations_center',category='google_email',workflow='email'))

@app.route('/integrations/google/disconnect',methods=['POST'])
@admin_required
def google_disconnect():
    data=_google_token_data(refresh=False); token=data.get('refresh_token') or data.get('access_token')
    if token:
        try: requests.post('https://oauth2.googleapis.com/revoke',params={'token':token},timeout=15)
        except Exception: pass
    _encrypted_setting_set('google_oauth_token',''); flash('Google connection removed.','success')
    return redirect(url_for('integrations_center',category='google_email',workflow='email'))

@app.route('/admin/google/settings',methods=['POST'])
@admin_required
def google_settings():
    set_setting('google_drive_folder_id',request.form.get('google_drive_folder_id','').strip())
    set_setting('google_drive_auto_backup','1' if request.form.get('google_drive_auto_backup')=='1' else '0')
    flash('Google Drive settings saved.','success'); return redirect(url_for('integrations_center',category='google_email',workflow='email'))

@app.route('/drive',methods=['GET','POST'])
@permission_required('drive')
def drive_workspace():
    if request.method=='GET': return redirect(url_for('integrations_center',category='google_drive',workflow='drive'))
    f=request.files.get('file')
    if not f or not f.filename: flash('Choose a file to upload.','danger'); return redirect(url_for('integrations_center',category='google_drive',workflow='drive'))
    data=f.read(); row,err=google_drive_upload_bytes(data,f.filename,f.mimetype or 'application/octet-stream','manual',current_user()); flash(('Uploaded to Google Drive.' if row else err),('success' if row else 'danger')); return redirect(url_for('integrations_center',category='google_drive',workflow='drive'))

@app.route('/drive/files/<file_id>/download')
@permission_required('drive')
def drive_file_download(file_id):
    if not re.fullmatch(r'[A-Za-z0-9_-]{10,220}',file_id): abort(400)
    headers=_google_headers()
    if not headers: abort(503)
    meta=requests.get(f'https://www.googleapis.com/drive/v3/files/{file_id}',headers=headers,params={'fields':'name,mimeType'},timeout=25)
    mime=meta.json().get('mimeType','') if meta.ok else ''
    exports={
        'application/vnd.google-apps.document':('application/pdf','.pdf'),
        'application/vnd.google-apps.spreadsheet':('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','.xlsx'),
        'application/vnd.google-apps.presentation':('application/vnd.openxmlformats-officedocument.presentationml.presentation','.pptx'),
        'application/vnd.google-apps.drawing':('application/pdf','.pdf'),
    }
    if mime in exports:
        export_mime,extension=exports[mime]
        content=requests.get(f'https://www.googleapis.com/drive/v3/files/{file_id}/export',headers=headers,params={'mimeType':export_mime},timeout=90)
    else:
        export_mime,extension=mime,''
        content=requests.get(f'https://www.googleapis.com/drive/v3/files/{file_id}',headers=headers,params={'alt':'media'},timeout=90)
    if not (meta.ok and content.ok): abort(502)
    info=meta.json(); name=info.get('name') or 'drive-file'
    if extension and not name.lower().endswith(extension): name+=extension
    return send_file(io.BytesIO(content.content),download_name=name,mimetype=export_mime or 'application/octet-stream',as_attachment=True)

def _gmail_plain_text(part):
    if not isinstance(part,dict): return ''
    mime=part.get('mimeType',''); data=(part.get('body') or {}).get('data','')
    if mime=='text/plain' and data:
        try: return _b64url_decode(data).decode('utf-8','replace')
        except Exception: return ''
    for child in part.get('parts',[]) or []:
        text=_gmail_plain_text(child)
        if text: return text
    return ''

@app.route('/email',methods=['GET','POST'])
@permission_required('email')
def email_workspace():
    if request.method=='GET': return redirect(url_for('integrations_center',category='google_email',workflow='email'))
    headers=_google_headers()
    if not headers: flash('Connect Google in Integrations Center first.','danger'); return redirect(url_for('integrations_center',category='google_email',workflow='email'))
    to=request.form.get('to','').strip(); subject=request.form.get('subject','').strip(); body=request.form.get('body','').strip()
    if not (to and body): flash('Recipient and message are required.','danger'); return redirect(url_for('integrations_center',category='google_email',workflow='email'))
    mail=EmailMessage(); mail['To']=to; mail['Subject']=subject; mail.set_content(body); raw=base64.urlsafe_b64encode(mail.as_bytes()).decode('ascii').rstrip('=')
    r=requests.post('https://gmail.googleapis.com/gmail/v1/users/me/messages/send',headers=dict(headers,**{'Content-Type':'application/json'}),json={'raw':raw},timeout=30); flash(('Email sent.' if r.ok else 'Email send failed: '+r.text[:300]),('success' if r.ok else 'danger')); return redirect(url_for('integrations_center',category='google_email',workflow='email'))

@app.route('/email/messages/<message_id>')
@permission_required('email')
def email_message(message_id):
    if not re.fullmatch(r'[A-Za-z0-9_-]{8,220}',message_id): abort(400)
    headers=_google_headers()
    if not headers: abort(503)
    r=requests.get(f'https://gmail.googleapis.com/gmail/v1/users/me/messages/{message_id}',headers=headers,params={'format':'full'},timeout=30)
    if not r.ok: abort(502)
    data=r.json(); hs={x.get('name','').lower():x.get('value','') for x in (data.get('payload',{}).get('headers',[]) or [])}
    message={'from':hs.get('from',''),'to':hs.get('to',''),'subject':hs.get('subject','(no subject)'),'date':hs.get('date',''),'body':_gmail_plain_text(data.get('payload') or {}) or data.get('snippet','')}
    return render_template('email_message.html',message=message)


# ===== Tesla OS 27 • Central Integrations Center =====
INTEGRATION_CATEGORY_LABELS={
    'ai':'AI Services','whatsapp':'WhatsApp','google_email':'Email','google_drive':'Google Drive','food':'Food Partners',
    'electricity':'Electricity & Bharat Connect','banking':'Banking Portals','billing':'Billing / RentOK','payments':'Payments','webhooks':'Webhooks & APIs'
}

def _integration_legacy_settings_snapshot():
    keys=('food_webhook_token','query_webhook_token','google_oauth_token')
    return {k:setting(k,'') for k in keys}

def _integration_workflow_data(workflow):
    workflow=(workflow or '').strip()
    if workflow=='whatsapp':
        return {'configured':whatsapp_cloud_configured(),'messages':WhatsAppMessage.query.order_by(WhatsAppMessage.created_at.desc()).limit(150).all()}
    if workflow=='email':
        headers=_google_headers(); messages=[]; error=''
        if headers:
            try:
                listing=requests.get('https://gmail.googleapis.com/gmail/v1/users/me/messages',headers=headers,params={'maxResults':15},timeout=25)
                for item in (listing.json().get('messages',[]) if listing.ok else []):
                    r=requests.get(f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{item['id']}",headers=headers,params={'format':'metadata','metadataHeaders':['From','Subject','Date']},timeout=20)
                    if not r.ok: continue
                    data=r.json(); hs={x.get('name','').lower():x.get('value','') for x in (data.get('payload',{}).get('headers',[]) or [])}
                    messages.append({'id':item['id'],'from':hs.get('from',''),'subject':hs.get('subject','(no subject)'),'date':hs.get('date',''),'snippet':data.get('snippet','')})
                if not listing.ok: error=listing.text[:400]
            except Exception as exc: error=str(exc)
        return {'connected':bool(headers),'messages':messages,'error':error}
    if workflow=='drive':
        headers=_google_headers(); live=[]; error=''
        if headers:
            try:
                folder=ensure_google_drive_folder(); params={'pageSize':100,'orderBy':'modifiedTime desc','fields':'files(id,name,mimeType,size,modifiedTime,webViewLink)','q':f"'{folder}' in parents and trashed=false" if folder else 'trashed=false'}
                r=requests.get('https://www.googleapis.com/drive/v3/files',headers=headers,params=params,timeout=25)
                if r.ok: live=r.json().get('files',[])
                else: error=r.text[:400]
            except Exception as exc: error=str(exc)
        return {'connected':bool(headers),'files':live,'error':error,'folder_id':setting('google_drive_folder_id','')}
    if workflow in ('food_connections','food_portals'):
        ensure_default_food_integrations(); return {'integrations':FoodIntegration.query.order_by(FoodIntegration.platform,FoodIntegration.display_name).all(),'official':OFFICIAL_FOOD_PORTALS}
    if workflow=='electricity_connections':
        return {'providers':_electricity_provider_rows(include_inactive=False)}
    if workflow=='bank_portals':
        return {'banks':BANK_PORTALS}
    if workflow=='billing_portal': return {}
    if workflow=='ai':
        return {'configured':bool(os.getenv('OPENAI_API_KEY','').strip()) or bool(IntegrationSecretRef.query.join(IntegrationConnection).join(IntegrationProvider).filter(IntegrationProvider.category=='ai',IntegrationConnection.active.is_(True)).first())}
    if workflow=='webhooks': return {}
    return {}

def _integration_center_context(active_category=None,active_provider=None,workflow=None):
    user=current_user(); admin=bool(user and (user.role or '').lower()=='admin'); modules=user_permissions(user)
    rows=IntegrationProvider.query.filter_by(active=True).order_by(IntegrationProvider.category,IntegrationProvider.display_name).all()
    allowed=[]
    for key,label in INTEGRATION_CATEGORY_LABELS.items():
        if user_can_access_category(modules,key,is_admin=admin): allowed.append({'key':key,'label':label})
    allowed_keys={x['key'] for x in allowed}
    if active_category not in allowed_keys: active_category=(allowed[0]['key'] if allowed else None)
    providers=[p for p in rows if p.category==active_category and p.category in allowed_keys]
    provider=None
    if active_provider:
        provider=next((p for p in providers if p.provider_key==active_provider or str(p.id)==str(active_provider)),None)
    if not provider and providers: provider=providers[0]
    connections=IntegrationConnection.query.filter_by(provider_id=provider.id,active=True).order_by(IntegrationConnection.id.desc()).all() if provider else []
    env=dict(os.environ); settings_snapshot=_integration_legacy_settings_snapshot(); db_state={'food_integrations':bool(FoodIntegration.query.first())}
    provider_status={p.provider_key:legacy_connection_status(p.provider_key,env=env,settings=settings_snapshot,db_state=db_state) for p in rows}
    for c in connections:
        provider_status.setdefault(provider.provider_key,{}); provider_status[provider.provider_key]['configured']=True
    workflow_urls={'whatsapp':url_for('integrations_center',category='whatsapp',workflow='whatsapp'),'google_email':url_for('integrations_center',category='google_email',workflow='email'),'google_drive':url_for('integrations_center',category='google_drive',workflow='drive'),'food':url_for('food'),'electricity':url_for('electricity_studio'),'banking':url_for('banking_suite'),'billing':url_for('billing')}
    workflow_templates={'whatsapp':'integration_workflow_whatsapp.html','email':'integration_workflow_google.html','drive':'integration_workflow_google.html','food_portals':'integration_workflow_food.html','food_connections':'integration_workflow_food.html','electricity_connections':'integration_workflow_electricity.html','bank_portals':'integration_workflow_banking.html','billing_portal':'integration_workflow_billing.html','ai':'integration_workflow_ai.html','webhooks':'integration_workflow_webhooks.html'}
    return dict(categories=allowed,active_category=active_category,providers=providers,active_provider=provider,connections=connections,provider_status=provider_status,workflow_url=workflow_urls.get(active_category,''),workflow=workflow,workflow_template=workflow_templates.get(workflow,''),workflow_data=_integration_workflow_data(workflow),portal_embed_url='')

@app.route('/integrations')
@login_required
def integrations_center():
    values={k:v for k,v in {'category':request.args.get('category'),'provider':request.args.get('provider'),'workflow':request.args.get('workflow')}.items() if v}
    return redirect(settings_pane_url('internet-accounts', **values))

@app.route('/integrations/connections/save',methods=['POST'])
@admin_required
def integration_connection_save():
    pid=(request.form.get('provider_id') or '').strip(); provider=db.session.get(IntegrationProvider,int(pid)) if pid.isdigit() else None
    if not provider: abort(404)
    cid=(request.form.get('id') or '').strip(); row=db.session.get(IntegrationConnection,int(cid)) if cid.isdigit() else None
    if row and row.provider_id!=provider.id: abort(400)
    if not row:
        row=IntegrationConnection(provider_id=provider.id,display_name='Connection',created_by_user_id=current_user().id); db.session.add(row)
    row.display_name=(request.form.get('display_name') or provider.display_name)[:180]; row.property_scope=(request.form.get('property_scope') or '')[:180]
    config={}
    if request.form.get('portal_url'): config['portal_url']=request.form.get('portal_url')
    try: config=normalize_nonsecret_config(config)
    except ValueError as exc: flash(str(exc),'danger'); return redirect(url_for('integrations_center',category=provider.category,provider=provider.provider_key))
    row.nonsecret_config_json=json.dumps(config,separators=(',',':')); row.status='configured'; row.updated_by_user_id=current_user().id
    db.session.flush(); record_audit('integration_connection_saved','integration_connection',row.id,module='integrations',meta={'provider_id':provider.id,'category':provider.category}); db.session.commit(); flash('Integration connection saved.','success')
    return redirect(url_for('integrations_center',category=provider.category,provider=provider.provider_key))

@app.route('/integrations/connections/<int:connection_id>/secret',methods=['POST'])
@admin_required
def integration_connection_secret_save(connection_id):
    row=db.session.get(IntegrationConnection,connection_id) or abort(404); admin=current_user()
    if not check_password_hash(admin.password_hash,request.form.get('admin_password','')): flash('Administrator password is required to change integration secrets.','danger'); return redirect(url_for('integrations_center',category=row.provider.category,provider=row.provider.provider_key))
    try: name=validate_integration_secret_name(request.form.get('secret_name',''))
    except ValueError as exc: flash(str(exc),'danger'); return redirect(url_for('integrations_center',category=row.provider.category,provider=row.provider.provider_key))
    value=request.form.get('secret_value') or ''
    if not value: flash('Enter the new secret value.','danger'); return redirect(url_for('integrations_center',category=row.provider.category,provider=row.provider.provider_key))
    master=os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()
    if not master: flash('LIVENZA_VAULT_MASTER_KEY is not configured.','danger'); return redirect(url_for('integrations_center',category=row.provider.category,provider=row.provider.provider_key))
    ref=IntegrationSecretRef.query.filter_by(connection_id=row.id,secret_name=name).first(); vault=db.session.get(VaultSecret,ref.vault_secret_id) if ref else None
    if not vault:
        vault=VaultSecret(secret_type='operational_api_secret',label=f'{row.provider.display_name} • {name}',ciphertext='',nonce='',created_by_user_id=admin.id); db.session.add(vault); db.session.flush()
        if not ref: ref=IntegrationSecretRef(connection_id=row.id,secret_name=name,vault_secret_id=vault.id); db.session.add(ref)
        else: ref.vault_secret_id=vault.id
    vault.secret_type='operational_api_secret'; vault.updated_by_user_id=admin.id; vault.ciphertext,vault.nonce=encrypt_secret(json.dumps({'secret':value}),master)
    row.status='configured'; record_audit('integration_secret_replace','integration_connection',row.id,module='integrations',meta={'provider_id':row.provider_id,'secret_name':name}); db.session.commit(); flash('Integration secret encrypted and saved.','success')
    return redirect(url_for('integrations_center',category=row.provider.category,provider=row.provider.provider_key))

@app.route('/integrations/connections/<int:connection_id>/secret/<name>/reveal',methods=['POST'])
@admin_required
def integration_connection_secret_reveal(connection_id,name):
    row=db.session.get(IntegrationConnection,connection_id) or abort(404); admin=current_user()
    if not check_password_hash(admin.password_hash,request.form.get('admin_password','')): return jsonify(ok=False,error='Administrator password did not match.'),403
    try: name=validate_integration_secret_name(name)
    except ValueError: abort(400)
    ref=IntegrationSecretRef.query.filter_by(connection_id=row.id,secret_name=name).first() or abort(404); vault=db.session.get(VaultSecret,ref.vault_secret_id) or abort(404)
    try: payload=json.loads(decrypt_secret(vault.ciphertext,vault.nonce,os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()))
    except Exception: return jsonify(ok=False,error='Integration secret could not be decrypted.'),422
    record_audit('integration_secret_reveal','integration_connection',row.id,module='integrations',meta={'provider_id':row.provider_id,'secret_name':name}); db.session.commit(); response=jsonify(ok=True,secret=payload.get('secret','')); response.headers['Cache-Control']='no-store, private'; return response

@app.route('/integrations/connections/<int:connection_id>/test',methods=['POST'])
@admin_required
def integration_connection_test(connection_id):
    row=db.session.get(IntegrationConnection,connection_id) or abort(404); provider=row.provider; now=datetime.datetime.utcnow(); status='configured'; message='Configuration is present; provider-side authorization is not probed automatically.'
    config={}
    try: config=json.loads(row.nonsecret_config_json or '{}')
    except Exception: config={}
    if provider.category in ('banking','billing'):
        status='portal_available' if provider_workflow_url(provider,{'nonsecret_config':config}) else 'needs_configuration'; message='Official portal is available.' if status=='portal_available' else 'Add an official portal URL.'
    elif provider.category in ('food','webhooks'):
        status='configured' if (config or provider.portal_url) else 'needs_configuration'; message='Workflow configuration is present.' if status=='configured' else 'No endpoint or portal is configured.'
    else:
        legacy=legacy_connection_status(provider.provider_key,env=os.environ,settings=_integration_legacy_settings_snapshot(),db_state={'food_integrations':bool(FoodIntegration.query.first())})
        refs=IntegrationSecretRef.query.filter_by(connection_id=row.id).count(); status='configured' if (refs or config) else ('configured_legacy' if legacy.get('configured') else 'needs_configuration'); message='Native integration configuration is present.' if status=='configured' else ('Legacy configuration detected.' if status=='configured_legacy' else 'Connection needs configuration.')
    row.last_test_status=status; row.last_test_message=message[:500]; row.last_tested_at=now
    if status in ('configured','configured_legacy','portal_available'):
        row.last_success_status=status; row.last_success_message=message[:500]; row.last_success_at=now; row.status=status
    record_audit('integration_connection_test','integration_connection',row.id,module='integrations',status=status,meta={'provider_id':provider.id}); db.session.commit(); flash(message,'success' if status!='needs_configuration' else 'warning')
    return redirect(url_for('integrations_center',category=provider.category,provider=provider.provider_key))

@app.route('/integrations/connections/<int:connection_id>/archive',methods=['POST'])
@admin_required
def integration_connection_archive(connection_id):
    row=db.session.get(IntegrationConnection,connection_id) or abort(404); row.active=False; row.status='archived'; record_audit('integration_connection_archived','integration_connection',row.id,module='integrations',meta={'provider_id':row.provider_id}); db.session.commit(); flash('Integration connection archived.','success'); return redirect(url_for('integrations_center',category=row.provider.category))

@app.route('/integrations/providers/<int:provider_id>/portal')
@login_required
def integration_provider_portal(provider_id):
    provider=db.session.get(IntegrationProvider,provider_id) or abort(404); user=current_user(); admin=(user.role or '').lower()=='admin'
    if not user_can_access_category(user_permissions(user),provider.category,is_admin=admin): abort(403)
    target=provider_workflow_url(provider)
    if not target: flash('No official portal URL is configured for this provider.','warning'); return redirect(url_for('integrations_center',category=provider.category,provider=provider.provider_key))
    if provider.embed_mode=='inline':
        context=_integration_center_context(provider.category,provider.provider_key); context['portal_embed_url']=target; return render_template('integrations.html',**context)
    return redirect(target)

# ===== Tesla OS 27 • Permission-aware Letterhead sources =====

def _letterhead_actor_is_admin(actor):
    return bool(actor and (actor.role or '').lower()=='admin')

def _letterhead_master_candidate(kind,row,actor):
    allowed=can_access('agreements',actor)
    payload=_master_payload(row) if allowed else {}
    facts={
        'full_name': payload.get('legal_name') or payload.get('full_name') or row.legal_name or row.profile_name,
        'profile_name': row.profile_name,
        'mobile': payload.get('primary_mobile') or row.primary_mobile,
        'email': payload.get('email') or row.email,
        'address': payload.get('permanent_address') or payload.get('current_address') or payload.get('registered_address') or '',
        'city': row.city, 'state': row.state, 'country': row.country,
        'aadhaar': payload.get('aadhaar_no') or payload.get('aadhaar') or '',
        'pan': payload.get('pan_no') or payload.get('pan') or '',
        'passport': payload.get('passport_no') or payload.get('passport') or '',
        'visa': payload.get('visa_no') or payload.get('visa') or '',
        'account_number': payload.get('bank_account_no') or payload.get('account_number') or '',
        'ifsc': payload.get('ifsc') or '',
    }
    docs=[]
    if allowed and _letterhead_actor_is_admin(actor):
        q=MasterDocument.query.filter_by(active=True)
        q=q.filter_by(tenant_master_id=row.id) if kind=='tenant' else q.filter_by(landlord_master_id=row.id)
        docs=[str(d.id) for d in q.all()]
    return {'id':row.id,'display_label':f"{row.profile_name} — {row.city or 'Livenza'}",'facts':facts,'protected_document_ids':docs,'allowed':bool(allowed)}

def _letterhead_source_loaders(actor):
    def tenant_loader(query):
        if not can_access('agreements',actor): return []
        needle=(query.get('name') or query.get('q') or '').strip()
        q=TenantMaster.query.filter_by(active=True)
        if needle: q=q.filter(or_(TenantMaster.profile_name.ilike(f'%{needle}%'),TenantMaster.legal_name.ilike(f'%{needle}%')))
        return [_letterhead_master_candidate('tenant',row,actor) for row in q.order_by(TenantMaster.updated_at.desc()).limit(20).all()]
    def landlord_loader(query):
        if not can_access('agreements',actor): return []
        needle=(query.get('name') or query.get('q') or '').strip()
        q=LandlordMaster.query.filter_by(active=True)
        if needle: q=q.filter(or_(LandlordMaster.profile_name.ilike(f'%{needle}%'),LandlordMaster.legal_name.ilike(f'%{needle}%')))
        return [_letterhead_master_candidate('landlord',row,actor) for row in q.order_by(LandlordMaster.updated_at.desc()).limit(20).all()]
    def room_loader(query):
        if not can_access('rooms',actor): return []
        needle=(query.get('room') or query.get('q') or '').strip(); q=Room.query
        if needle: q=q.filter(or_(Room.room_no.ilike(f'%{needle}%'),Room.property_name.ilike(f'%{needle}%')))
        return [{'id':r.id,'display_label':f'{r.property_name} — Room {r.room_no}','facts':{'property_name':r.property_name,'room_no':r.room_no,'city':r.city,'premises':r.premises,'room_type':r.room_type},'allowed':True} for r in q.limit(20).all()]
    def agreement_loader(query):
        if not can_access('agreements',actor): return []
        needle=(query.get('q') or query.get('name') or '').strip(); q=Agreement.query
        if needle: q=q.filter(Agreement.name.ilike(f'%{needle}%'))
        out=[]
        for row in q.order_by(Agreement.updated_at.desc()).limit(20).all():
            d=row.data; out.append({'id':row.id,'display_label':row.name,'facts':{'agreement_name':row.name,'preset':row.preset,'property_name':d.get('property_name',''),'room_no':d.get('room_no',''),'rent':d.get('monthly_rent','') or d.get('rent',''),'tenant_name':d.get('tenant_name',''),'landlord_name':d.get('landlord_name','')},'allowed':True})
        return out
    def billing_loader(query):
        if not can_access('electricity',actor): return []
        rows=ElectricityBill.query.order_by(ElectricityBill.due_date.desc()).limit(20).all()
        return [{'id':r.id,'display_label':f'Electricity bill {r.bill_no or r.id}','facts':{'bill_no':r.bill_no,'bill_date':str(r.bill_date or ''),'due_date':str(r.due_date or ''),'total_due':r.total_due,'status':r.status},'allowed':True} for r in rows]
    def user_loader(query):
        needle=(query.get('name') or query.get('q') or '').strip()
        q=User.query
        if not _letterhead_actor_is_admin(actor): q=q.filter(User.id==actor.id)
        elif needle: q=q.filter(or_(User.full_name.ilike(f'%{needle}%'),User.username.ilike(f'%{needle}%')))
        return [{'id':r.id,'display_label':r.full_name or r.username,'facts':{'full_name':r.full_name,'username':r.username,'role':r.role},'allowed':True} for r in q.limit(20).all()]
    return {'tenant':tenant_loader,'landlord':landlord_loader,'room':room_loader,'agreement':agreement_loader,'billing':billing_loader,'user':user_loader}

def _letterhead_resolve_sources(actor,query):
    return resolve_sources(actor,query,_letterhead_source_loaders(actor))

def _letterhead_protected_permission(actor,source_kind,source_id):
    if source_kind!='master_document' or not _letterhead_actor_is_admin(actor): return False
    try: doc=db.session.get(MasterDocument,int(source_id))
    except Exception: return False
    return bool(doc and doc.active)

def _letterhead_read_protected_source(actor,source_kind,source_id,document_id=None):
    if not can_access_protected_source(actor,source_kind,source_id,_letterhead_protected_permission): abort(403)
    if source_kind=='master_document':
        doc=db.session.get(MasterDocument,int(source_id)) or abort(404)
        raw=decrypt_blob(doc.ciphertext,doc.nonce,_master_key())
        record_audit('letterhead_protected_source_read',source_kind,doc.id,module='letterhead',meta={'source_kind':source_kind,'source_id':str(doc.id),'purpose':'letterhead_ai_drafting','document_id':document_id}); db.session.commit(); return raw,doc.mime_type
    abort(404)

# ===== Tesla OS 27 • Letterhead Studio editor/review =====

def _letterhead_encrypt_bytes(raw):
    ciphertext,nonce=encrypt_blob(raw,_master_key())
    return json.dumps({'v':'v1','ciphertext':ciphertext,'nonce':nonce},separators=(',',':')).encode('utf-8')

def _letterhead_decrypt_bytes(packed):
    try:
        data=json.loads(bytes(packed or b'').decode('utf-8'))
        return decrypt_blob(data['ciphertext'],data['nonce'],_master_key())
    except Exception as exc:
        raise ValueError('Protected Letterhead asset could not be decrypted.') from exc

def _letterhead_can_edit_document(document,actor):
    return bool(document and actor and (document.creator_user_id==actor.id or has_capability('letterhead_vault_all',actor)))

def _letterhead_current_revision(document):
    return db.session.get(LetterheadDocument,document.id).current_revision_id and db.session.get(LetterheadDocumentRevision,document.current_revision_id)

def _letterhead_attachment_access(actor,source_kind,source_id):
    if source_kind=='master_document': return _letterhead_protected_permission(actor,source_kind,source_id)
    if source_kind=='letterhead_asset':
        try: asset=db.session.get(LetterheadAsset,int(source_id))
        except Exception: return False
        return bool(asset and asset.is_active and (asset.owner_user_id==actor.id or (actor.role or '').lower()=='admin' or has_capability('letterhead_vault_all',actor)))
    return False

def _letterhead_review_errors(document,revision,actor):
    errors=[]
    try: content=json.loads(revision.structured_content_json or '{}')
    except Exception: content={}; errors.append('Document content is not valid structured data.')
    for key in validate_structured_content(content): errors.append('Missing or invalid '+key.replace('_',' ')+'.')
    if not revision.template_version_id: errors.append('Select a published letterhead template.')
    else:
        version=db.session.get(LetterheadTemplateVersion,revision.template_version_id)
        if not version or not template_is_usable(version,actor,{'property_ref':document.property_ref,'entity_ref':document.entity_ref,'document_family':document.document_family}): errors.append('Selected letterhead template is not currently published or permitted.')
    if revision.signature_asset_id:
        signature=db.session.get(SignatureAsset,revision.signature_asset_id)
        if not signature or not signature_is_usable(signature,actor,{'property_ref':document.property_ref,'entity_ref':document.entity_ref,'document_family':document.document_family},datetime.date.today()): errors.append('Selected signature/seal is expired, revoked, or not permitted.')
    for link in DocumentAttachmentLink.query.filter_by(revision_id=revision.id,approved_by_user=True).all():
        if not _letterhead_attachment_access(actor,link.source_kind,link.source_id): errors.append('An approved supporting attachment is no longer accessible.')
    return errors

def _letterhead_integration_readiness():
    ai_ready = bool((_letterhead_ai_provider_config().get('api_key') or '').strip())
    return {
        'ai': {'label': 'AI', 'ready': ai_ready, 'category': 'ai', 'provider': 'openai'},
        'email': {'label': 'Email', 'ready': _letterhead_email_ready(), 'category': 'google_email', 'provider': 'google_email'},
        'whatsapp': {'label': 'WhatsApp', 'ready': _letterhead_whatsapp_ready(), 'category': 'whatsapp', 'provider': 'whatsapp_cloud'},
    }


@app.route('/letterhead')
@permission_required('letterhead')
def letterhead_studio():
    actor=current_user(); all_docs=has_capability('letterhead_vault_all',actor)
    q=LetterheadDocument.query if all_docs else LetterheadDocument.query.filter_by(creator_user_id=actor.id)
    recent=q.order_by(LetterheadDocument.updated_at.desc()).limit(20).all()
    finalized=q.filter(LetterheadDocument.finalized_revision_id.isnot(None)).order_by(LetterheadDocument.updated_at.desc()).limit(8).all()
    pending=LetterheadTemplateVersion.query.filter_by(lifecycle_state='submitted').count() if (actor.role or '').lower()=='admin' else 0
    accessible_ids=[d.current_revision_id for d in recent if d.current_revision_id]
    failed=DocumentDelivery.query.filter(DocumentDelivery.revision_id.in_(accessible_ids),DocumentDelivery.state=='failed').count() if accessible_ids else 0
    published=LetterheadTemplateVersion.query.filter_by(lifecycle_state='published').order_by(LetterheadTemplateVersion.id.desc()).all()
    return render_template('letterhead_studio.html',recent_documents=recent,finalized_documents=finalized,pending_templates=pending,delivery_failures=failed,templates=published,ai_pending=session.get('letterhead_ai_pending'),ai_missing=(session.get('letterhead_ai_pending') or {}).get('missing',[]),integration_readiness=_letterhead_integration_readiness())

@app.route('/letterhead/documents/new',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_document_new():
    actor=current_user(); title=(request.form.get('title') or 'Untitled Livenza Letter').strip()[:240]; family=classify_request(request.form.get('document_type') or title)
    property_ref=(request.form.get('property_or_entity') or '').strip()[:160]
    content={'document_family':family,'title':title,'date':datetime.date.today().isoformat(),'addressee':'To Whom It May Concern','subject':title,'body_sections':[{'type':'paragraph','text':''}],'property_or_entity':property_ref,'source_record_ids':[],'suggested_attachment_ids':[],'source_summary':[]}
    document=LetterheadDocument(title=title,document_family=family,lifecycle_state='draft',creator_user_id=actor.id,property_ref=property_ref,source_refs_json='[]'); db.session.add(document); db.session.flush()
    revision=LetterheadDocumentRevision(document_id=document.id,revision_no=1,structured_content_json=json.dumps(content,ensure_ascii=False),status='draft'); db.session.add(revision); db.session.flush(); document.current_revision_id=revision.id
    record_audit('letterhead_document_created','letterhead_document',document.id,module='letterhead',meta={'document_family':family}); db.session.commit(); return redirect(url_for('letterhead_editor_page',document_id=document.id))

@app.route('/letterhead/supporting-files/upload',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_supporting_upload():
    actor=current_user(); upload=request.files.get('file')
    if not upload or not upload.filename: flash('Choose a supporting file.','danger'); return redirect(request.referrer or url_for('letterhead_studio'))
    raw=upload.read(); mime=(upload.mimetype or 'application/octet-stream')[:80]
    if len(raw)>12*1024*1024: flash('Supporting files must be 12 MB or smaller.','danger'); return redirect(request.referrer or url_for('letterhead_studio'))
    try: packed=_letterhead_encrypt_bytes(raw)
    except Exception as exc: flash(str(exc),'danger'); return redirect(request.referrer or url_for('letterhead_studio'))
    asset=LetterheadAsset(asset_kind='supporting_document',owner_user_id=actor.id,mime_type=mime,encrypted_asset=packed,sha256=hashlib.sha256(raw).hexdigest(),display_name=secure_filename(upload.filename)[:240] or 'supporting-document',is_active=True); db.session.add(asset); db.session.flush()
    did=(request.form.get('document_id') or '').strip()
    if did.isdigit():
        document=db.session.get(LetterheadDocument,int(did))
        if document and _letterhead_can_edit_document(document,actor):
            revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id)
            if revision: db.session.add(DocumentAttachmentLink(revision_id=revision.id,source_kind='letterhead_asset',source_id=str(asset.id),suggested_by_ai=False,approved_by_user=True,approved_by_user_id=actor.id,approved_at=datetime.datetime.utcnow()))
    record_audit('letterhead_supporting_asset_uploaded','letterhead_asset',asset.id,module='letterhead',meta={'mime_type':mime,'size_bytes':len(raw)}); db.session.commit(); flash('Supporting file encrypted and saved.','success'); return redirect(request.referrer or url_for('letterhead_studio'))

@app.route('/letterhead/documents/<int:document_id>/edit')
@capability_required('letterhead_use')
def letterhead_editor_page(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    try: content=json.loads(revision.structured_content_json or '{}')
    except Exception: content={}
    versions=[v for v in LetterheadTemplateVersion.query.filter_by(lifecycle_state='published').order_by(LetterheadTemplateVersion.id.desc()).all() if template_is_usable(v,actor,{'property_ref':document.property_ref,'entity_ref':document.entity_ref,'document_family':document.document_family})]
    signatures=[x for x in SignatureAsset.query.filter_by(is_active=True).order_by(SignatureAsset.signatory_name).all() if signature_is_usable(x,actor,{'property_ref':document.property_ref,'entity_ref':document.entity_ref,'document_family':document.document_family},datetime.date.today())]
    links=DocumentAttachmentLink.query.filter_by(revision_id=revision.id).order_by(DocumentAttachmentLink.id).all()
    suggestions=[]
    for link in links:
        label=f'{link.source_kind} #{link.source_id}'
        if link.source_kind=='master_document':
            try:
                d=db.session.get(MasterDocument,int(link.source_id)); label=d.display_label if d else label
            except Exception: pass
        elif link.source_kind=='letterhead_asset':
            try:
                a=db.session.get(LetterheadAsset,int(link.source_id)); label=a.display_name if a else label
            except Exception: pass
        suggestions.append({'link':link,'label':label,'accessible':_letterhead_attachment_access(actor,link.source_kind,link.source_id)})
    return render_template('letterhead_editor.html',document=document,revision=revision,content=content,template_versions=versions,signatures=signatures,attachments=suggestions,can_ai=has_capability('letterhead_ai',actor))

@app.route('/letterhead/documents/<int:document_id>/autosave',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_autosave(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    if revision.status not in ('draft','review_required'): return jsonify(ok=False,error='Finalized revisions cannot be edited.'),409
    data=request.get_json(silent=True) or {}
    content=data.get('content')
    if not isinstance(content,dict): return jsonify(ok=False,error='Structured content is required.'),400
    # Persist structured JSON only; browser HTML is never authoritative.
    revision.structured_content_json=json.dumps(content,ensure_ascii=False,separators=(',',':')); document.title=str(content.get('title') or document.title)[:240]; document.property_ref=str(content.get('property_or_entity') or '')[:160]; document.updated_at=datetime.datetime.utcnow()
    tid=str(data.get('template_version_id') or '')
    if tid.isdigit():
        version=db.session.get(LetterheadTemplateVersion,int(tid))
        if not version or not template_is_usable(version,actor,{'property_ref':document.property_ref,'entity_ref':document.entity_ref,'document_family':document.document_family}): return jsonify(ok=False,error='Template is not published or permitted.'),400
        revision.template_version_id=version.id
    sid=str(data.get('signature_asset_id') or '')
    if sid.isdigit():
        signature=db.session.get(SignatureAsset,int(sid))
        if not signature or not signature_is_usable(signature,actor,{'property_ref':document.property_ref,'entity_ref':document.entity_ref,'document_family':document.document_family},datetime.date.today()): return jsonify(ok=False,error='Signature is not permitted.'),400
        revision.signature_asset_id=signature.id
    elif sid=='': revision.signature_asset_id=None
    if revision.status=='review_required': revision.status='draft'; document.lifecycle_state='draft'
    db.session.commit(); return jsonify(ok=True,revision_id=revision.id,updated_at=document.updated_at.isoformat()+'Z')

@app.route('/letterhead/documents/<int:document_id>/attachments/decision',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_attachment_decision(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404); data=request.get_json(silent=True) or request.form
    kind=str(data.get('source_kind') or ''); sid=str(data.get('source_id') or ''); approved=str(data.get('approved') or '').lower() in ('1','true','yes','on')
    if approved and not _letterhead_attachment_access(actor,kind,sid): return jsonify(ok=False,error='Attachment is no longer accessible.'),403
    link=DocumentAttachmentLink.query.filter_by(revision_id=revision.id,source_kind=kind,source_id=sid).first()
    if not link: link=DocumentAttachmentLink(revision_id=revision.id,source_kind=kind,source_id=sid); db.session.add(link)
    link.approved_by_user=approved; link.approved_by_user_id=actor.id if approved else None; link.approved_at=datetime.datetime.utcnow() if approved else None
    record_audit('letterhead_attachment_decision','letterhead_document',document.id,module='letterhead',meta={'source_kind':kind,'source_id':sid,'approved':approved}); db.session.commit(); return jsonify(ok=True,approved=approved)

@app.route('/letterhead/documents/<int:document_id>/request-review',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_request_review(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404); errors=_letterhead_review_errors(document,revision,actor)
    if errors:
        for error in errors[:6]: flash(error,'danger')
        return redirect(url_for('letterhead_editor_page',document_id=document.id))
    revision.status='review_required'; document.lifecycle_state='review_required'; record_audit('letterhead_review_requested','letterhead_document',document.id,module='letterhead'); db.session.commit(); return redirect(url_for('letterhead_final_review',document_id=document.id))

@app.route('/letterhead/documents/<int:document_id>/review')
@capability_required('letterhead_use')
def letterhead_final_review(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    if document.lifecycle_state!='review_required' or revision.status!='review_required': flash('Send the document to Final Review from the editor first.','warning'); return redirect(url_for('letterhead_editor_page',document_id=document.id))
    errors=_letterhead_review_errors(document,revision,actor)
    if errors:
        for error in errors[:6]: flash(error,'danger')
        revision.status='draft'; document.lifecycle_state='draft'; db.session.commit(); return redirect(url_for('letterhead_editor_page',document_id=document.id))
    content=json.loads(revision.structured_content_json or '{}'); version=db.session.get(LetterheadTemplateVersion,revision.template_version_id); template=db.session.get(LetterheadTemplate,version.template_id) if version else None; signature=db.session.get(SignatureAsset,revision.signature_asset_id) if revision.signature_asset_id else None
    links=DocumentAttachmentLink.query.filter_by(revision_id=revision.id,approved_by_user=True).all()
    return render_template('letterhead_final_review.html',document=document,revision=revision,content=content,template_version=version,template=template,signature=signature,attachments=links)

# ===== Tesla OS 27 • Immutable Letterhead PDF issuance =====

def _letterhead_layout_asset(layout,location):
    value=None
    if location=='logo': value=(layout.get('header') or {}).get('logo_asset_id')
    elif location=='watermark': value=layout.get('watermark_asset_id')
    elif location=='background': value=layout.get('background_asset_id')
    if value in (None,''): return None
    try: asset=db.session.get(LetterheadAsset,int(value))
    except Exception: return None
    if not asset or not asset.is_active: return None
    try: raw=_letterhead_decrypt_bytes(asset.encrypted_asset)
    except Exception: return None
    return {'bytes':raw,'mime_type':asset.mime_type}

def _letterhead_render_input(document,revision,reference_number):
    content=json.loads(revision.structured_content_json or '{}')
    version=db.session.get(LetterheadTemplateVersion,revision.template_version_id) or abort(409)
    try: layout=json.loads(version.layout_json or '{}')
    except Exception: layout={}
    signature_payload=None
    if revision.signature_asset_id:
        signature=db.session.get(SignatureAsset,revision.signature_asset_id) or abort(409)
        raw=_letterhead_decrypt_bytes(signature.encrypted_asset)
        signature_payload={'bytes':raw,'mime_type':signature.mime_type,'name':signature.signatory_name,'designation':signature.designation}
    return {'template':layout,'document':content,'reference_number':reference_number,'signature':signature_payload,'logo':_letterhead_layout_asset(layout,'logo'),'watermark':_letterhead_layout_asset(layout,'watermark'),'background':_letterhead_layout_asset(layout,'background')}

def _letterhead_annexures(revision,actor):
    result=[]
    for link in DocumentAttachmentLink.query.filter_by(revision_id=revision.id,approved_by_user=True).order_by(DocumentAttachmentLink.id).all():
        if not _letterhead_attachment_access(actor,link.source_kind,link.source_id): raise PermissionError('An approved attachment is no longer accessible.')
        if link.source_kind=='master_document':
            doc=db.session.get(MasterDocument,int(link.source_id)) or abort(404); raw=decrypt_blob(doc.ciphertext,doc.nonce,_master_key()); result.append((doc.mime_type,raw))
        elif link.source_kind=='letterhead_asset':
            asset=db.session.get(LetterheadAsset,int(link.source_id)) or abort(404); raw=_letterhead_decrypt_bytes(asset.encrypted_asset); result.append((asset.mime_type,raw))
    return result

def _letterhead_fiscal_year(on_date=None):
    d=on_date or datetime.date.today(); start=d.year if d.month>=4 else d.year-1; return f'{start}-{str(start+1)[-2:]}'

def _letterhead_allocate_sequence(sequence_key):
    query=DocumentSequence.query.filter_by(sequence_key=sequence_key)
    row=query.with_for_update().first() if db.engine.dialect.name=='postgresql' else query.first()
    if not row:
        row=DocumentSequence(sequence_key=sequence_key,next_value=2); db.session.add(row); db.session.flush(); return 1
    value=max(1,int(row.next_value or 1)); row.next_value=value+1; db.session.flush(); return value

def _letterhead_prior_reference(document,revision):
    prior=LetterheadDocumentRevision.query.filter(LetterheadDocumentRevision.document_id==document.id,LetterheadDocumentRevision.id!=revision.id,LetterheadDocumentRevision.status=='finalized',LetterheadDocumentRevision.reference_number!='').order_by(LetterheadDocumentRevision.revision_no.asc()).first()
    if not prior: return ''
    return re.sub(r'-R\d+$','',prior.reference_number or '')

@app.route('/letterhead/documents/<int:document_id>/finalize',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_document_finalize(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    if document.lifecycle_state!='review_required' or revision.status!='review_required': abort(409)
    errors=_letterhead_review_errors(document,revision,actor)
    if errors:
        for error in errors[:6]: flash(error,'danger')
        return redirect(url_for('letterhead_editor_page',document_id=document.id))
    try:
        annexures=_letterhead_annexures(revision,actor)
        # Rendering must succeed before a number can be consumed.
        draft_pdf=render_letterhead_pdf(_letterhead_render_input(document,revision,'PENDING FINAL REFERENCE'))
        if annexures: merge_annexures(draft_pdf,annexures)
        prior_base=_letterhead_prior_reference(document,revision)
        if prior_base:
            reference_number=f'{prior_base}-R{max(1,revision.revision_no-1)}'
        else:
            fy=_letterhead_fiscal_year(); key=f'{document.property_ref or document.entity_ref or "LIVENZA"}|{document.document_family}|{fy}'
            seq=_letterhead_allocate_sequence(key); prefix=build_reference_prefix(document.property_ref or document.entity_ref or 'Livenza',document.document_family,fy); reference_number=format_reference_number(prefix,seq)
        final_pdf=render_letterhead_pdf(_letterhead_render_input(document,revision,reference_number))
        if annexures: final_pdf=merge_annexures(final_pdf,annexures)
        packed=_letterhead_encrypt_bytes(final_pdf); now=datetime.datetime.utcnow()
        revision.encrypted_pdf=packed; revision.pdf_sha256=sha256_bytes(final_pdf); revision.reference_number=reference_number; revision.status='finalized'; revision.approved_by_user_id=actor.id; revision.approved_at=now; revision.finalized_at=now
        document.lifecycle_state='finalized'; document.finalized_revision_id=revision.id; document.current_revision_id=revision.id; document.updated_at=now
        record_audit('letterhead_document_finalized','letterhead_document',document.id,module='letterhead',meta={'revision_id':revision.id,'reference_number':reference_number,'template_version_id':revision.template_version_id,'attachment_count':len(annexures)})
        db.session.commit()
    except Exception as exc:
        db.session.rollback(); flash('PDF finalization failed. The document remains in Final Review: '+str(exc)[:220],'danger'); return redirect(url_for('letterhead_final_review',document_id=document.id))
    flash('Document approved, finalized and stored in the Document Vault.','success'); return redirect(url_for('letterhead_vault_detail',document_id=document.id))

@app.route('/letterhead/documents/<int:document_id>/revise',methods=['POST'])
@capability_required('letterhead_use')
def letterhead_document_revise(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    source=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    if source.status!='finalized': abort(409)
    next_no=(db.session.query(func.max(LetterheadDocumentRevision.revision_no)).filter_by(document_id=document.id).scalar() or 0)+1
    revision=LetterheadDocumentRevision(document_id=document.id,revision_no=next_no,structured_content_json=source.structured_content_json,template_version_id=source.template_version_id,signature_asset_id=source.signature_asset_id,status='draft'); db.session.add(revision); db.session.flush()
    for link in DocumentAttachmentLink.query.filter_by(revision_id=source.id).all(): db.session.add(DocumentAttachmentLink(revision_id=revision.id,source_kind=link.source_kind,source_id=link.source_id,suggested_by_ai=link.suggested_by_ai,approved_by_user=False))
    document.current_revision_id=revision.id; document.lifecycle_state='draft'; document.updated_at=datetime.datetime.utcnow(); record_audit('letterhead_document_revision_created','letterhead_document',document.id,module='letterhead',meta={'source_revision_id':source.id,'revision_id':revision.id,'revision_no':next_no}); db.session.commit(); flash('A new editable revision was created. The previously issued PDF is unchanged.','success'); return redirect(url_for('letterhead_editor_page',document_id=document.id))


@app.route('/letterhead/documents/<int:document_id>/send/email',methods=['POST'])
@capability_required('letterhead_email_send')
def letterhead_document_send_email(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.finalized_revision_id) if document.finalized_revision_id else None
    if not revision or revision.status!='finalized': abort(409)
    recipient=(request.form.get('recipient') or '').strip()
    if not recipient or '@' not in recipient: flash('Enter a valid email recipient.','danger'); return redirect(url_for('letterhead_vault_detail',document_id=document.id))
    payload=_letterhead_delivery_payload(document,revision,recipient,'email'); result=letterhead_send_email(actor,payload,provider=_letterhead_email_provider if _letterhead_email_ready() else None)
    _letterhead_record_delivery(revision,'email',recipient,result,1); db.session.commit(); flash(('Email accepted by the configured provider.' if result.ok else 'Email delivery failed. You can retry from Document Vault.'),('success' if result.ok else 'danger')); return redirect(url_for('letterhead_vault_detail',document_id=document.id))


@app.route('/letterhead/documents/<int:document_id>/send/whatsapp',methods=['POST'])
@capability_required('letterhead_whatsapp_send')
def letterhead_document_send_whatsapp(document_id):
    actor=current_user(); document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if not _letterhead_can_edit_document(document,actor): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.finalized_revision_id) if document.finalized_revision_id else None
    if not revision or revision.status!='finalized': abort(409)
    recipient=wa_number(request.form.get('recipient',''))
    if not recipient: flash('Enter a valid WhatsApp recipient.','danger'); return redirect(url_for('letterhead_vault_detail',document_id=document.id))
    payload=_letterhead_delivery_payload(document,revision,recipient,'whatsapp'); result=letterhead_send_whatsapp(actor,payload,provider=_letterhead_whatsapp_provider if _letterhead_whatsapp_ready() else None)
    _letterhead_record_delivery(revision,'whatsapp',recipient,result,1); db.session.commit(); flash(('WhatsApp document accepted by the configured provider.' if result.ok else 'WhatsApp delivery failed. You can retry from Document Vault.'),('success' if result.ok else 'danger')); return redirect(url_for('letterhead_vault_detail',document_id=document.id))


@app.route('/letterhead/deliveries/<int:delivery_id>/retry',methods=['POST'])
@permission_required('letterhead')
def letterhead_delivery_retry(delivery_id):
    prior=db.session.get(DocumentDelivery,delivery_id) or abort(404)
    if not can_retry_delivery(prior.state): abort(409)
    revision=db.session.get(LetterheadDocumentRevision,prior.revision_id) or abort(404); document=db.session.get(LetterheadDocument,revision.document_id) or abort(404); actor=current_user()
    if not _letterhead_can_edit_document(document,actor): abort(403)
    if prior.channel=='email':
        if not has_capability('letterhead_email_send',actor): abort(403)
        payload=_letterhead_delivery_payload(document,revision,prior.recipient,'email'); result=letterhead_send_email(actor,payload,provider=_letterhead_email_provider if _letterhead_email_ready() else None)
    elif prior.channel=='whatsapp':
        if not has_capability('letterhead_whatsapp_send',actor): abort(403)
        payload=_letterhead_delivery_payload(document,revision,prior.recipient,'whatsapp'); result=letterhead_send_whatsapp(actor,payload,provider=_letterhead_whatsapp_provider if _letterhead_whatsapp_ready() else None)
    else: abort(400)
    row=_letterhead_record_delivery(revision,prior.channel,prior.recipient,result,int(prior.attempt_no or 1)+1); db.session.commit(); flash(('Delivery retry accepted.' if result.ok else 'Delivery retry failed.'),('success' if result.ok else 'danger')); return redirect(url_for('letterhead_vault_detail',document_id=document.id))



# ===== Tesla OS 27 • Letterhead Document Vault =====

def _letterhead_can_view_document(document, actor):
    if not document or not actor:
        return False
    if document.creator_user_id == actor.id:
        return True
    return has_capability('letterhead_vault_all', actor)


def _letterhead_visible_documents_query(actor):
    query = LetterheadDocument.query.filter(LetterheadDocument.finalized_revision_id.isnot(None))
    if not has_capability('letterhead_vault_all', actor):
        query = query.filter(LetterheadDocument.creator_user_id == actor.id)
    return query


def _letterhead_json(value, fallback):
    try:
        parsed = json.loads(value or '')
        return parsed if isinstance(parsed, type(fallback)) else fallback
    except Exception:
        return fallback


def _letterhead_vault_rows(actor):
    query = _letterhead_visible_documents_query(actor)
    reference = (request.args.get('reference') or '').strip()
    family = (request.args.get('family') or '').strip()
    property_ref = (request.args.get('property') or '').strip()
    recipient = (request.args.get('recipient') or '').strip()
    delivery_state = (request.args.get('delivery_state') or '').strip().lower()
    subject = (request.args.get('subject') or '').strip()
    room = (request.args.get('room') or '').strip()
    creator = (request.args.get('creator') or '').strip()
    approver = (request.args.get('approver') or '').strip()
    template = (request.args.get('template') or '').strip()
    date_from = (request.args.get('date_from') or '').strip()
    date_to = (request.args.get('date_to') or '').strip()

    if family:
        query = query.filter(LetterheadDocument.document_family.ilike(f'%{family}%'))
    if property_ref:
        query = query.filter(LetterheadDocument.property_ref.ilike(f'%{property_ref}%'))
    if subject:
        query = query.filter(or_(LetterheadDocument.title.ilike(f'%{subject}%'), LetterheadDocument.entity_ref.ilike(f'%{subject}%')))
    if creator:
        ids = [u.id for u in User.query.filter(or_(User.full_name.ilike(f'%{creator}%'), User.username.ilike(f'%{creator}%'))).limit(50).all()]
        query = query.filter(LetterheadDocument.creator_user_id.in_(ids or [-1]))

    docs = query.order_by(LetterheadDocument.updated_at.desc()).all()
    rows = []
    for document in docs:
        revision = db.session.get(LetterheadDocumentRevision, document.finalized_revision_id)
        if not revision or revision.status != 'finalized':
            continue
        if reference and reference.lower() not in (revision.reference_number or '').lower():
            continue
        content = _letterhead_json(revision.structured_content_json, {})
        searchable_room = ' '.join(str(content.get(k) or '') for k in ('room','room_no','room_number'))
        if room and room.lower() not in searchable_room.lower():
            continue
        if date_from and revision.finalized_at and revision.finalized_at.date().isoformat() < date_from:
            continue
        if date_to and revision.finalized_at and revision.finalized_at.date().isoformat() > date_to:
            continue
        template_version = db.session.get(LetterheadTemplateVersion, revision.template_version_id) if revision.template_version_id else None
        template_row = db.session.get(LetterheadTemplate, template_version.template_id) if template_version else None
        if template and template.lower() not in ((template_row.name if template_row else '') or '').lower():
            continue
        approver_user = db.session.get(User, revision.approved_by_user_id) if revision.approved_by_user_id else None
        if approver:
            name = ((approver_user.full_name or approver_user.username) if approver_user else '')
            if approver.lower() not in name.lower():
                continue
        deliveries = DocumentDelivery.query.filter_by(revision_id=revision.id).order_by(DocumentDelivery.created_at.desc()).all()
        if recipient and not any(recipient.lower() in (d.recipient or '').lower() for d in deliveries):
            continue
        if delivery_state and not any((d.state or '').lower() == delivery_state for d in deliveries):
            continue
        creator_user = db.session.get(User, document.creator_user_id)
        rows.append({
            'document': document,
            'revision': revision,
            'content': content,
            'template_version': template_version,
            'template': template_row,
            'creator': creator_user,
            'approver': approver_user,
            'deliveries': deliveries,
        })
    return rows


@app.route('/letterhead/vault')
@permission_required('letterhead')
def letterhead_vault():
    actor = current_user()
    rows = _letterhead_vault_rows(actor)
    return render_template('letterhead_vault.html', vault_rows=rows, selected=None, filters=request.args, can_view_all=has_capability('letterhead_vault_all', actor))


@app.route('/letterhead/vault/<int:document_id>')
@permission_required('letterhead')
def letterhead_vault_detail(document_id):
    actor = current_user()
    document = db.session.get(LetterheadDocument, document_id) or abort(404)
    if not _letterhead_can_view_document(document, actor):
        abort(403)
    revisions = LetterheadDocumentRevision.query.filter_by(document_id=document.id).order_by(LetterheadDocumentRevision.revision_no.desc()).all()
    finalized = db.session.get(LetterheadDocumentRevision, document.finalized_revision_id) if document.finalized_revision_id else None
    if not finalized or finalized.status != 'finalized':
        abort(404)
    content = _letterhead_json(finalized.structured_content_json, {})
    template_version = db.session.get(LetterheadTemplateVersion, finalized.template_version_id) if finalized.template_version_id else None
    template_row = db.session.get(LetterheadTemplate, template_version.template_id) if template_version else None
    creator_user = db.session.get(User, document.creator_user_id)
    approver_user = db.session.get(User, finalized.approved_by_user_id) if finalized.approved_by_user_id else None
    attachments = DocumentAttachmentLink.query.filter_by(revision_id=finalized.id, approved_by_user=True).order_by(DocumentAttachmentLink.id).all()
    deliveries = DocumentDelivery.query.filter_by(revision_id=finalized.id).order_by(DocumentDelivery.created_at.desc()).all()
    selected = {
        'document': document,
        'revision': finalized,
        'content': content,
        'template_version': template_version,
        'template': template_row,
        'creator': creator_user,
        'approver': approver_user,
        'attachments': attachments,
        'revisions': revisions,
        'deliveries': deliveries,
    }
    return render_template('letterhead_vault.html', vault_rows=[], selected=selected, filters={}, can_view_all=has_capability('letterhead_vault_all', actor))


@app.route('/letterhead/documents/<int:document_id>/pdf')
@permission_required('letterhead')
def letterhead_document_pdf(document_id):
    actor = current_user()
    document = db.session.get(LetterheadDocument, document_id) or abort(404)
    if not _letterhead_can_view_document(document, actor):
        abort(403)
    revision = db.session.get(LetterheadDocumentRevision, document.finalized_revision_id) if document.finalized_revision_id else None
    if not revision or revision.status != 'finalized' or not revision.encrypted_pdf:
        abort(404)
    raw = _letterhead_decrypt_bytes(revision.encrypted_pdf)
    safe_ref = re.sub(r'[^A-Za-z0-9._-]+', '_', revision.reference_number or f'Livenza_Document_{document.id}').strip('._') or f'Livenza_Document_{document.id}'
    response = send_file(io.BytesIO(raw), mimetype='application/pdf', as_attachment=True, download_name=f'{safe_ref}.pdf', max_age=0)
    response.headers['Cache-Control'] = 'no-store, private'
    response.headers['Pragma'] = 'no-cache'
    record_audit('letterhead_pdf_downloaded', 'letterhead_document', document.id, module='letterhead', meta={'document_id': document.id, 'revision_id': revision.id, 'reference_number': revision.reference_number})
    db.session.commit()
    return response


# ===== Tesla OS 27 • Ask Livenza AI =====

def _letterhead_integration_secret(provider_key, secret_names):
    provider=IntegrationProvider.query.filter_by(provider_key=provider_key,active=True).first()
    if not provider: return ''
    connections=IntegrationConnection.query.filter_by(provider_id=provider.id,active=True).order_by(IntegrationConnection.id.desc()).all()
    master=os.getenv('LIVENZA_VAULT_MASTER_KEY','').strip()
    if not master: return ''
    for connection in connections:
        for name in secret_names:
            ref=IntegrationSecretRef.query.filter_by(connection_id=connection.id,secret_name=name).first()
            if not ref: continue
            vault=db.session.get(VaultSecret,ref.vault_secret_id)
            if not vault: continue
            try:
                payload=json.loads(decrypt_secret(vault.ciphertext,vault.nonce,master)); value=str(payload.get('secret') or '').strip()
                if value: return value
            except Exception: continue
    return ''

def _letterhead_ai_provider_config():
    key=_letterhead_integration_secret('openai',('openai_api_key','api_key')) or os.getenv('OPENAI_API_KEY','').strip()
    return {'api_key':key,'model':os.getenv('OPENAI_LETTERHEAD_MODEL',os.getenv('OPENAI_HELP_MODEL','gpt-5.6-luna'))}


def _letterhead_integration_connection_config(provider_key):
    provider=IntegrationProvider.query.filter_by(provider_key=provider_key,active=True).first()
    if not provider: return {}
    row=IntegrationConnection.query.filter_by(provider_id=provider.id,active=True).order_by(IntegrationConnection.id.desc()).first()
    if not row: return {}
    try:
        config=json.loads(row.nonsecret_config_json or '{}')
        return config if isinstance(config,dict) else {}
    except Exception:
        return {}


def _letterhead_email_ready():
    return bool(_google_token_data(refresh=False).get('access_token'))


def _letterhead_whatsapp_config():
    config=_letterhead_integration_connection_config('whatsapp_cloud')
    token=_letterhead_integration_secret('whatsapp_cloud',('whatsapp_access_token','access_token','api_token')) or os.getenv('WHATSAPP_CLOUD_TOKEN','').strip()
    phone_id=str(config.get('phone_number_id') or os.getenv('WHATSAPP_PHONE_NUMBER_ID','')).strip()
    return {'token':token,'phone_number_id':phone_id,'graph_version':os.getenv('WHATSAPP_GRAPH_VERSION','v23.0').strip() or 'v23.0'}


def _letterhead_whatsapp_ready():
    cfg=_letterhead_whatsapp_config(); return bool(cfg.get('token') and cfg.get('phone_number_id'))


def _letterhead_email_provider(payload):
    headers=_google_headers()
    if not headers: return {'accepted':False,'provider':'google_email','error_code':'integration_not_configured'}
    try:
        mail=EmailMessage(); mail['To']=payload['recipient']; mail['Subject']=payload.get('subject') or 'Livenza Life Document'
        mail.set_content(payload.get('message') or 'Please find the approved Livenza Life document attached.')
        mail.add_attachment(payload['pdf_bytes'],maintype='application',subtype='pdf',filename=payload.get('filename') or 'livenza-document.pdf')
        raw=base64.urlsafe_b64encode(mail.as_bytes()).decode('ascii').rstrip('=')
        r=requests.post('https://gmail.googleapis.com/gmail/v1/users/me/messages/send',headers=dict(headers,**{'Content-Type':'application/json'}),json={'raw':raw},timeout=30)
        if not r.ok: return {'accepted':False,'provider':'google_email','error_code':f'http_{r.status_code}'}
        try: reference=str(r.json().get('id') or '')
        except Exception: reference=''
        return {'accepted':True,'provider':'google_email','reference':reference}
    except Exception:
        return {'accepted':False,'provider':'google_email','error_code':'provider_exception'}


def _letterhead_whatsapp_provider(payload):
    cfg=_letterhead_whatsapp_config(); to=wa_number(payload.get('recipient',''))
    if not (cfg.get('token') and cfg.get('phone_number_id') and to): return {'accepted':False,'provider':'whatsapp_cloud','error_code':'integration_not_configured'}
    base=f"https://graph.facebook.com/{cfg['graph_version']}/{cfg['phone_number_id']}"; auth={'Authorization':f"Bearer {cfg['token']}"}
    try:
        upload=requests.post(base+'/media',headers=auth,data={'messaging_product':'whatsapp'},files={'file':(payload.get('filename') or 'livenza-document.pdf',payload['pdf_bytes'],'application/pdf')},timeout=30)
        if not upload.ok: return {'accepted':False,'provider':'whatsapp_cloud','error_code':f'upload_http_{upload.status_code}'}
        media_id=str(upload.json().get('id') or '')
        if not media_id: return {'accepted':False,'provider':'whatsapp_cloud','error_code':'upload_missing_media_id'}
        body={'messaging_product':'whatsapp','to':to,'type':'document','document':{'id':media_id,'filename':payload.get('filename') or 'livenza-document.pdf','caption':str(payload.get('message') or 'Approved Livenza Life document')[:1024]}}
        sent=requests.post(base+'/messages',headers=dict(auth,**{'Content-Type':'application/json'}),json=body,timeout=25)
        if not sent.ok: return {'accepted':False,'provider':'whatsapp_cloud','error_code':f'send_http_{sent.status_code}'}
        try: reference=str((sent.json().get('messages') or [{}])[0].get('id') or media_id)
        except Exception: reference=media_id
        return {'accepted':True,'provider':'whatsapp_cloud','reference':reference}
    except Exception:
        return {'accepted':False,'provider':'whatsapp_cloud','error_code':'provider_exception'}


def _letterhead_safe_pdf_filename(reference):
    name=re.sub(r'[^A-Za-z0-9._-]+','-',str(reference or 'livenza-document')).strip('-') or 'livenza-document'
    return name+'.pdf'


def _letterhead_delivery_payload(document,revision,recipient,channel):
    if revision.status!='finalized' or not revision.encrypted_pdf: raise ValueError('Only a finalized document can be sent.')
    pdf=_letterhead_decrypt_bytes(revision.encrypted_pdf)
    label=revision.reference_number or document.title
    return {'recipient':recipient,'pdf_bytes':pdf,'filename':_letterhead_safe_pdf_filename(label),'subject':f'{document.title} · {label}','message':f'Please find attached the approved Livenza Life document {label}.'}


def _letterhead_record_delivery(revision,channel,recipient,result,attempt_no):
    row=DocumentDelivery(revision_id=revision.id,channel=channel,recipient=recipient[:320],state=result.state,provider_name=result.provider_name[:80],provider_reference=result.provider_reference[:240],attempt_no=attempt_no,error_code=result.error_code[:120],completed_at=datetime.datetime.utcnow())
    db.session.add(row); db.session.flush()
    record_audit('letterhead_document_delivery','document_delivery',row.id,module='letterhead',status=('success' if result.ok else 'failed'),meta=audit_safe_metadata({'revision_id':revision.id,'channel':channel,'recipient':recipient,'provider_name':result.provider_name,'provider_reference':result.provider_reference,'state':result.state,'attempt_no':attempt_no,'error_code':result.error_code}))
    return row

def _letterhead_search_hint(text):
    text=str(text or '')
    m=re.search(r'\bfor\s+([A-Z][A-Za-z .\'-]{1,80})',text)
    return (m.group(1).strip(' .') if m else text[:80]).strip()

def _letterhead_merge_source_facts(candidates, extra=None):
    facts=dict(extra or {}); summaries=[]; source_ids=[]; attachments=set()
    for candidate in candidates:
        source_ids.append(f'{candidate.kind}:{candidate.record_id}')
        for key,value in (candidate.facts or {}).items():
            if value not in (None,'',[]) and key not in facts: facts[key]=value
        summaries.append(f'{candidate.display_label} from {candidate.kind.title()} record')
        for did in candidate.protected_document_ids: attachments.add(f'master_document:{did}')
    facts['source_record_ids']=source_ids; facts['source_summary']=summaries
    if not facts.get('property_or_entity'): facts['property_or_entity']=facts.get('property_name') or facts.get('entity_name') or ''
    return facts,attachments

@app.route('/letterhead/ai/start',methods=['POST'])
@capability_required('letterhead_ai')
def letterhead_ai_start():
    actor=current_user(); text=(request.form.get('request_text') or '').strip()
    if not text: flash('Tell Ask Livenza AI what document you need.','danger'); return redirect(url_for('letterhead_studio'))
    family=classify_request(text); hint=(request.form.get('search_name') or _letterhead_search_hint(text)).strip()
    candidates=_letterhead_resolve_sources(actor,{'q':hint,'name':hint})
    requested={'full_name','profile_name','mobile','email','address','city','state','country','property_name','room_no','premises','agreement_name','rent','bill_no','bill_date','due_date','total_due','status','role'}
    minimized=[minimize_for_ai(c,requested) for c in candidates]
    extra={'date':(request.form.get('date') or datetime.date.today().isoformat()),'property_or_entity':(request.form.get('property_or_entity') or '').strip(),'full_name':(request.form.get('full_name') or '').strip()}
    facts,allowed_attachments=_letterhead_merge_source_facts(candidates,extra)
    missing=missing_required_fields(family,facts)
    if missing:
        session['letterhead_ai_pending']={'request_text':text,'family':family,'hint':hint,'facts':facts,'missing':missing}
        return render_template('letterhead_studio.html',ai_pending=session['letterhead_ai_pending'],ai_missing=missing,templates=[],recent_documents=[],pending_templates=0,delivery_failures=0)
    allowed_source_ids={f'{c.kind}:{c.record_id}' for c in candidates}
    client=get_ai_client(actor,_letterhead_ai_provider_config())
    payload=None
    if client.available:
        try:
            payload=client.generate_json(build_ai_draft_request(text,family,minimized,allowed_attachments,extra_facts={k:v for k,v in facts.items() if k not in ('source_record_ids','source_summary')}))
            payload=parse_structured_draft(payload,allowed_source_ids,allowed_attachments)
        except Exception as exc:
            flash('AI drafting could not complete; a reviewable local draft was created instead. '+str(exc)[:180],'warning')
    if payload is None:
        payload=deterministic_draft(family,facts,text); payload['source_record_ids']=list(allowed_source_ids); payload=parse_structured_draft(payload,allowed_source_ids,allowed_attachments)
    document=LetterheadDocument(title=payload['title'],document_family=family,lifecycle_state='draft',creator_user_id=actor.id,property_ref=str(payload.get('property_or_entity') or '')[:160],source_refs_json=json.dumps(payload.get('source_record_ids') or [])); db.session.add(document); db.session.flush()
    revision=LetterheadDocumentRevision(document_id=document.id,revision_no=1,structured_content_json=json.dumps(payload,ensure_ascii=False),status='draft'); db.session.add(revision); db.session.flush(); document.current_revision_id=revision.id
    for aid in payload.get('suggested_attachment_ids',[]):
        kind,_,sid=aid.partition(':'); db.session.add(DocumentAttachmentLink(revision_id=revision.id,source_kind=kind,source_id=sid,suggested_by_ai=True,approved_by_user=False))
    record_audit('letterhead_ai_draft_created','letterhead_document',document.id,module='letterhead',meta=audit_safe_metadata({'document_family':family,'source_record_ids':payload.get('source_record_ids',[])})); db.session.commit(); session.pop('letterhead_ai_pending',None)
    return redirect(url_for('letterhead_editor_page',document_id=document.id))

@app.route('/letterhead/ai/resolve',methods=['POST'])
@capability_required('letterhead_ai')
def letterhead_ai_resolve():
    pending=session.get('letterhead_ai_pending') or {}; facts=pending.get('facts') or {}
    for field in pending.get('missing') or []:
        value=(request.form.get(field) or '').strip()
        if value: facts[field]=value
    session['letterhead_ai_pending']={**pending,'facts':facts,'missing':missing_required_fields(pending.get('family','custom'),facts)}
    if session['letterhead_ai_pending']['missing']:
        return redirect(url_for('letterhead_studio'))
    # Re-submit through the canonical start route contract using a server-side temporary form is avoided; create a deterministic reviewed draft here.
    payload=deterministic_draft(pending.get('family','custom'),facts,pending.get('request_text',''))
    document=LetterheadDocument(title=payload['title'],document_family=payload['document_family'],lifecycle_state='draft',creator_user_id=current_user().id,property_ref=str(payload.get('property_or_entity') or '')[:160],source_refs_json=json.dumps(payload.get('source_record_ids') or [])); db.session.add(document); db.session.flush()
    revision=LetterheadDocumentRevision(document_id=document.id,revision_no=1,structured_content_json=json.dumps(payload,ensure_ascii=False),status='draft'); db.session.add(revision); db.session.flush(); document.current_revision_id=revision.id; record_audit('letterhead_ai_clarification_resolved','letterhead_document',document.id,module='letterhead'); db.session.commit(); session.pop('letterhead_ai_pending',None); return redirect(url_for('letterhead_editor_page',document_id=document.id))

@app.route('/letterhead/documents/<int:document_id>/rewrite/<action>',methods=['POST'])
@capability_required('letterhead_ai')
def letterhead_ai_rewrite(document_id,action):
    document=db.session.get(LetterheadDocument,document_id) or abort(404)
    if document.creator_user_id!=current_user().id and not has_capability('letterhead_vault_all'): abort(403)
    revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    if revision.status not in ('draft','review_required'): abort(409)
    content=json.loads(revision.structured_content_json or '{}'); client=get_ai_client(current_user(),_letterhead_ai_provider_config())
    if not client.available: flash('AI integration is not configured in Integrations Center.','danger'); return redirect(url_for('letterhead_editor_page',document_id=document.id))
    try: content=rewrite_action(content,action,client)
    except Exception as exc: flash('Rewrite failed: '+str(exc)[:220],'danger'); return redirect(url_for('letterhead_editor_page',document_id=document.id))
    revision.structured_content_json=json.dumps(content,ensure_ascii=False); document.updated_at=datetime.datetime.utcnow(); record_audit('letterhead_ai_rewrite','letterhead_document',document.id,module='letterhead',meta={'action':action}); db.session.commit(); flash('AI rewrite applied. Review the wording before finalizing.','success'); return redirect(url_for('letterhead_editor_page',document_id=document.id))

@app.route('/letterhead/documents/<int:document_id>/regenerate',methods=['POST'])
@capability_required('letterhead_ai')
def letterhead_ai_regenerate(document_id):
    document=db.session.get(LetterheadDocument,document_id) or abort(404); revision=db.session.get(LetterheadDocumentRevision,document.current_revision_id) or abort(404)
    if document.creator_user_id!=current_user().id and not has_capability('letterhead_vault_all'): abort(403)
    content=json.loads(revision.structured_content_json or '{}'); client=get_ai_client(current_user(),_letterhead_ai_provider_config())
    if not client.available: flash('AI integration is not configured in Integrations Center.','danger'); return redirect(url_for('letterhead_editor_page',document_id=document.id))
    try: content=rewrite_action(content,'make_formal',client)
    except Exception as exc: flash('Regeneration failed: '+str(exc)[:220],'danger'); return redirect(url_for('letterhead_editor_page',document_id=document.id))
    revision.structured_content_json=json.dumps(content,ensure_ascii=False); record_audit('letterhead_ai_regenerated','letterhead_document',document.id,module='letterhead'); db.session.commit(); return redirect(url_for('letterhead_editor_page',document_id=document.id))

@app.route('/letterhead/templates')
@permission_required('letterhead')
def letterhead_templates_page():
    actor=current_user(); templates=LetterheadTemplate.query.order_by(LetterheadTemplate.updated_at.desc()).all()
    versions={}
    for row in LetterheadTemplateVersion.query.order_by(LetterheadTemplateVersion.template_id,LetterheadTemplateVersion.version_no.desc()).all(): versions.setdefault(row.template_id,[]).append(row)
    assets=LetterheadAsset.query.filter_by(is_active=True).order_by(LetterheadAsset.created_at.desc()).limit(50).all()
    signatures=SignatureAsset.query.filter_by(is_active=True).order_by(SignatureAsset.signatory_name).all()
    return render_template('letterhead_templates.html',templates=templates,versions=versions,assets=assets,signatures=signatures,can_author=has_capability('letterhead_template_author',actor),can_submit=has_capability('letterhead_template_submit',actor))

@app.route('/letterhead/templates/<int:template_id>/versions/<int:version_id>/edit')
@capability_required('letterhead_template_author')
def letterhead_template_editor_page(template_id,version_id):
    template,version=_letterhead_template_version(template_id,version_id)
    if version.lifecycle_state!='draft': flash('Published/submitted template versions are immutable. Duplicate the template to create a new draft.','warning'); return redirect(url_for('letterhead_templates_page'))
    try: layout=json.loads(version.layout_json or '{}')
    except Exception: layout={}
    try: scope=json.loads(version.scope_json or '{}')
    except Exception: scope={}
    assets=LetterheadAsset.query.filter_by(is_active=True).order_by(LetterheadAsset.created_at.desc()).all()
    return render_template('letterhead_template_editor.html',template=template,version=version,layout=layout,scope=scope,assets=assets,can_submit=has_capability('letterhead_template_submit'))

@app.route('/letterhead/templates/<int:template_id>/versions/<int:version_id>/preview.json')
@permission_required('letterhead')
def letterhead_template_preview_json(template_id,version_id):
    template,version=_letterhead_template_version(template_id,version_id)
    try: layout=json.loads(version.layout_json or '{}')
    except Exception: layout={}
    return jsonify(ok=True,template={'id':template.id,'name':template.name,'status':template.status},version={'id':version.id,'version_no':version.version_no,'state':version.lifecycle_state,'layout':layout})

@app.route('/letterhead/template-assets/upload',methods=['POST'])
@capability_required('letterhead_template_author')
def letterhead_template_asset_upload():
    actor=current_user(); upload=request.files.get('file')
    if not upload or not upload.filename: flash('Choose an artwork file.','danger'); return redirect(url_for('letterhead_templates_page'))
    raw=upload.read(); mime=(upload.mimetype or 'application/octet-stream')[:80]
    try: validate_template_asset(upload.filename,mime,len(raw)); packed=_letterhead_encrypt_bytes(raw)
    except Exception as exc: flash(str(exc),'danger'); return redirect(url_for('letterhead_templates_page'))
    asset=LetterheadAsset(asset_kind=(request.form.get('asset_kind') or 'template_artwork')[:40],owner_user_id=actor.id,mime_type=mime,encrypted_asset=packed,sha256=hashlib.sha256(raw).hexdigest(),display_name=secure_filename(upload.filename)[:240] or 'letterhead-artwork',is_active=True); db.session.add(asset); db.session.flush(); record_audit('letterhead_template_asset_uploaded','letterhead_asset',asset.id,module='letterhead',meta={'mime_type':mime,'size_bytes':len(raw)}); db.session.commit(); flash('Template artwork encrypted and added to the library.','success'); return redirect(url_for('letterhead_templates_page'))

@app.route('/letterhead/signatures/upload',methods=['POST'])
@admin_required
def letterhead_signature_upload():
    upload=request.files.get('file'); actor=current_user()
    if not upload or not upload.filename: flash('Choose a signature or seal image.','danger'); return redirect(url_for('letterhead_templates_page'))
    raw=upload.read(); mime=(upload.mimetype or 'application/octet-stream')[:80]
    if mime not in ('image/png','image/jpeg','image/webp') or len(raw)>5*1024*1024: flash('Signature/seal must be PNG, JPG or WebP up to 5 MB.','danger'); return redirect(url_for('letterhead_templates_page'))
    try: packed=_letterhead_encrypt_bytes(raw)
    except Exception as exc: flash(str(exc),'danger'); return redirect(url_for('letterhead_templates_page'))
    scope={'roles':request.form.getlist('roles'),'properties':[x.strip() for x in (request.form.get('properties') or '').split(',') if x.strip()],'entities':[],'document_families':[x.strip() for x in (request.form.get('document_families') or '').split(',') if x.strip()]}
    def date_or_none(value):
        try: return datetime.date.fromisoformat(value) if value else None
        except Exception: return None
    row=SignatureAsset(asset_kind=(request.form.get('asset_kind') or 'signature')[:24],signatory_name=(request.form.get('signatory_name') or 'Authorized Signatory')[:160],designation=(request.form.get('designation') or '')[:160],scope_json=json.dumps(scope,separators=(',',':')),encrypted_asset=packed,mime_type=mime,effective_date=date_or_none(request.form.get('effective_date')),expires_at=date_or_none(request.form.get('expires_at')),is_active=True,created_by_user_id=actor.id); db.session.add(row); db.session.flush(); record_audit('letterhead_signature_asset_uploaded','signature_asset',row.id,module='letterhead',meta={'asset_kind':row.asset_kind,'signatory_name':row.signatory_name}); db.session.commit(); flash('Protected signature/seal saved.','success'); return redirect(url_for('letterhead_templates_page'))

@app.route('/letterhead/signatures/<int:signature_id>/revoke',methods=['POST'])
@admin_required
def letterhead_signature_revoke(signature_id):
    row=db.session.get(SignatureAsset,signature_id) or abort(404); row.is_active=False; row.revoked_at=datetime.datetime.utcnow(); record_audit('letterhead_signature_revoked','signature_asset',row.id,module='letterhead',meta={'signatory_name':row.signatory_name}); db.session.commit(); flash('Signature/seal revoked.','success'); return redirect(url_for('letterhead_templates_page'))

# ===== Tesla OS 27 • Letterhead Template lifecycle =====

def _letterhead_template_version(template_id, version_id=None):
    template=db.session.get(LetterheadTemplate,template_id) or abort(404)
    version=db.session.get(LetterheadTemplateVersion,version_id) if version_id else None
    if version and version.template_id != template.id: abort(404)
    return template,version

@app.route('/letterhead/templates/create',methods=['POST'])
@capability_required('letterhead_template_author')
def letterhead_template_create():
    actor=current_user(); name=(request.form.get('name') or 'Untitled Letterhead').strip()[:160]
    slug=re.sub(r'[^a-z0-9]+','-',(request.form.get('slug') or name).lower()).strip('-')[:180] or ('template-'+uuid.uuid4().hex[:8])
    base=slug; n=2
    while LetterheadTemplate.query.filter_by(slug=slug).first(): slug=f'{base}-{n}'; n+=1
    row=LetterheadTemplate(name=name,slug=slug,entity_scope=(request.form.get('entity_scope') or '')[:160],property_scope=(request.form.get('property_scope') or '')[:160],document_family_scope=(request.form.get('document_family_scope') or '')[:160],status='draft',created_by_user_id=actor.id)
    db.session.add(row); db.session.flush()
    layout={'page_size':'A4','margins_mm':{'top':38,'right':18,'bottom':24,'left':18},'header':{'brand':'Livenza Life','show_logo':True},'footer':{'text':'Livenza Life LLP','page_numbers':True}}
    version=LetterheadTemplateVersion(template_id=row.id,version_no=1,lifecycle_state='draft',layout_json=json.dumps(layout),scope_json='{}',content_hash=hashlib.sha256(json.dumps(layout,sort_keys=True).encode()).hexdigest())
    db.session.add(version); record_audit('letterhead_template_created','letterhead_template',row.id,module='letterhead',meta={'version_no':1}); db.session.commit()
    flash('Letterhead template draft created.','success'); return redirect(url_for('letterhead_template_editor_page',template_id=row.id,version_id=version.id))

@app.route('/letterhead/templates/<int:template_id>/versions/<int:version_id>/save',methods=['POST'])
@capability_required('letterhead_template_author')
def letterhead_template_save(template_id,version_id):
    template,version=_letterhead_template_version(template_id,version_id)
    if version.lifecycle_state!='draft': abort(409)
    try: layout=json.loads(request.form.get('layout_json') or '{}'); scope=json.loads(request.form.get('scope_json') or '{}')
    except Exception: flash('Template layout or scope is invalid JSON.','danger'); return redirect(url_for('letterhead_template_editor_page',template_id=template.id,version_id=version.id))
    version.layout_json=json.dumps(layout,ensure_ascii=False,separators=(',',':')); version.scope_json=json.dumps(scope,ensure_ascii=False,separators=(',',':')); version.content_hash=hashlib.sha256(version.layout_json.encode()).hexdigest(); template.updated_at=datetime.datetime.utcnow()
    record_audit('letterhead_template_edited','letterhead_template',template.id,module='letterhead',meta={'version_no':version.version_no}); db.session.commit(); flash('Template draft saved.','success'); return redirect(url_for('letterhead_template_editor_page',template_id=template.id,version_id=version.id))

@app.route('/letterhead/templates/<int:template_id>/versions/<int:version_id>/submit',methods=['POST'])
@capability_required('letterhead_template_submit')
def letterhead_template_submit(template_id,version_id):
    template,version=_letterhead_template_version(template_id,version_id)
    if version.lifecycle_state!='draft': abort(409)
    version.lifecycle_state='submitted'; version.submitted_by_user_id=current_user().id; version.submitted_at=datetime.datetime.utcnow(); template.status='submitted'
    record_audit('letterhead_template_submitted','letterhead_template',template.id,module='letterhead',meta={'version_no':version.version_no}); db.session.commit(); flash('Template submitted for Admin approval.','success'); return redirect(url_for('letterhead_templates_page'))

@app.route('/letterhead/templates/<int:template_id>/versions/<int:version_id>/publish',methods=['POST'])
@admin_required
def letterhead_template_publish(template_id,version_id):
    template,version=_letterhead_template_version(template_id,version_id)
    if version.lifecycle_state!='submitted': abort(409)
    try:
        normalized_layout=json.dumps(json.loads(version.layout_json or '{}'),ensure_ascii=False,sort_keys=True,separators=(',',':'))
    except Exception: abort(400)
    version.layout_json=normalized_layout; version.content_hash=hashlib.sha256(normalized_layout.encode('utf-8')).hexdigest()
    previous=db.session.get(LetterheadTemplateVersion,template.current_published_version_id) if template.current_published_version_id else None
    if previous and previous.id != version.id and previous.lifecycle_state=='published': previous.lifecycle_state='superseded'
    version.lifecycle_state='published'; version.published_by_user_id=current_user().id; version.published_at=datetime.datetime.utcnow(); template.current_published_version_id=version.id; template.status='published'
    record_audit('letterhead_template_published','letterhead_template',template.id,module='letterhead',meta={'version_no':version.version_no}); db.session.commit(); flash('Template published.','success'); return redirect(url_for('letterhead_templates_page'))

@app.route('/letterhead/templates/<int:template_id>/versions/<int:version_id>/reject',methods=['POST'])
@admin_required
def letterhead_template_reject(template_id,version_id):
    template,version=_letterhead_template_version(template_id,version_id)
    if version.lifecycle_state!='submitted': abort(409)
    version.lifecycle_state='draft'; version.rejection_comment=(request.form.get('comment') or '')[:2000]; template.status='draft'; record_audit('letterhead_template_rejected','letterhead_template',template.id,module='letterhead',meta={'version_no':version.version_no}); db.session.commit(); flash('Template returned to Draft.','warning'); return redirect(url_for('letterhead_templates_page'))

@app.route('/letterhead/templates/<int:template_id>/archive',methods=['POST'])
@admin_required
def letterhead_template_archive(template_id):
    template=db.session.get(LetterheadTemplate,template_id) or abort(404); template.status='archived'
    if template.current_published_version_id:
        version=db.session.get(LetterheadTemplateVersion,template.current_published_version_id)
        if version and version.lifecycle_state in ('published','superseded'): version.lifecycle_state='archived'
    record_audit('letterhead_template_archived','letterhead_template',template.id,module='letterhead'); db.session.commit(); flash('Template archived.','success'); return redirect(url_for('letterhead_templates_page'))

@app.route('/letterhead/templates/<int:template_id>/duplicate',methods=['POST'])
@capability_required('letterhead_template_author')
def letterhead_template_duplicate(template_id):
    template=db.session.get(LetterheadTemplate,template_id) or abort(404)
    source=db.session.get(LetterheadTemplateVersion,template.current_published_version_id) if template.current_published_version_id else LetterheadTemplateVersion.query.filter_by(template_id=template.id).order_by(LetterheadTemplateVersion.version_no.desc()).first()
    if not source: abort(404)
    version_no=next_template_version_no([v.version_no for v in LetterheadTemplateVersion.query.filter_by(template_id=template.id).all()])
    version=LetterheadTemplateVersion(template_id=template.id,version_no=version_no,lifecycle_state='draft',layout_json=source.layout_json,scope_json=source.scope_json,content_hash=source.content_hash); db.session.add(version); template.status='draft'; record_audit('letterhead_template_duplicated','letterhead_template',template.id,module='letterhead',meta={'source_version':source.version_no,'version_no':version_no}); db.session.commit(); flash('New editable draft version created.','success'); return redirect(url_for('letterhead_template_editor_page',template_id=template.id,version_id=version.id))

@app.route('/settings', methods=['GET','POST'])
@login_required
def settings_page():
    user=current_user()
    if request.method=='POST':
        if not user or (user.role or '').lower()!='admin': abort(403)
        for k in _system_settings_server_keys():
            if k in request.form:
                vals=request.form.getlist(k); val=(vals[-1] if vals else '').strip()
                if k=='default_google_review_url' and val:
                    val=normalize_google_review_url(val)
                    if not val:
                        flash('Default Google Review Link is invalid.','danger'); return redirect(settings_pane_url('automations'))
                set_setting(k,val)
        flash('Settings saved.','success'); return redirect(settings_pane_url('automations'))
    return redirect(settings_pane_url(default_settings_pane(user)))

@app.route('/settings/<pane>')
@login_required
def system_settings_pane(pane):
    user=current_user(); panes=allowed_settings_panes(user); allowed={item['key'] for item in panes}
    if pane not in allowed: abort(403 if any(item['key']==pane for item in SYSTEM_SETTINGS_PANES) else 404)
    selected=next(item for item in panes if item['key']==pane)
    context=settings_pane_context(pane,user)
    return render_template('system_settings.html',settings_panes=panes,selected_settings_pane=pane,selected_settings=selected,**context)

@app.route('/admin')
@admin_required
def admin_panel():
    return redirect(settings_pane_url('users-groups'))

@app.route('/admin/kiosk/settings',methods=['POST'])
@admin_required
def kiosk_settings():
    u=current_user(); password=request.form.get('admin_password','')
    if not check_password_hash(u.password_hash,password):
        flash('Administrator password is required to change Windows/kiosk lock settings.','danger'); return redirect(url_for('admin_panel')+'#kiosk-security')
    enabled=request.form.get('kiosk_mode_enabled')=='1'; new_pin=request.form.get('kiosk_pin','').strip()
    if enabled and not (setting('kiosk_pin_hash','') or len(new_pin)>=6):
        flash('Set a kiosk PIN with at least 6 characters before enabling lock mode.','danger'); return redirect(url_for('admin_panel')+'#kiosk-security')
    if new_pin:
        if len(new_pin)<6:
            flash('Kiosk PIN must contain at least 6 characters.','danger'); return redirect(url_for('admin_panel')+'#kiosk-security')
        set_setting('kiosk_pin_hash',generate_password_hash(new_pin))
    set_setting('kiosk_mode_enabled','1' if enabled else '0')
    session['kiosk_unlocked']=not enabled
    flash(('Kiosk lock enabled. Unlock with the kiosk PIN or user password.' if enabled else 'Kiosk lock disabled.'),'success')
    return redirect(url_for('kiosk_lock') if enabled else url_for('admin_panel')+'#kiosk-security')

@app.route('/admin/kiosk/windows/<kind>')
@admin_required
def kiosk_windows_download(kind):
    names={'enable':'Enable-LivenzaKiosk.ps1','disable':'Disable-LivenzaKiosk.ps1','guide':'README_WINDOWS_KIOSK.md'}
    name=names.get(kind) or abort(404)
    return send_file(os.path.join(BASE_DIR,'windows-kiosk',name),as_attachment=True,download_name=name)

@app.route('/admin/users/save', methods=['POST'])
@admin_required
def admin_user_save():
    uid=request.form.get('id')
    u=db.session.get(User,int(uid)) if uid else None
    username=request.form.get('username','').strip()
    if not username:
        flash('Login ID is required.','danger'); return redirect(url_for('admin_panel'))
    if not u:
        if User.query.filter_by(username=username).first():
            flash('That Login ID already exists.','danger'); return redirect(url_for('admin_panel'))
        password=request.form.get('password','')
        if len(password)<8:
            flash('A new user password must contain at least 8 characters.','danger'); return redirect(url_for('admin_panel'))
        u=User(username=username,password_hash=generate_password_hash(password)); db.session.add(u)
    else:
        other=User.query.filter(User.username==username,User.id!=u.id).first()
        if other:
            flash('That Login ID already exists.','danger'); return redirect(url_for('admin_panel'))
        u.username=username
        if request.form.get('password'):
            u.password_hash=generate_password_hash(request.form['password'])
    u.full_name=request.form.get('full_name','').strip()
    if request.files.get('photo') and request.files['photo'].filename:
        raw,error=_upload_image_bytes(request.files['photo'])
        if not error:
            profile=profile_photo_data_uri_from_bytes(raw)
            avatar,mode,_=create_live_avatar(raw,prefer_ai=request.form.get('use_ai_avatar','1')=='1')
            if profile: u.photo_data_uri=profile
            u.avatar_data_uri=avatar or ''; u.avatar_generation_mode=mode or 'default'; u.avatar_updated_at=datetime.datetime.utcnow()
            record_audit('mascot_generated' if avatar else 'mascot_defaulted','user',u.id if u.id else None,module='identity',meta={'mode':u.avatar_generation_mode})
    last4=''.join(ch for ch in request.form.get('aadhaar_last4','') if ch.isdigit())[-4:]
    if last4: u.aadhaar_last4=last4
    if request.form.get('aadhaar_name') is not None: u.aadhaar_name=request.form.get('aadhaar_name','').strip()
    if request.form.get('aadhaar_verification_ref') is not None: u.aadhaar_verification_ref=request.form.get('aadhaar_verification_ref','').strip()
    u.role=request.form.get('role','manager') if request.form.get('role') in ('admin','manager') else 'manager'
    u.active=request.form.get('active')=='1'
    u.webauthn_enabled=request.form.get('webauthn_enabled')=='1'
    pattern=_pattern_value(request.form.get('pattern',''))
    if request.form.get('clear_pattern')=='1': u.pattern_hash=''
    elif pattern: u.pattern_hash=generate_password_hash('pattern:'+pattern)
    elif request.form.get('pattern'):
        flash('Pattern was not changed: connect at least 4 different dots.','warning')
    perms=[m for m in MODULES if request.form.get(f'perm_{m}')=='1']
    capabilities=[key for key in LETTERHEAD_CAPABILITIES if request.form.get(f'cap_{key}')=='1'] if u.role!='admin' else list(LETTERHEAD_CAPABILITIES)
    if 'letterhead' in perms and 'letterhead_use' not in capabilities: capabilities.append('letterhead_use')
    u.permissions_json=json.dumps(perms)
    u.capabilities_json=json.dumps(sorted(set(capabilities)))
    db.session.flush(); record_audit('user_permissions_changed','user',u.id,module='identity',meta={'role':u.role,'permissions':perms,'letterhead_capabilities':sorted(set(capabilities))}); db.session.commit(); flash('User access saved.','success'); return redirect(url_for('admin_panel'))

@app.route('/admin/users/<int:uid>/delete', methods=['POST'])
@admin_required
def admin_user_delete(uid):
    u=db.session.get(User,uid) or abort(404)
    if u.id==current_user().id:
        flash('You cannot delete the account currently signed in.','danger')
    else:
        db.session.delete(u); db.session.commit(); flash('User deleted.','success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/users/<int:uid>/avatar', methods=['POST'])
@admin_required
def admin_user_avatar(uid):
    u=db.session.get(User,uid) or abort(404)
    upload=request.files.get('photo') or request.files.get('avatar_photo')
    action=(request.form.get('avatar_action') or '').strip()
    if action=='reset':
        u.avatar_data_uri=''; u.avatar_generation_mode='default'; u.avatar_updated_at=datetime.datetime.utcnow(); record_audit('mascot_reset','user',u.id,module='identity'); db.session.commit()
        flash(f'Default Livenza mascot restored for {u.full_name or u.username}.','success')
        return redirect(url_for('admin_panel')+f'#user-{u.id}')
    if action=='remove_photo':
        u.photo_data_uri=''; u.avatar_data_uri=''; u.avatar_generation_mode='default'; u.avatar_updated_at=datetime.datetime.utcnow(); record_audit('mascot_photo_removed','user',u.id,module='identity'); db.session.commit()
        flash(f'Photo and mascot removed for {u.full_name or u.username}.','success')
        return redirect(url_for('admin_panel')+f'#user-{u.id}')
    if upload and upload.filename:
        raw,error=_upload_image_bytes(upload)
    elif action=='regenerate' and u.photo_data_uri:
        raw=_data_uri_bytes(u.photo_data_uri); error='' if raw else 'Upload the source photo again before regenerating the mascot.'
    else:
        raw,error=b'', 'Choose a source photo or use Regenerate from Saved Photo.'
    if error:
        flash(error,'danger')
        return redirect(url_for('admin_panel')+f'#user-{u.id}')
    profile=profile_photo_data_uri_from_bytes(raw)
    avatar,mode,message=create_live_avatar(raw,prefer_ai=request.form.get('use_ai_avatar','1')=='1')
    if profile:
        u.photo_data_uri=profile
    u.avatar_data_uri=avatar or ''; u.avatar_generation_mode=mode or 'default'; u.avatar_updated_at=datetime.datetime.utcnow()
    record_audit('mascot_generated' if avatar else 'mascot_defaulted','user',u.id,module='identity',meta={'mode':u.avatar_generation_mode})
    db.session.commit()
    flash(f'{u.full_name or u.username}: {message}',('success' if avatar else 'warning'))
    return redirect(url_for('admin_panel')+f'#user-{u.id}')

@app.route('/admin/cities/save', methods=['POST'])
@admin_required
def admin_city_save():
    cid=request.form.get('id'); c=db.session.get(City,int(cid)) if cid else None
    name=request.form.get('name','').strip()
    if not name:
        flash('City name is required.','danger'); return redirect(url_for('admin_panel'))
    if not c:
        c=City.query.filter(func.lower(City.name)==name.lower()).first()
        if not c: c=City(name=name); db.session.add(c)
    else: c.name=name
    c.code=request.form.get('code','').strip().upper(); c.active=request.form.get('active')=='1'
    db.session.commit(); flash('City saved.','success'); return redirect(url_for('admin_panel'))

@app.route('/admin/cities/<int:cid>/delete', methods=['POST'])
@admin_required
def admin_city_delete(cid):
    c=db.session.get(City,cid) or abort(404); db.session.delete(c); db.session.commit(); flash('City removed.','success'); return redirect(url_for('admin_panel'))


def ensure_v150_user_columns():
    """Small compatibility bridge; production schema is also provided as a migration."""
    existing={c['name'] for c in inspect(db.engine).get_columns('user')}
    statements=[]
    if 'pattern_hash' not in existing: statements.append('ALTER TABLE "user" ADD COLUMN pattern_hash TEXT DEFAULT \'\'')
    if 'webauthn_enabled' not in existing: statements.append('ALTER TABLE "user" ADD COLUMN webauthn_enabled BOOLEAN NOT NULL DEFAULT FALSE')
    if 'webauthn_enrolled_at' not in existing: statements.append('ALTER TABLE "user" ADD COLUMN webauthn_enrolled_at TIMESTAMP NULL')
    for sql in statements: db.session.execute(db.text(sql))
    if statements: db.session.commit()

def ensure_v1512_user_columns():
    """Compatibility bridge for personal live-avatar storage."""
    existing={c['name'] for c in inspect(db.engine).get_columns('user')}
    statements=[]
    if 'avatar_data_uri' not in existing: statements.append('ALTER TABLE "user" ADD COLUMN avatar_data_uri TEXT DEFAULT \'\'')
    if 'avatar_generation_mode' not in existing: statements.append('ALTER TABLE "user" ADD COLUMN avatar_generation_mode VARCHAR(40) DEFAULT \'\'')
    if 'avatar_updated_at' not in existing: statements.append('ALTER TABLE "user" ADD COLUMN avatar_updated_at TIMESTAMP NULL')
    for sql in statements: db.session.execute(db.text(sql))
    if statements: db.session.commit()

def ensure_v190_user_columns():
    """Compatibility bridge for Letterhead capability storage on existing databases."""
    existing={c['name'] for c in inspect(db.engine).get_columns('user')}
    if 'capabilities_json' not in existing:
        db.session.execute(db.text('ALTER TABLE "user" ADD COLUMN capabilities_json TEXT DEFAULT \'[]\'')); db.session.commit()

def ensure_integration_provider_seed():
    try:
        rows=load_integration_catalog(os.path.join(BASE_DIR,'data','integration_providers.json'))
        return seed_integration_providers(db.session,IntegrationProvider,rows)
    except Exception as exc:
        db.session.rollback()
        print('Integration provider seed warning:',exc)
        return {'created':0,'updated':0}

def ensure_letterhead_starter_templates():
    admin=User.query.filter(func.lower(User.role)=='admin').order_by(User.id).first() or User.query.order_by(User.id).first()
    if not admin: return {'created':0,'existing':0}
    created=existing=0; now=datetime.datetime.utcnow()
    for item in starter_template_definitions():
        row=LetterheadTemplate.query.filter_by(slug=item['slug']).first()
        if row: existing+=1; continue
        row=LetterheadTemplate(name=item['name'],slug=item['slug'],document_family_scope=item['document_family'],status='published',created_by_user_id=admin.id); db.session.add(row); db.session.flush()
        layout=json.dumps(item.get('layout') or {},ensure_ascii=False,sort_keys=True,separators=(',',':'))
        version=LetterheadTemplateVersion(template_id=row.id,version_no=1,lifecycle_state='published',layout_json=layout,scope_json='{}',content_hash=hashlib.sha256(layout.encode()).hexdigest(),published_by_user_id=None,published_at=now); db.session.add(version); db.session.flush(); row.current_published_version_id=version.id
        record_audit('letterhead_template_seeded','letterhead_template',row.id,module='letterhead',meta={'source':'system_seed','slug':row.slug,'version_no':1}); created+=1
    if created: db.session.commit()
    return {'created':created,'existing':existing}

def bootstrap():
    os.makedirs(os.path.join(BASE_DIR,'instance'),exist_ok=True)
    db.create_all()
    ensure_v150_user_columns()
    ensure_v1512_user_columns()
    ensure_v190_user_columns()
    ensure_electricity_provider_seed()
    ensure_integration_provider_seed()
    migrate_legacy_party_profiles()
    if User.query.count()==0:
        username=os.getenv('ADMIN_USERNAME','admin').strip() or 'admin'
        password=os.getenv('ADMIN_PASSWORD','')
        if not password:
            password='ChangeMeNow!2026'
            print('WARNING: ADMIN_PASSWORD was not set. Temporary password: ChangeMeNow!2026')
        db.session.add(User(username=username,password_hash=generate_password_hash(password),role='admin')); db.session.commit()
    ensure_letterhead_starter_templates()

with app.app_context(): bootstrap()

if __name__=='__main__':
    app.run(host='0.0.0.0',port=int(os.getenv('PORT','5000')),debug=os.getenv('FLASK_DEBUG')=='1')
